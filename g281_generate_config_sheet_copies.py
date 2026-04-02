from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter


OUTPUT_PATH = Path("g281_data_config_sheet_copies.json")
SOURCE_DIR = Path("BOM核对")
SHEET_NAME = "配置清单"
CONTRACT_VERSION = "g281.config-sheet-copies.v1"


@dataclass(frozen=True)
class WorkbookSpec:
    key: str
    label: str
    version_tag: str
    filename_hint: str


WORKBOOK_SPECS = (
    WorkbookSpec(
        key="quote",
        label="报价版配置清单",
        version_tag="quote",
        filename_hint="V01-11.3",
    ),
    WorkbookSpec(
        key="fixed",
        label="定点版配置清单",
        version_tag="fixed",
        filename_hint="V05-2026.01.04",
    ),
)


def discover_workbook(spec: WorkbookSpec) -> Path:
    candidates = []
    for path in SOURCE_DIR.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if spec.filename_hint in path.name:
            candidates.append(path)
    if not candidates:
        raise FileNotFoundError(
            f"Unable to locate workbook for {spec.key} with filename hint '{spec.filename_hint}' in {SOURCE_DIR}."
        )
    if len(candidates) > 1:
        candidates.sort(key=lambda item: item.name)
    return candidates[0]


def normalize_sheet_name(value: str) -> str:
    return "".join(str(value).split()).lower()


def discover_sheet_name(sheet_names: list[str], preferred: str) -> str:
    if preferred in sheet_names:
        return preferred

    normalized_target = normalize_sheet_name(preferred)
    for name in sheet_names:
        if normalize_sheet_name(name) == normalized_target:
            return name

    # Practical fallback: both BOM files place the config list at sheet index 0.
    return sheet_names[0]


def serialize_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def clean_descriptor_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("Values must be of type"):
        return None
    return value


def serialize_color(color: Color | None) -> dict[str, Any] | None:
    if color is None:
        return None
    payload = {
        "type": clean_descriptor_value(serialize_scalar(color.type)),
        "rgb": clean_descriptor_value(serialize_scalar(color.rgb)),
        "indexed": clean_descriptor_value(serialize_scalar(color.indexed)),
        "theme": clean_descriptor_value(serialize_scalar(color.theme)),
        "tint": clean_descriptor_value(serialize_scalar(color.tint)),
        "auto": clean_descriptor_value(serialize_scalar(color.auto)),
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_side(side: Any) -> dict[str, Any] | None:
    if side is None:
        return None
    payload = {
        "style": side.style,
        "color": serialize_color(side.color),
    }
    compact = {key: value for key, value in payload.items() if value not in (None, "")}
    return compact or None


def serialize_font(font: Any) -> dict[str, Any]:
    payload = {
        "name": font.name,
        "size": font.sz,
        "bold": font.b,
        "italic": font.i,
        "underline": font.u,
        "strike": font.strike,
        "color": serialize_color(font.color),
        "family": font.family,
        "charset": font.charset,
        "scheme": font.scheme,
        "vertAlign": font.vertAlign,
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_fill(fill: Any) -> dict[str, Any]:
    payload = {
        "fillType": fill.fill_type,
        "patternType": getattr(fill, "patternType", None),
        "fgColor": serialize_color(getattr(fill, "fgColor", None)),
        "bgColor": serialize_color(getattr(fill, "bgColor", None)),
        "type": getattr(fill, "type", None),
        "degree": getattr(fill, "degree", None),
        "top": serialize_color(getattr(fill, "top", None)),
        "bottom": serialize_color(getattr(fill, "bottom", None)),
        "left": serialize_color(getattr(fill, "left", None)),
        "right": serialize_color(getattr(fill, "right", None)),
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_border(border: Any) -> dict[str, Any]:
    payload = {
        "left": serialize_side(border.left),
        "right": serialize_side(border.right),
        "top": serialize_side(border.top),
        "bottom": serialize_side(border.bottom),
        "diagonal": serialize_side(border.diagonal),
        "vertical": serialize_side(border.vertical),
        "horizontal": serialize_side(border.horizontal),
        "diagonalDown": border.diagonalDown,
        "diagonalUp": border.diagonalUp,
        "outline": border.outline,
    }
    return {key: value for key, value in payload.items() if value not in (None, "")}


def serialize_alignment(alignment: Any) -> dict[str, Any]:
    payload = {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "textRotation": alignment.textRotation,
        "wrapText": alignment.wrapText,
        "shrinkToFit": alignment.shrinkToFit,
        "indent": alignment.indent,
        "relativeIndent": alignment.relativeIndent,
        "justifyLastLine": alignment.justifyLastLine,
        "readingOrder": alignment.readingOrder,
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_protection(protection: Any) -> dict[str, Any]:
    payload = {
        "locked": protection.locked,
        "hidden": protection.hidden,
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_style(cell: Any) -> dict[str, Any]:
    return {
        "numberFormat": cell.number_format,
        "font": serialize_font(cell.font),
        "fill": serialize_fill(cell.fill),
        "border": serialize_border(cell.border),
        "alignment": serialize_alignment(cell.alignment),
        "protection": serialize_protection(cell.protection),
    }


def serialize_row_dimensions(worksheet: Any) -> tuple[dict[str, Any], list[int]]:
    rows: dict[str, Any] = {}
    hidden_rows: list[int] = []
    for row_idx, dim in worksheet.row_dimensions.items():
        payload = {
            "height": dim.height,
            "hidden": bool(dim.hidden),
            "outlineLevel": dim.outlineLevel,
            "collapsed": bool(dim.collapsed),
            "styleId": dim.style_id,
            "customHeight": bool(dim.customHeight),
            "customFormat": bool(dim.customFormat),
        }
        compact = {key: value for key, value in payload.items() if value not in (None, False, 0)}
        if dim.hidden:
            hidden_rows.append(int(row_idx))
        if compact:
            rows[str(row_idx)] = compact
    return rows, sorted(set(hidden_rows))


def serialize_column_dimensions(worksheet: Any) -> tuple[dict[str, Any], list[str]]:
    columns: dict[str, Any] = {}
    hidden_columns: list[str] = []
    for key, dim in worksheet.column_dimensions.items():
        column_key = str(key)
        payload = {
            "min": dim.min,
            "max": dim.max,
            "width": dim.width,
            "hidden": bool(dim.hidden),
            "outlineLevel": dim.outlineLevel,
            "collapsed": bool(dim.collapsed),
            "styleId": dim.style_id,
            "bestFit": bool(dim.bestFit),
            "customWidth": bool(dim.customWidth),
        }
        compact = {attr: value for attr, value in payload.items() if value not in (None, False, 0)}
        if dim.hidden:
            if dim.min and dim.max:
                hidden_columns.extend(get_column_letter(index) for index in range(dim.min, dim.max + 1))
            else:
                hidden_columns.append(column_key)
        if compact:
            columns[column_key] = compact
    return columns, sorted(set(hidden_columns))


def serialize_page_setup(worksheet: Any) -> dict[str, Any]:
    setup = worksheet.page_setup
    payload = {
        "orientation": setup.orientation,
        "paperSize": setup.paperSize,
        "fitToWidth": setup.fitToWidth,
        "fitToHeight": setup.fitToHeight,
        "scale": setup.scale,
        "useFirstPageNumber": setup.useFirstPageNumber,
        "firstPageNumber": setup.firstPageNumber,
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_sheet_view(worksheet: Any) -> dict[str, Any]:
    view = worksheet.sheet_view
    payload = {
        "showGridLines": view.showGridLines,
        "showRowColHeaders": view.showRowColHeaders,
        "showZeros": view.showZeros,
        "tabSelected": view.tabSelected,
        "zoomScale": view.zoomScale,
        "topLeftCell": view.topLeftCell,
        "rightToLeft": view.rightToLeft,
        "view": view.view,
    }
    return {key: value for key, value in payload.items() if value is not None}


def serialize_data_validations(worksheet: Any) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    container = worksheet.data_validations
    if container is None:
        return output
    for rule in container.dataValidation:
        payload = {
            "sqref": str(rule.sqref),
            "type": rule.type,
            "operator": rule.operator,
            "allowBlank": rule.allowBlank,
            "showDropDown": rule.showDropDown,
            "showInputMessage": rule.showInputMessage,
            "showErrorMessage": rule.showErrorMessage,
            "formula1": rule.formula1,
            "formula2": rule.formula2,
            "errorTitle": rule.errorTitle,
            "error": rule.error,
            "promptTitle": rule.promptTitle,
            "prompt": rule.prompt,
        }
        output.append({key: value for key, value in payload.items() if value is not None})
    return output


def serialize_cells(worksheet: Any, display_worksheet: Any | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    style_pool: dict[int, dict[str, Any]] = {}
    cells: list[dict[str, Any]] = []
    ordered_cells = sorted(worksheet._cells.values(), key=lambda cell: (cell.row, cell.column))

    for cell in ordered_cells:
        style_id = int(getattr(cell, "style_id", 0) or 0)
        if style_id not in style_pool:
            style_pool[style_id] = serialize_style(cell)

        payload: dict[str, Any] = {
            "r": cell.row,
            "c": cell.column,
            "addr": cell.coordinate,
            "type": cell.data_type,
            "styleId": style_id,
        }

        if cell.data_type == "f":
            payload["formula"] = str(cell.value) if cell.value is not None else ""
            if display_worksheet is not None:
                payload["displayValue"] = serialize_scalar(display_worksheet[cell.coordinate].value)
        else:
            payload["value"] = serialize_scalar(cell.value)

        if cell.comment is not None:
            payload["comment"] = {
                "text": cell.comment.text,
                "author": cell.comment.author,
            }
        if cell.hyperlink is not None:
            payload["hyperlink"] = str(cell.hyperlink.target or cell.hyperlink.ref or "")

        cells.append(payload)

    style_pool_json = {str(style_id): style for style_id, style in style_pool.items()}
    return cells, style_pool_json


def snapshot_sheet(worksheet: Any, display_worksheet: Any | None = None) -> dict[str, Any]:
    row_dimensions, hidden_rows = serialize_row_dimensions(worksheet)
    column_dimensions, hidden_columns = serialize_column_dimensions(worksheet)
    cells, style_pool = serialize_cells(worksheet, display_worksheet)

    freeze = worksheet.freeze_panes
    freeze_panes = freeze.coordinate if hasattr(freeze, "coordinate") else (str(freeze) if freeze else None)

    return {
        "sheetName": worksheet.title,
        "dimension": worksheet.calculate_dimension(),
        "maxRow": worksheet.max_row,
        "maxColumn": worksheet.max_column,
        "freezePanes": freeze_panes,
        "mergedRanges": [str(cell_range) for cell_range in worksheet.merged_cells.ranges],
        "rowDimensions": row_dimensions,
        "columnDimensions": column_dimensions,
        "hiddenRows": hidden_rows,
        "hiddenColumns": hidden_columns,
        "defaultRowHeight": worksheet.sheet_format.defaultRowHeight,
        "defaultColWidth": worksheet.sheet_format.defaultColWidth,
        "sheetView": serialize_sheet_view(worksheet),
        "pageSetup": serialize_page_setup(worksheet),
        "autoFilter": str(worksheet.auto_filter.ref) if worksheet.auto_filter and worksheet.auto_filter.ref else None,
        "dataValidations": serialize_data_validations(worksheet),
        "cells": cells,
        "stylePool": style_pool,
        "metadata": {
            "sheetState": worksheet.sheet_state,
            "tabColor": serialize_color(worksheet.sheet_properties.tabColor),
            "showGridLines": worksheet.sheet_view.showGridLines,
            "printTitleRows": worksheet.print_title_rows,
            "printTitleCols": worksheet.print_title_cols,
        },
    }


def build_payload() -> dict[str, Any]:
    versions: dict[str, Any] = {}
    workbook_sources: dict[str, str] = {}

    for spec in WORKBOOK_SPECS:
        workbook_path = discover_workbook(spec)
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
        display_workbook = load_workbook(workbook_path, data_only=True, read_only=False)
        sheet_name = discover_sheet_name(workbook.sheetnames, SHEET_NAME)
        worksheet = workbook[sheet_name]
        display_worksheet = display_workbook[sheet_name]

        versions[spec.key] = {
            "key": spec.key,
            "label": spec.label,
            "versionTag": spec.version_tag,
            "workbook": workbook_path.name,
            "workbookPath": str(workbook_path.as_posix()),
            "sheetName": sheet_name,
            "sheetIndex": workbook.sheetnames.index(sheet_name),
            "sheetCount": len(workbook.sheetnames),
            "snapshot": snapshot_sheet(worksheet, display_worksheet),
        }
        workbook_sources[spec.key] = workbook_path.name
        workbook.close()
        display_workbook.close()

    return {
        "meta": {
            "generator": "g281_generate_config_sheet_copies.py",
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "contractVersion": CONTRACT_VERSION,
            "sourceDirectory": str(SOURCE_DIR.as_posix()),
            "sheetName": SHEET_NAME,
            "versionOrder": ["tt", "fixed", "quote"],
            "availableVersions": list(versions.keys()),
            "workbooks": workbook_sources,
            "notes": [
                "This payload preserves sheet content and formatting metadata for config-sheet version management.",
                "Workbook formulas are exported as formula strings with style references in stylePool.",
            ],
        },
        "versions": versions,
    }


def main() -> None:
    payload = build_payload()
    OUTPUT_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[ok] wrote {OUTPUT_PATH}")
    print(f"[ok] versions: {', '.join(payload.get('versions', {}).keys())}")


if __name__ == "__main__":
    main()
