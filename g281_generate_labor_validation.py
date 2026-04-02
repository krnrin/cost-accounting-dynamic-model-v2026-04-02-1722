from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_INDEX = {
    "summary": 0,
    "assessment": 1,
    "quote_logic": 3,
    "labor_base": 4,
    "efficiency": 5,
}

SUMMARY_COLUMN = {
    "quote": "E",
    "fixed": "C",
}

NUMBER_RE = re.compile(r"(-?\d+(?:\.\d+)?)")


def discover_workbook(pattern: str) -> Path:
    matches = [
        Path(name)
        for name in os.listdir(".")
        if name.endswith(".xlsx") and not name.startswith("~$") and pattern in name and "导线联动" not in name
    ]
    if not matches:
        raise FileNotFoundError(f"Workbook containing '{pattern}' not found in current directory.")
    return sorted(matches)[0]


def collapse_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\s+", " ", text)


def raw_error_text(value: Any) -> str:
    text = collapse_text(value)
    if text.startswith("#"):
        return text
    return ""


def numeric_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = collapse_text(value).replace(",", "")
    if not text or text.startswith("#"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def round_number(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def parse_rate_from_header(value: Any) -> float | None:
    match = NUMBER_RE.search(collapse_text(value))
    if not match:
        return None
    return round_number(float(match.group(1)), 6)


def note_join(*parts: str) -> str:
    return "；".join(part for part in parts if collapse_text(part))


def make_item(
    item_key: str,
    label: str,
    unit: str,
    source_sheet: str,
    source_cell: str,
    raw_value: Any,
    formula_value: Any = "",
    *,
    numeric_override: float | None = None,
    note: str = "",
) -> dict[str, Any]:
    numeric = round_number(numeric_override if numeric_override is not None else numeric_value(raw_value))
    raw_text = collapse_text(raw_value)
    formula_text = collapse_text(formula_value)
    value: Any = numeric if numeric is not None else (raw_text or "-")
    value_type = "number" if numeric is not None else ("error" if raw_text.startswith("#") else "text")
    return {
        "itemKey": item_key,
        "label": label,
        "unit": unit,
        "value": value,
        "numericValue": numeric,
        "valueType": value_type,
        "sourceSheet": source_sheet,
        "sourceCell": source_cell,
        "source": f"{source_sheet}!{source_cell}",
        "formula": formula_text,
        "note": collapse_text(note),
        "rawValue": raw_text,
    }


def find_row(ws, needles: list[str], columns: tuple[str, ...] = ("B", "A", "C"), limit: int = 80) -> int:
    for row in range(1, min(ws.max_row, limit) + 1):
        haystack = " | ".join(collapse_text(ws[f"{column}{row}"].value) for column in columns)
        if all(needle in haystack for needle in needles):
            return row
    raise KeyError(f"Unable to find row matching: {' / '.join(needles)}")


def workbook_context(path: Path) -> dict[str, Any]:
    value_book = load_workbook(path, read_only=True, data_only=True)
    formula_book = load_workbook(path, read_only=True, data_only=False)
    sheets = {}
    for alias, index in SHEET_INDEX.items():
        title = value_book.sheetnames[index]
        sheets[alias] = {
            "title": title,
            "values": value_book[title],
            "formulas": formula_book[formula_book.sheetnames[index]],
        }
    return {
        "path": path,
        "sheets": sheets,
    }


def summary_metric(ctx: dict[str, Any], kind: str, item_key: str, label_needles: list[str], unit: str, note: str = "") -> dict[str, Any]:
    ws_values = ctx["sheets"]["summary"]["values"]
    ws_formulas = ctx["sheets"]["summary"]["formulas"]
    row = find_row(ws_values, label_needles)
    column = SUMMARY_COLUMN[kind]
    return make_item(
        item_key,
        collapse_text(ws_values[f"B{row}"].value) or collapse_text(ws_values[f"A{row}"].value),
        unit,
        ctx["sheets"]["summary"]["title"],
        f"{column}{row}",
        ws_values[f"{column}{row}"].value,
        ws_formulas[f"{column}{row}"].value,
        note=note,
    )


def assessment_metric(
    ctx: dict[str, Any],
    row: int,
    item_key: str,
    unit: str,
    *,
    note: str = "",
    numeric_override: float | None = None,
) -> dict[str, Any]:
    ws_values = ctx["sheets"]["assessment"]["values"]
    ws_formulas = ctx["sheets"]["assessment"]["formulas"]
    label = collapse_text(ws_values[f"B{row}"].value) or collapse_text(ws_values[f"A{row}"].value)
    description = collapse_text(ws_values[f"C{row}"].value) or collapse_text(ws_values[f"D{row}"].value)
    return make_item(
        item_key,
        label,
        unit,
        ctx["sheets"]["assessment"]["title"],
        f"F{row}",
        ws_values[f"F{row}"].value,
        ws_formulas[f"F{row}"].value,
        numeric_override=numeric_override,
        note=note_join(note, description),
    )


def labor_base_metric(ctx: dict[str, Any], cell: str, item_key: str, label: str, unit: str, note: str = "") -> dict[str, Any]:
    ws_values = ctx["sheets"]["labor_base"]["values"]
    ws_formulas = ctx["sheets"]["labor_base"]["formulas"]
    return make_item(
        item_key,
        label,
        unit,
        ctx["sheets"]["labor_base"]["title"],
        cell,
        ws_values[cell].value,
        ws_formulas[cell].value,
        note=note,
    )


def efficiency_metric(ctx: dict[str, Any], cell: str, item_key: str, label: str, unit: str, note: str = "") -> dict[str, Any]:
    ws_values = ctx["sheets"]["efficiency"]["values"]
    ws_formulas = ctx["sheets"]["efficiency"]["formulas"]
    return make_item(
        item_key,
        label,
        unit,
        ctx["sheets"]["efficiency"]["title"],
        cell,
        ws_values[cell].value,
        ws_formulas[cell].value,
        note=note,
    )


def quote_logic_rates(ctx: dict[str, Any]) -> dict[str, dict[str, Any]]:
    ws_values = ctx["sheets"]["quote_logic"]["values"]
    ws_formulas = ctx["sheets"]["quote_logic"]["formulas"]
    direct_rate = parse_rate_from_header(ws_values["J2"].value)
    mfg_rate = round_number(numeric_value(ws_values["K20"].value))
    return {
        "directRate": make_item(
            "apply.directRate",
            "前工程费率",
            "元/h",
            ctx["sheets"]["quote_logic"]["title"],
            "J2",
            ws_values["J2"].value,
            ws_formulas["J2"].value,
            numeric_override=direct_rate,
            note="按客户报价逻辑直接人工列头取值。",
        ),
        "manufacturingRate": make_item(
            "apply.manufacturingRate",
            "制造费率",
            "元/h",
            ctx["sheets"]["quote_logic"]["title"],
            "K20",
            ws_values["K20"].value,
            ws_formulas["K20"].value,
            numeric_override=mfg_rate,
            note="按客户报价逻辑制造费汇总取值。",
        ),
    }


def manufacturing_rows(ctx: dict[str, Any], fallback_ctx: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    ws_values = ctx["sheets"]["assessment"]["values"]
    items = [
        assessment_metric(ctx, 24, "manufacturing.indirect", "元/PCS"),
        assessment_metric(ctx, 25, "manufacturing.supplies", "元/PCS"),
        assessment_metric(ctx, 26, "manufacturing.materials", "元/PCS"),
        assessment_metric(ctx, 27, "manufacturing.building", "元/PCS"),
        assessment_metric(ctx, 28, "manufacturing.warehouse", "元/PCS"),
        assessment_metric(ctx, 29, "manufacturing.other", "元/PCS"),
    ]
    raw_material_loss = ws_values["F30"].value
    material_loss_note = ""
    material_loss_value = numeric_value(raw_material_loss)
    if material_loss_value is None and raw_error_text(raw_material_loss):
        base_material_cost = numeric_value(ws_values["F15"].value)
        if base_material_cost is not None:
            material_loss_value = base_material_cost * 0.005
            material_loss_note = f"Excel 原值 {raw_error_text(raw_material_loss)}，按单套材料成本 * 0.5% 修复。"
        elif fallback_ctx is not None:
            fallback_ws = fallback_ctx["sheets"]["assessment"]["values"]
            fallback_loss = numeric_value(fallback_ws["F30"].value)
            if fallback_loss is not None:
                material_loss_value = fallback_loss
                material_loss_note = f"Excel 原值 {raw_error_text(raw_material_loss)}，按对侧核算表同逻辑行 F30 回填。"
    items.append(
        assessment_metric(
            ctx,
            30,
            "manufacturing.material_loss",
            "元/PCS",
            note=material_loss_note,
            numeric_override=material_loss_value,
        )
    )
    total_raw = ws_values["F23"].value
    total_value = numeric_value(total_raw)
    total_note = ""
    if total_value is None:
        subtotal = 0.0
        for item in items:
            if item["numericValue"] is None:
                subtotal = None
                break
            subtotal += float(item["numericValue"])
        if subtotal is not None:
            total_value = subtotal
            total_note = f"Excel 原值 {raw_error_text(total_raw) or '-'}，按 F24:F30 汇总修复。"
    items.insert(
        0,
        assessment_metric(
            ctx,
            23,
            "manufacturing.total",
            "元/PCS",
            note=total_note,
            numeric_override=total_value,
        ),
    )
    return items


def build_snapshot(ctx: dict[str, Any], kind: str, fallback_ctx: dict[str, Any] | None = None) -> dict[str, Any]:
    rates = quote_logic_rates(ctx)
    direct_per_set = assessment_metric(ctx, 16, "snapshot.directPerSet", "元/PCS")
    manufacturing_total = manufacturing_rows(ctx, fallback_ctx)[0]
    direct_rate = rates["directRate"]["numericValue"]
    manufacturing_rate = rates["manufacturingRate"]["numericValue"]
    direct_hours = round_number(
        float(direct_per_set["numericValue"]) / float(direct_rate)
        if direct_per_set["numericValue"] is not None and direct_rate not in (None, 0)
        else None
    )
    manufacturing_hours = round_number(
        float(manufacturing_total["numericValue"]) / float(manufacturing_rate)
        if manufacturing_total["numericValue"] is not None and manufacturing_rate not in (None, 0)
        else None
    )
    return {
        "kind": kind,
        "label": "报价工时版" if kind == "quote" else "定点工时版",
        "directHours": direct_hours,
        "directRate": direct_rate,
        "manufacturingHours": manufacturing_hours,
        "manufacturingRate": manufacturing_rate,
        "directLaborPerSet": direct_per_set["numericValue"],
        "manufacturingPerSet": manufacturing_total["numericValue"],
        "sources": {
            "directHours": f"{direct_per_set['source']} / {rates['directRate']['source']}",
            "directRate": rates["directRate"]["source"],
            "manufacturingHours": f"{manufacturing_total['source']} / {rates['manufacturingRate']['source']}",
            "manufacturingRate": rates["manufacturingRate"]["source"],
        },
        "note": "程序应用值按《客户报价逻辑》的前工程费率与制造费率折算。",
    }


def align_items(quote_items: list[dict[str, Any]], fixed_items: list[dict[str, Any]]) -> dict[str, Any]:
    quote_map = {item["itemKey"]: item for item in quote_items}
    fixed_map = {item["itemKey"]: item for item in fixed_items}
    order = [item["itemKey"] for item in quote_items]
    for item in fixed_items:
        if item["itemKey"] not in quote_map:
            order.append(item["itemKey"])

    aligned = []
    matched = 0
    quote_only = 0
    fixed_only = 0
    difference_count = 0

    for item_key in order:
        quote = quote_map.get(item_key)
        fixed = fixed_map.get(item_key)
        if quote and fixed:
            status = "matched"
            matched += 1
        elif quote:
            status = "quote_only"
            quote_only += 1
        else:
            status = "fixed_only"
            fixed_only += 1

        if quote and fixed:
            quote_numeric = quote.get("numericValue")
            fixed_numeric = fixed.get("numericValue")
            if quote_numeric is not None and fixed_numeric is not None:
                if abs(float(quote_numeric) - float(fixed_numeric)) > 1e-6:
                    difference_count += 1
            elif quote.get("value") != fixed.get("value"):
                difference_count += 1
        else:
            difference_count += 1

        aligned.append(
            {
                "status": status,
                "quote": quote,
                "fixed": fixed,
            }
        )

    return {
        "aligned": aligned,
        "summary": {
            "quoteCount": len(quote_items),
            "fixedCount": len(fixed_items),
            "matchedCount": matched,
            "quoteOnlyCount": quote_only,
            "fixedOnlyCount": fixed_only,
            "differenceCount": difference_count,
        },
    }


def build_scope(scope_id: str, scope_label: str, hint: str, quote_groups: list[dict[str, Any]], fixed_groups: list[dict[str, Any]]) -> dict[str, Any]:
    fixed_map = {group["key"]: group for group in fixed_groups}
    order = [group["key"] for group in quote_groups]
    for group in fixed_groups:
        if group["key"] not in order:
            order.append(group["key"])

    groups = []
    scope_summary = {
        "quoteCount": 0,
        "fixedCount": 0,
        "matchedCount": 0,
        "quoteOnlyCount": 0,
        "fixedOnlyCount": 0,
        "differenceCount": 0,
    }

    for key in order:
        quote_group = next((group for group in quote_groups if group["key"] == key), None)
        fixed_group = fixed_map.get(key)
        quote_items = quote_group["items"] if quote_group else []
        fixed_items = fixed_group["items"] if fixed_group else []
        aligned = align_items(quote_items, fixed_items)
        groups.append(
            {
                "key": key,
                "label": quote_group["label"] if quote_group else fixed_group["label"],
                "meta": quote_group.get("meta") if quote_group else fixed_group.get("meta", ""),
                "aligned": aligned["aligned"],
                "summary": aligned["summary"],
            }
        )
        for summary_key in scope_summary:
            scope_summary[summary_key] += aligned["summary"][summary_key]

    return {
        "scopeId": scope_id,
        "scopeLabel": scope_label,
        "hint": hint,
        "groups": groups,
        "summary": scope_summary,
    }


def workbook_groups(ctx: dict[str, Any], kind: str, fallback_ctx: dict[str, Any] | None = None) -> dict[str, list[dict[str, Any]]]:
    rates = quote_logic_rates(ctx)
    snapshot = build_snapshot(ctx, kind, fallback_ctx)
    direct_items = [
        assessment_metric(ctx, 16, "direct.total", "元/PCS", note="程序工时换算基准。"),
        assessment_metric(ctx, 17, "direct.open_line", "元/PCS"),
        assessment_metric(ctx, 18, "direct.common", "元/PCS"),
        assessment_metric(ctx, 19, "direct.assembly", "元/PCS"),
    ]
    manufacturing_items = manufacturing_rows(ctx, fallback_ctx)

    summary_items = [
        make_item(
            "snapshot.directHours",
            "前工程工时",
            "h/套",
            ctx["sheets"]["assessment"]["title"],
            "F16/J2",
            snapshot["directHours"],
            "",
            note="按单套直接人工 / 客户报价逻辑前工程费率换算。",
        ),
        rates["directRate"],
        make_item(
            "snapshot.manufacturingHours",
            "后工程工时",
            "h/套",
            ctx["sheets"]["assessment"]["title"],
            "F23/K20",
            snapshot["manufacturingHours"],
            "",
            note="按单套制造费用 / 客户报价逻辑制造费率换算。",
        ),
        rates["manufacturingRate"],
        summary_metric(
            ctx,
            kind,
            "summary.cycleSeconds",
            ["回路工时"],
            "s/回路",
            note="报价版取第一轮报价列，定点版取客户目标/定点列。",
        ),
        summary_metric(
            ctx,
            kind,
            "summary.operatingLaborRate",
            ["运营成本工时费"],
            "元/h",
            note="报价版取第一轮报价列，定点版取客户目标/定点列。",
        ),
        summary_metric(
            ctx,
            kind,
            "summary.directLaborRate",
            ["运营成本直接人工费率"],
            "元/h",
            note="报价版取第一轮报价列，定点版取客户目标/定点列。",
        ),
        assessment_metric(ctx, 23, "summary.manufacturingPerSet", "元/PCS", numeric_override=manufacturing_items[0]["numericValue"], note=manufacturing_items[0]["note"]),
    ]

    ws_base_values = ctx["sheets"]["labor_base"]["values"]
    ws_eff_values = ctx["sheets"]["efficiency"]["values"]
    base_rate = numeric_value(ws_base_values["E8"].value)
    efficiency = numeric_value(ws_eff_values["D17"].value)
    derived_rate = round_number(base_rate / efficiency if base_rate is not None and efficiency not in (None, 0) else None)
    baseline_items = [
        labor_base_metric(ctx, "E8", "baseline.stdAssembly", "K3 工厂标准工时费", "元/h", "项目评估汇总 F17:F19 实际引用。"),
        efficiency_metric(ctx, "D17", "baseline.efficiency", "K3 工厂效率", "系数", "项目评估汇总按工厂效率折算。"),
        make_item(
            "baseline.derivedDirectRate",
            "折算直接人工费率",
            "元/h",
            ctx["sheets"]["labor_base"]["title"],
            "E8/D17",
            derived_rate,
            "",
            note="按标准工时费 / 工厂效率折算。",
        ),
        labor_base_metric(ctx, "E10", "baseline.indirect", "间接人工费率", "元/h"),
        labor_base_metric(ctx, "E11", "baseline.supplies", "低值易耗品费率", "元/h"),
        labor_base_metric(ctx, "E12", "baseline.materials", "机物料消耗费率", "元/h"),
        labor_base_metric(ctx, "E13", "baseline.building", "厂房分摊费率", "元/h"),
        labor_base_metric(ctx, "E14", "baseline.warehouse", "自动化仓/仓库分摊费率", "元/h"),
        labor_base_metric(ctx, "E15", "baseline.other", "其他制费费率", "元/h"),
    ]

    return {
        "scenario": [
            {
                "key": "scenario.mapping",
                "label": "程序工时映射",
                "meta": "把核算表里的直接人工/制造费用折算为程序当前的 4 个工时参数。",
                "items": summary_items,
            }
        ],
        "directLabor": [
            {
                "key": "direct.cost",
                "label": "直接人工拆分",
                "meta": "来源：项目评估汇总（昆山90%）",
                "items": direct_items,
            }
        ],
        "manufacturing": [
            {
                "key": "manufacturing.cost",
                "label": "制造费用拆分",
                "meta": "来源：项目评估汇总（昆山90%）",
                "items": manufacturing_items,
            }
        ],
        "baseline": [
            {
                "key": "baseline.rates",
                "label": "工时费基准引用",
                "meta": "来源：运营工时费报价基准 + 工厂效率",
                "items": baseline_items,
            }
        ],
    }


def build_payload(quote_workbook: Path, fixed_workbook: Path) -> dict[str, Any]:
    quote_ctx = workbook_context(quote_workbook)
    fixed_ctx = workbook_context(fixed_workbook)
    quote_groups = workbook_groups(quote_ctx, "quote", fixed_ctx)
    fixed_groups = workbook_groups(fixed_ctx, "fixed", quote_ctx)

    scope_order = ["scenario", "directLabor", "manufacturing", "baseline"]
    scope_meta = {
        "scenario": {
            "label": "程序映射",
            "hint": "应用报价版 / 定点版后，会直接回填主页面的前工程工时、前工程费率、后工程工时、制造费率 4 个输入项。",
        },
        "directLabor": {
            "label": "直接人工",
            "hint": "按项目评估汇总的单套直接人工及前后工程拆分，对齐报价核算与定点核算。",
        },
        "manufacturing": {
            "label": "制造费用",
            "hint": "按项目评估汇总的制造费用拆分对齐展示；若 Excel 原值为 #VALUE!，程序按同页逻辑做可追溯修复。",
        },
        "baseline": {
            "label": "工时费基准",
            "hint": "保留运营工时费报价基准和工厂效率的原始引用，便于追溯工时费率来源。",
        },
    }

    comparisons = {
        scope_id: build_scope(
            scope_id,
            scope_meta[scope_id]["label"],
            scope_meta[scope_id]["hint"],
            quote_groups[scope_id],
            fixed_groups[scope_id],
        )
        for scope_id in scope_order
    }

    return {
        "meta": {
            "quoteWorkbook": quote_workbook.name,
            "fixedWorkbook": fixed_workbook.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "layer": "JSON 数据层",
        },
        "scopeOrder": scope_order,
        "versionSnapshots": {
            "quote": build_snapshot(quote_ctx, "quote", fixed_ctx),
            "fixed": build_snapshot(fixed_ctx, "fixed", quote_ctx),
        },
        "comparisons": comparisons,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate labor validation data from quote/fixed E281 workbooks.")
    parser.add_argument("--quote", type=Path, default=None, help="Quote workbook path")
    parser.add_argument("--fixed", type=Path, default=None, help="Fixed workbook path")
    parser.add_argument("--out", type=Path, default=Path("g281_data_labor_validation.json"), help="Output JSON path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    quote_path = args.quote or discover_workbook("报价核算")
    fixed_path = args.fixed or discover_workbook("定点核算")
    payload = build_payload(quote_path, fixed_path)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(args.out)


if __name__ == "__main__":
    main()
