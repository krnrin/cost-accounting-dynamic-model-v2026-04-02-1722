from __future__ import annotations

import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import BaseExtractor
from .utils import (
    collapse_text,
    discover_workbook,
    numeric_value,
    round_number,
)

ASSESSMENT_SHEET = "项目评估汇总（昆山90%）"
YEAR_COLUMNS = ["F", "G", "H", "I", "J", "K"]
YEAR_VALUES = [2026, 2027, 2028, 2029, 2030, 2031]
SNAPSHOT_SHEETS = [
    ASSESSMENT_SHEET, "运营工时费报价基准", "设备投资明细", "项目专用模具", "项目工装投入 ", "研发费用 ", "包装物流费用", "配置明细",
]

VERSION_SPECS = {
    "quote": {"pattern": "报价核算", "label": "报价版"},
    "fixed": {"pattern": "定点核算", "label": "定点版"},
}

TOTAL_CELLS = {
    "volume": "E5", "revenue": "E9", "profit": "E10", "margin": "E11", "cost": "E14", "material": "E15",
    "directLabor": "E16", "equipment": "E20", "manufacturing": "E23", "rnd": "E31", "packaging": "E32",
}

ANNUAL_ROWS = {
    "asp": 6, "revenue": 9, "cost": 14, "material": 15, "directLabor": 16, "equipment": 20, "manufacturing": 23, "rnd": 31, "packaging": 32,
}

class FinancialVersionsExtractor(BaseExtractor):
    def get_config(self) -> dict[str, Any]:
        return {
            "name": "financial_versions",
            "description": "Financial version tracking (quote vs fixed)",
            "workbook_pattern": "核算",
            "default_output": "g281_data_financial_versions.json",
        }

    def extract(self, input_path: Path) -> dict[str, Any]:
        directory = input_path if input_path.is_dir() else input_path.parent
        versions = {}
        for version_key, spec in VERSION_SPECS.items():
            workbook_path = discover_workbook(spec["pattern"], directory)
            versions[version_key] = self.extract_version_payload(version_key, workbook_path)

        generated_at = datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")
        return {
            "meta": {
                "generator": "FinancialVersionsExtractor",
                "generatedAt": generated_at,
                "sheetName": ASSESSMENT_SHEET,
                "note": "报价/定点精确口径来自《项目评估汇总（昆山90%）》，用于程序基线验证与版本化 ASP/成本参考。",
            },
            "versionOrder": list(VERSION_SPECS.keys()),
            "versions": versions,
        }

    def extract_version_payload(self, version_key: str, workbook_path: Path) -> dict[str, Any]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        worksheet = workbook[ASSESSMENT_SHEET]
        workbook_formula = load_workbook(workbook_path, data_only=False, read_only=True)

        def read_row(ws, row_index: int) -> list[float | None]:
            return [round_number(numeric_value(ws[f"{col}{row_index}"].value)) for col in YEAR_COLUMNS]

        totals = {key: round_number(numeric_value(worksheet[cell].value)) or 0.0 for key, cell in TOTAL_CELLS.items()}
        annual_raw = {key: [0.0 if v is None else float(v) for v in read_row(worksheet, row_idx)] for key, row_idx in ANNUAL_ROWS.items()}
        annual_volume = [0.0 if v is None else float(v) for v in read_row(worksheet, 5)]

        total_volume = totals["volume"] or sum(annual_volume)
        def per_set_calc(total_val: float) -> float:
            return round_number(total_val / total_volume) or 0.0 if total_volume else 0.0

        per_set = {
            "revenue": per_set_calc(totals["revenue"]), "cost": per_set_calc(totals["cost"]),
            "profit": per_set_calc(totals["profit"]), "margin": round_number(totals["margin"]) or 0.0,
            "material": per_set_calc(totals["material"]), "directLabor": per_set_calc(totals["directLabor"]),
            "equipment": per_set_calc(totals["equipment"]), "manufacturing": per_set_calc(totals["manufacturing"]),
            "rnd": per_set_calc(totals["rnd"]), "packaging": per_set_calc(totals["packaging"]),
        }

        annual_cost = [round_number((annual_raw["cost"][i] or 0.0) * (annual_volume[i] or 0.0)) or 0.0 for i in range(len(YEAR_VALUES))]
        annual_profit = [round_number((annual_raw["revenue"][i] or 0.0) - annual_cost[i]) or 0.0 for i in range(len(YEAR_VALUES))]
        annual_margin = [round_number(annual_profit[i] / annual_raw["revenue"][i]) if annual_raw["revenue"][i] else 0.0 for i in range(len(YEAR_VALUES))]

        seed = self.capture_assessment_workbook_seed(workbook, workbook_formula, workbook_path)
        return {
            "key": version_key, "label": VERSION_SPECS[version_key]["label"], "workbook": workbook_path.name,
            "sheetName": ASSESSMENT_SHEET, "years": YEAR_VALUES, "volumes": annual_volume, "asp": annual_raw["asp"],
            "totals": totals, "perSet": per_set,
            "annual": {
                "revenue": annual_raw["revenue"], "cost": annual_cost, "profit": annual_profit, "margin": annual_margin,
                "costPerSet": annual_raw["cost"], "materialPerSet": annual_raw["material"],
                "directLaborPerSet": annual_raw["directLabor"], "equipmentPerSet": annual_raw["equipment"],
                "manufacturingPerSet": annual_raw["manufacturing"], "rndPerSet": annual_raw["rnd"],
                "packagingPerSet": annual_raw["packaging"],
            },
            "cells": {"totals": TOTAL_CELLS, "annualRows": ANNUAL_ROWS},
            "assessmentWorkbookSeed": seed,
        }

    def capture_assessment_workbook_seed(self, data_workbook: Any, formula_workbook: Any, workbook_path: Path) -> dict[str, Any]:
        sheet_order = [s for s in SNAPSHOT_SHEETS if s in data_workbook.sheetnames and s in formula_workbook.sheetnames]
        return {
            "workbookName": workbook_path.name, "sourceFileName": workbook_path.name,
            "sourcePath": str(workbook_path.resolve()), "sheetOrder": sheet_order,
            "sheets": [self.capture_sheet_snapshot(data_workbook[s], formula_workbook[s]) for s in sheet_order],
        }

    def capture_sheet_snapshot(self, data_ws: Any, formula_ws: Any) -> dict[str, Any]:
        cells = []
        for r in range(1, formula_ws.max_row + 1):
            for c in range(1, formula_ws.max_column + 1):
                d_cell = data_ws.cell(row=r, column=c); f_cell = formula_ws.cell(row=r, column=c)
                val = d_cell.value; f_val = getattr(f_cell.value, "text", str(f_cell.value)) if f_cell.value is not None else None
                if val is None and f_val is None: continue
                cells.append({"address": f_cell.coordinate, "row": r, "column": c, "dataType": f_cell.data_type, "value": val, "formula": f_val})
        return {"sheetName": formula_ws.title, "maxRow": formula_ws.max_row, "maxColumn": formula_ws.max_column, "cells": cells}
