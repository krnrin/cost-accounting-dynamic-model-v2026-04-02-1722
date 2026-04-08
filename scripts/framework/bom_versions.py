from __future__ import annotations

import json
import math
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import BaseExtractor
from .utils import (
    to_text,
    numeric_value,
    discover_workbook,
)

DETAIL_SHEET = "二次物料明细"
MASTER_PATH = Path("g281_data_master.json")
TZ = timezone(timedelta(hours=8))

WIRE_KEYWORDS = ("导线", "线缆", "电缆")
TAPE_KEYWORDS = ("胶带",)

class BomVersionsExtractor(BaseExtractor):
    def get_config(self) -> dict[str, Any]:
        return {
            "name": "bom_versions",
            "description": "BOM version tracking and comparison",
            "workbook_pattern": "BOM",
            "default_output": "g281_data_bom_versions.json",
        }

    def extract(self, input_path: Path) -> dict[str, Any]:
        directory = input_path if input_path.is_dir() else input_path.parent
        
        base_defaults = self.load_master_defaults()
        quote_bom = discover_workbook("报价BOM", directory)
        fixed_bom = discover_workbook("定点BOM", directory)
        tt_reference = discover_workbook("G281 TT.xlsx", directory)
        tt_current = discover_workbook("G281 TT_", directory)

        quote_metrics = self.load_detail_metrics(quote_bom)
        fixed_metrics = self.load_detail_metrics(fixed_bom)
        tt_metrics = self.load_detail_metrics(tt_current)
        tt_changes = self.compare_tt_harness_lengths(tt_reference, tt_current)

        payload = {
            "meta": {
                "generator": "BomVersionsExtractor",
                "generatedAt": datetime.now(TZ).isoformat(timespec="seconds"),
                "quoteWorkbook": quote_bom.name,
                "fixedWorkbook": fixed_bom.name,
                "ttWorkbook": tt_current.name,
                "ttReferenceWorkbook": tt_reference.name,
                "layer": "JSON 数据层",
                "note": "BOM 版本快照来自报价 BOM、定点 BOM 和 TT 实际开线长度回填工作簿；材料系数按《二次物料明细》M 类总用量口径估算。",
            },
            "versionSnapshots": {
                "quote": self.build_snapshot("quote", "报价版BOM", quote_bom, base_defaults, quote_metrics, quote_metrics),
                "fixed": self.build_snapshot("fixed", "定点版BOM", fixed_bom, base_defaults, fixed_metrics, quote_metrics),
                "tt": self.build_snapshot("tt", "TT版BOM", tt_current, base_defaults, tt_metrics, quote_metrics),
            },
            "ttActualLengthChanges": tt_changes,
        }
        payload["versionSnapshots"]["tt"]["actualLengthChangeSummary"] = {
            "changedHarnessCount": tt_changes["changedHarnessCount"],
            "changedRowCount": tt_changes["changedRowCount"],
            "deltaQtyTotal": tt_changes["deltaQtyTotal"],
        }
        return payload

    def load_master_defaults(self) -> dict[str, float]:
        if not MASTER_PATH.exists():
            return {k: 0.0 for k in ["wireDrawing", "wireEat", "wireHidden", "tapeDiameter", "tapeWidth", "tapeOverlap"]}
        payload = json.loads(MASTER_PATH.read_text(encoding="utf-8"))
        defaults = payload.get("bomDefaults", {})
        return {
            "wireDrawing": float(defaults.get("wireDrawing", 0) or 0),
            "wireEat": float(defaults.get("wireEat", 0) or 0),
            "wireHidden": float(defaults.get("wireHidden", 0) or 0),
            "tapeDiameter": float(defaults.get("tapeDiameter", 0) or 0),
            "tapeWidth": float(defaults.get("tapeWidth", 0) or 0),
            "tapeOverlap": float(defaults.get("tapeOverlap", 0) or 0),
        }

    def load_detail_metrics(self, workbook_path: Path) -> dict[str, Any]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        sheet = workbook[DETAIL_SHEET]
        metrics = {
            "totalMeter": 0.0, "wireMeter": 0.0, "tapeMeter": 0.0, "tubeMeter": 0.0,
            "meterRows": 0, "wireRows": 0, "tapeRows": 0, "tubeRows": 0,
            "samples": [],
        }
        for row_index in range(3, sheet.max_row + 1):
            part_number = str(sheet.cell(row_index, 1).value or "").strip()
            part_name = str(sheet.cell(row_index, 2).value or "").strip()
            quantity = numeric_value(sheet.cell(row_index, 3).value)
            unit = str(sheet.cell(row_index, 4).value or "").strip()
            if quantity is None or unit.upper() != "M":
                continue
            category = self.classify_meter_item(part_number, part_name, unit)
            metrics["totalMeter"] += quantity
            metrics["meterRows"] += 1
            if category == "wire":
                metrics["wireMeter"] += quantity; metrics["wireRows"] += 1
            elif category == "tape":
                metrics["tapeMeter"] += quantity; metrics["tapeRows"] += 1
            else:
                metrics["tubeMeter"] += quantity; metrics["tubeRows"] += 1
            if len(metrics["samples"]) < 8:
                metrics["samples"].append({
                    "partNumber": part_number, "partName": part_name,
                    "quantity": round(quantity, 6), "unit": unit,
                    "category": category, "sourceCell": f"C{row_index}",
                })
        return metrics

    def classify_meter_item(self, part_number: str, part_name: str, unit: str) -> str:
        if unit.upper() != "M": return "other"
        haystack = f"{part_number} {part_name}"
        if any(keyword in haystack for keyword in WIRE_KEYWORDS): return "wire"
        if any(keyword in haystack for keyword in TAPE_KEYWORDS): return "tape"
        return "tube"

    def build_snapshot(self, kind: str, label: str, workbook_path: Path, base_defaults: dict[str, float], current_metrics: dict[str, Any], quote_metrics: dict[str, Any]) -> dict[str, Any]:
        quote_total = float(quote_metrics.get("totalMeter", 0) or 0)
        quote_wire = float(quote_metrics.get("wireMeter", 0) or 0)
        quote_tape = float(quote_metrics.get("tapeMeter", 0) or 0)
        quote_tube = float(quote_metrics.get("tubeMeter", 0) or 0)
        total_meter = float(current_metrics.get("totalMeter", 0) or 0)
        wire_meter = float(current_metrics.get("wireMeter", 0) or 0)
        tape_meter = float(current_metrics.get("tapeMeter", 0) or 0)
        tube_meter = float(current_metrics.get("tubeMeter", 0) or 0)

        material_factor = total_meter / quote_total if quote_total > 0 else 1
        wire_factor = wire_meter / quote_wire if quote_wire > 0 else 1
        tape_factor = tape_meter / quote_tape if quote_tape > 0 else 1
        tube_factor = tube_meter / quote_tube if quote_tube > 0 else 1

        return {
            "kind": kind, "label": label, "workbook": workbook_path.name, "detailSheet": DETAIL_SHEET,
            "materialFactor": round(material_factor, 6), "wireFactor": round(wire_factor, 6),
            "tapeFactor": round(tape_factor, 6), "tubeFactor": round(tube_factor, 6),
            "totalMeter": round(total_meter, 6), "wireMeter": round(wire_meter, 6),
            "tapeMeter": round(tape_meter, 6), "tubeMeter": round(tube_meter, 6),
            "meterRows": int(current_metrics.get("meterRows", 0) or 0),
            "wireRows": int(current_metrics.get("wireRows", 0) or 0),
            "tapeRows": int(current_metrics.get("tapeRows", 0) or 0),
            "tubeRows": int(current_metrics.get("tubeRows", 0) or 0),
            "draft": self.derive_draft(base_defaults, current_metrics, quote_metrics),
            "samples": current_metrics.get("samples", []),
            "sources": {
                "detailSheet": f"{workbook_path.name}!{DETAIL_SHEET}",
                "metricRule": "按《二次物料明细》中单位=M的零件统计导线/胶带/套管总用量；材料系数按总 M 类用量相对报价版的比值估算。",
            },
        }

    def derive_draft(self, base_defaults: dict[str, float], current_metrics: dict[str, Any], quote_metrics: dict[str, Any]) -> dict[str, float]:
        base_drawing = base_defaults["wireDrawing"]
        base_eat = base_defaults["wireEat"]
        base_hidden = base_defaults["wireHidden"]
        base_diameter = base_defaults["tapeDiameter"]
        base_width = base_defaults["tapeWidth"]
        base_overlap = base_defaults["tapeOverlap"]
        base_total_mm = base_drawing + base_eat + base_hidden

        quote_wire = float(quote_metrics.get("wireMeter", 0) or 0)
        current_wire = float(current_metrics.get("wireMeter", 0) or 0)
        quote_tape = float(quote_metrics.get("tapeMeter", 0) or 0)
        current_tape = float(current_metrics.get("tapeMeter", 0) or 0)

        wire_factor = current_wire / quote_wire if quote_wire > 0 else 1
        derived_total_mm = base_total_mm * wire_factor
        derived_drawing = max(derived_total_mm - base_eat - base_hidden, 0)
        
        # derive_overlap logic
        def tape_per_mm(diameter: float, width: float, overlap_pct: float) -> float:
            pitch = width * (1 - overlap_pct / 100)
            circumference = math.pi * diameter
            if pitch <= 0: return 0
            return math.sqrt(circumference * circumference + pitch * pitch) / pitch

        def derive_overlap_val(tape_d: float, tape_w: float, base_o: float, w_f: float, t_f: float) -> float:
            if tape_d <= 0 or tape_w <= 0 or w_f <= 0 or t_f <= 0: return base_o
            base_p_mm = tape_per_mm(tape_d, tape_w, base_o)
            if base_p_mm <= 1: return base_o
            target_p_mm = base_p_mm * (t_f / w_f)
            if target_p_mm <= 1: return base_o
            circumference = math.pi * tape_d
            denominator = max(target_p_mm * target_p_mm - 1, 1e-9)
            pitch = circumference / math.sqrt(denominator)
            overlap = (1 - pitch / tape_w) * 100
            if not math.isfinite(overlap): return base_o
            return max(0, min(95, overlap))

        derived_overlap = derive_overlap_val(base_diameter, base_width, base_overlap, wire_factor, (current_tape / quote_tape if quote_tape > 0 else wire_factor))

        return {
            "bomWireDrawing": round(derived_drawing, 3), "bomWireEat": round(base_eat, 3), "bomWireHidden": round(base_hidden, 3),
            "bomTapeDiameter": round(base_diameter, 3), "bomTapeWidth": round(base_width, 3), "bomTapeOverlap": round(derived_overlap, 3),
        }

    def compare_tt_harness_lengths(self, reference_path: Path, current_path: Path) -> dict[str, Any]:
        old_wb = load_workbook(reference_path, data_only=True, read_only=True)
        new_wb = load_workbook(current_path, data_only=True, read_only=True)
        changes: list[dict[str, Any]] = []

        for sheet_name in new_wb.sheetnames:
            if not sheet_name.isdigit() or sheet_name not in old_wb.sheetnames: continue
            old_sheet = old_wb[sheet_name]; new_sheet = new_wb[sheet_name]
            max_row = max(old_sheet.max_row or 0, new_sheet.max_row or 0)
            for row_index in range(5, max_row + 1):
                old_val = numeric_value(old_sheet.cell(row_index, 10).value)
                new_val = numeric_value(new_sheet.cell(row_index, 10).value)
                if old_val == new_val: continue
                unit = str(new_sheet.cell(row_index, 11).value or old_sheet.cell(row_index, 11).value or "").strip()
                if unit.upper() != "M": continue
                changes.append({
                    "harnessNo": sheet_name, "cell": f"J{row_index}",
                    "partNumber": str(new_sheet.cell(row_index, 3).value or old_sheet.cell(row_index, 3).value or "").strip(),
                    "partName": str(new_sheet.cell(row_index, 4).value or old_sheet.cell(row_index, 4).value or "").strip(),
                    "function": str(new_sheet.cell(row_index, 2).value or old_sheet.cell(row_index, 2).value or "").strip(),
                    "sap": str(new_sheet.cell(row_index, 6).value or old_sheet.cell(row_index, 6).value or "").strip(),
                    "oldQty": old_val, "newQty": new_val, "deltaQty": round((new_val or 0) - (old_val or 0), 6), "unit": unit,
                })

        harnesses = sorted({item["harnessNo"] for item in changes})
        old_total = sum(float(item["oldQty"] or 0) for item in changes)
        new_total = sum(float(item["newQty"] or 0) for item in changes)
        return {
            "referenceWorkbook": reference_path.name, "currentWorkbook": current_path.name,
            "changedHarnessCount": len(harnesses), "changedRowCount": len(changes),
            "oldQtyTotal": round(old_total, 6), "newQtyTotal": round(new_total, 6),
            "deltaQtyTotal": round(new_total - old_total, 6), "harnesses": harnesses, "rows": changes,
        }
