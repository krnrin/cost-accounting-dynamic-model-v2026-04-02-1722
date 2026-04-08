from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import BaseExtractor
from .utils import (
    collapse_text,
    discover_workbook,
    numeric_value,
)

CODE_RE = re.compile(r"^[A-Z]\d{4}$")
SECTION_RE = re.compile(r"^[一二三四五六七八九十]+、")

class CapitalValidationExtractor(BaseExtractor):
    def get_config(self) -> dict[str, Any]:
        return {
            "name": "capital",
            "description": "Equipment and capital investment validation",
            "workbook_pattern": "核算",
            "default_output": "g281_data_capital_validation.json",
            "sheet_configs": [
                {
                    "scopeId": "equipment",
                    "scopeLabel": "设备投资",
                    "sheetIndex": 6,
                    "itemNameHeader": "设备名称",
                    "specHeader": "规格型号",
                    "qtyFieldLabel": "新增数量",
                    "amountFieldLabel": "新增金额",
                    "itemNameCol": 4,
                    "specCol": 5,
                    "unitCol": 6,
                    "demandQtyCol": 7,
                    "demandAmountCol": 8,
                    "reuseQtyCol": 9,
                    "reuseAmountCol": 10,
                    "newQtyCol": 11,
                    "newAmountCol": 12,
                    "unitPriceCol": 13,
                    "noteCol": 14,
                },
                {
                    "scopeId": "tooling",
                    "scopeLabel": "专用模具",
                    "sheetIndex": 7,
                    "itemNameHeader": "名称",
                    "specHeader": "型号",
                    "qtyFieldLabel": "新增数量",
                    "amountFieldLabel": "新增金额",
                    "itemNameCol": 4,
                    "specCol": 5,
                    "unitCol": 6,
                    "demandQtyCol": 7,
                    "demandAmountCol": 8,
                    "reuseQtyCol": 9,
                    "reuseAmountCol": 10,
                    "newQtyCol": 11,
                    "newAmountCol": 12,
                    "unitPriceCol": 14,
                    "noteCol": 13,
                },
                {
                    "scopeId": "fixtures",
                    "scopeLabel": "项目工装",
                    "sheetIndex": 8,
                    "itemNameHeader": "名称",
                    "specHeader": "规格型号",
                    "qtyFieldLabel": "预估/实际数量",
                    "amountFieldLabel": "预估/实际金额",
                    "itemNameCol": 4,
                    "specCol": 5,
                    "unitCol": 6,
                    "demandQtyCol": 7,
                    "demandAmountCol": 8,
                    "reuseQtyCol": 9,
                    "reuseAmountCol": 10,
                    "newQtyCol": 11,
                    "newAmountCol": 12,
                    "unitPriceCol": 14,
                    "noteCol": 13,
                },
            ]
        }

    def extract(self, input_path: Path) -> dict[str, Any]:
        directory = input_path if input_path.is_dir() else input_path.parent
        quote_path = discover_workbook("报价核算", directory)
        fixed_path = discover_workbook("定点核算", directory)
        return self.build_payload(quote_path, fixed_path)

    def build_payload(self, quote_workbook: Path, fixed_workbook: Path) -> dict[str, Any]:
        config = self.get_config()
        comparisons: dict[str, Any] = {}
        scope_order = []

        for cfg in config["sheet_configs"]:
            quote_scope = self.parse_scope(quote_workbook, cfg)
            fixed_scope = self.parse_scope(fixed_workbook, cfg)
            comparison = self.align_scope(quote_scope, fixed_scope)
            comparisons[cfg["scopeId"]] = comparison
            scope_order.append(cfg["scopeId"])

        return {
            "meta": {
                "quoteWorkbook": quote_workbook.name,
                "fixedWorkbook": fixed_workbook.name,
                "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
                "layer": "JSON 数据层",
            },
            "scopeOrder": scope_order,
            "comparisons": comparisons,
            "unmatchedScopes": [scope_id for scope_id in scope_order if comparisons[scope_id]["summary"]["quoteOnlyCount"] or comparisons[scope_id]["summary"]["fixedOnlyCount"]],
        }

    def parse_scope(self, workbook_path: Path, config: dict[str, Any]) -> dict[str, Any]:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet_name = workbook.sheetnames[config["sheetIndex"]]
        ws = workbook[sheet_name]

        sections: list[dict[str, Any]] = []
        current_section_label = collapse_text(ws["A2"].value) or config["scopeLabel"]
        
        def normalize_key(*parts: Any) -> str:
            values = [collapse_text(part).upper().replace(" ", "") for part in parts if collapse_text(part)]
            return "|".join(values)

        def build_section_key(sheet_id: str, label: str) -> str:
            return normalize_key(sheet_id, label) or f"{sheet_id}-default"

        current_section_key = build_section_key(config["scopeId"], current_section_label)
        current_items: list[dict[str, Any]] = []
        occurrence_map: dict[str, int] = defaultdict(int)

        def flush_section() -> None:
            if not current_items: return
            sections.append({"key": current_section_key, "label": current_section_label, "items": current_items.copy()})
            current_items.clear()
            occurrence_map.clear()

        for row in ws.iter_rows(min_row=1, values_only=True):
            first = collapse_text(row[0] if len(row) > 0 else "")
            if SECTION_RE.match(first) or (first.endswith("投入") and not CODE_RE.match(first) and "编号" not in first):
                flush_section()
                current_section_label = first
                current_section_key = build_section_key(config["scopeId"], current_section_label)
                continue

            if not CODE_RE.match(first): continue

            category = collapse_text(row[1] if len(row) > 1 else "")
            investment_name = collapse_text(row[2] if len(row) > 2 else "")
            item_name = collapse_text(row[config["itemNameCol"] - 1] if len(row) >= config["itemNameCol"] else "")
            spec = collapse_text(row[config["specCol"] - 1] if len(row) >= config["specCol"] else "")
            unit = collapse_text(row[config["unitCol"] - 1] if len(row) >= config["unitCol"] else "")
            base_key = normalize_key(first, investment_name, item_name, spec)
            if not base_key: continue
            occurrence_map[base_key] += 1
            item_key = f"{base_key}#{occurrence_map[base_key]:02d}"

            item = {
                "itemKey": item_key, "code": first, "category": category, "investmentName": investment_name,
                "itemName": item_name, "spec": spec, "unit": unit,
                "demandQty": numeric_value(row[config["demandQtyCol"] - 1] if len(row) >= config["demandQtyCol"] else None),
                "demandAmount": numeric_value(row[config["demandAmountCol"] - 1] if len(row) >= config["demandAmountCol"] else None),
                "reuseQty": numeric_value(row[config["reuseQtyCol"] - 1] if len(row) >= config["reuseQtyCol"] else None),
                "reuseAmount": numeric_value(row[config["reuseAmountCol"] - 1] if len(row) >= config["reuseAmountCol"] else None),
                "newQty": numeric_value(row[config["newQtyCol"] - 1] if len(row) >= config["newQtyCol"] else None),
                "newAmount": numeric_value(row[config["newAmountCol"] - 1] if len(row) >= config["newAmountCol"] else None),
                "unitPrice": numeric_value(row[config["unitPriceCol"] - 1] if len(row) >= config["unitPriceCol"] else None),
                "note": collapse_text(row[config["noteCol"] - 1] if len(row) >= config["noteCol"] else ""),
                "sectionLabel": current_section_label, "sheetName": sheet_name,
                "qtyLabel": config["qtyFieldLabel"], "amountLabel": config["amountFieldLabel"],
            }
            current_items.append(item)

        flush_section()
        total_new_amount = sum(float(item["newAmount"] or 0) for section in sections for item in section["items"])
        total_new_qty = sum(float(item["newQty"] or 0) for section in sections for item in section["items"] if item["newQty"] is not None)

        return {
            "scopeId": config["scopeId"], "scopeLabel": config["scopeLabel"], "sheetName": sheet_name,
            "itemNameHeader": config["itemNameHeader"], "specHeader": config["specHeader"],
            "qtyFieldLabel": config["qtyFieldLabel"], "amountFieldLabel": config["amountFieldLabel"],
            "sections": sections,
            "summary": {
                "sectionCount": len(sections),
                "itemCount": sum(len(section["items"]) for section in sections),
                "totalNewAmount": round(total_new_amount, 6),
                "totalNewQty": round(total_new_qty, 6),
            },
        }

    def align_scope(self, quote_scope: dict[str, Any], fixed_scope: dict[str, Any]) -> dict[str, Any]:
        quote_sections = OrderedDict((section["key"], section) for section in quote_scope["sections"])
        fixed_sections = OrderedDict((section["key"], section) for section in fixed_scope["sections"])
        section_order = list(quote_sections.keys())
        for key in fixed_sections.keys():
            if key not in quote_sections: section_order.append(key)

        comparison_groups = []
        total_matched = 0; total_quote_only = 0; total_fixed_only = 0

        for section_key in section_order:
            quote_section = quote_sections.get(section_key)
            fixed_section = fixed_sections.get(section_key)
            label = quote_section["label"] if quote_section else fixed_section["label"]
            quote_items = quote_section["items"] if quote_section else []
            fixed_items = fixed_section["items"] if fixed_section else []
            fixed_map = OrderedDict((item["itemKey"], item) for item in fixed_items)
            aligned = []; matched = 0; quote_only = 0; fixed_only = 0

            for quote_item in quote_items:
                fixed_item = fixed_map.pop(quote_item["itemKey"], None)
                status = "matched" if fixed_item else "quote_only"
                if status == "matched": matched += 1
                else: quote_only += 1
                aligned.append({"status": status, "quote": quote_item, "fixed": fixed_item})

            for fixed_item in fixed_map.values():
                fixed_only += 1
                aligned.append({"status": "fixed_only", "quote": None, "fixed": fixed_item})

            total_matched += matched; total_quote_only += quote_only; total_fixed_only += fixed_only
            comparison_groups.append({
                "key": section_key, "label": label, "quoteCount": len(quote_items), "fixedCount": len(fixed_items),
                "matchedCount": matched, "quoteOnlyCount": quote_only, "fixedOnlyCount": fixed_only,
                "quoteAmount": round(sum(float(item["newAmount"] or 0) for item in quote_items), 6),
                "fixedAmount": round(sum(float(item["newAmount"] or 0) for item in fixed_items), 6),
                "aligned": aligned,
            })

        return {
            "scopeId": quote_scope["scopeId"], "scopeLabel": quote_scope["scopeLabel"],
            "quoteSheet": quote_scope["sheetName"], "fixedSheet": fixed_scope["sheetName"],
            "quoteSummary": quote_scope["summary"], "fixedSummary": fixed_scope["summary"],
            "summary": {
                "sectionCount": len(comparison_groups), "quoteCount": quote_scope["summary"]["itemCount"],
                "fixedCount": fixed_scope["summary"]["itemCount"], "matchedCount": total_matched,
                "quoteOnlyCount": total_quote_only, "fixedOnlyCount": total_fixed_only,
                "quoteAmount": quote_scope["summary"]["totalNewAmount"],
                "fixedAmount": fixed_scope["summary"]["totalNewAmount"],
                "deltaAmount": round(float(fixed_scope["summary"]["totalNewAmount"] or 0) - float(quote_scope["summary"]["totalNewAmount"] or 0), 6),
                "quoteQty": quote_scope["summary"]["totalNewQty"], "fixedQty": fixed_scope["summary"]["totalNewQty"],
            },
            "fields": {
                "itemNameHeader": quote_scope["itemNameHeader"], "specHeader": quote_scope["specHeader"],
                "qtyFieldLabel": quote_scope["qtyFieldLabel"], "amountFieldLabel": quote_scope["amountFieldLabel"],
            },
            "groups": comparison_groups,
        }
