from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

def discover_workbook(pattern: str, directory: str = ".") -> Path:
    """Find workbook in directory by pattern."""
    matches = [
        Path(directory) / name
        for name in os.listdir(directory)
        if name.endswith(".xlsx") and not name.startswith("~$") and pattern in name and "导线联动" not in name
    ]
    if not matches:
        raise FileNotFoundError(f"Workbook containing '{pattern}' not found in directory '{directory}'.")
    return sorted(matches)[0]

def collapse_text(value: Any) -> str:
    """Normalize whitespace and strip text."""
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\s+", " ", text)

def to_text(value: Any) -> str:
    """Basic string conversion and strip."""
    return str(value or "").strip()

def numeric_value(value: Any) -> float | None:
    """Convert value to float or None."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = collapse_text(value).replace(",", "")
    if not text or text.startswith("#"):
        return None
    try:
        return float(text)
    except ValueError:
        return None

def round_number(value: float | None, digits: int = 6) -> float | None:
    """Round value to specified digits."""
    if value is None:
        return None
    return round(float(value), digits)

def note_join(*parts: str) -> str:
    """Join parts with Chinese semicolon."""
    return "；".join(part for part in parts if collapse_text(part))

def write_json(data: dict[str, Any], path: Path, dry_run: bool = False):
    """Write dictionary to JSON file."""
    output = json.dumps(data, ensure_ascii=False, indent=2)
    if dry_run:
        print(f"Dry run: skip writing {len(output)} chars to {path}")
    else:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(output, encoding="utf-8")
        print(f"Exported data to {path}")

def make_item(
    item_key: str,
    label: str,
    unit: str,
    source_sheet: str,
    source_cell: str,
    raw_value: Any,
    formula_value: Any = "",
    *,
    numeric_override: float | None = None,
    note: str = "",
    display_label: str = "",
) -> dict[str, Any]:
    """Helper to create standard data item."""
    numeric = round_number(numeric_override if numeric_override is not None else numeric_value(raw_value))
    raw_text = collapse_text(raw_value)
    formula_text = collapse_text(formula_value)
    value: Any = numeric if numeric is not None else (raw_text or "-")
    value_type = "number" if numeric is not None else ("error" if raw_text.startswith("#") else "text")
    return {
        "itemKey": item_key,
        "label": label,
        "displayLabel": display_label or label,
        "unit": unit,
        "value": value,
        "numericValue": numeric,
        "valueType": value_type,
        "sourceSheet": source_sheet,
        "sourceCell": source_cell,
        "source": f"{source_sheet}!{source_cell}",
        "formula": formula_text,
        "note": collapse_text(note),
        "rawValue": raw_text,
    }

def workbook_context(path: Path, sheet_indices: dict[str, int]) -> dict[str, Any]:
    """Load both data and formula workbooks."""
    value_book = load_workbook(path, read_only=True, data_only=True)
    formula_book = load_workbook(path, read_only=True, data_only=False)
    sheets = {}
    for alias, index in sheet_indices.items():
        if index >= len(value_book.sheetnames):
            continue
        title = value_book.sheetnames[index]
        sheets[alias] = {
            "title": title,
            "values": value_book[title],
            "formulas": formula_book[formula_book.sheetnames[index]],
        }
    return {
        "path": path,
        "sheets": sheets,
        "value_book": value_book,
        "formula_book": formula_book,
    }
