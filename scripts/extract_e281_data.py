"""
从 E281 参考表格提取数据，输出 JSON 供 seed 脚本使用。
数据来源：
  - BOM: 参考表格/E281/E281项目 报价BOM V01-11.3.xlsx
  - 报价: 参考表格/E281/吉利E281报价核算.xlsx
"""
import json, re, sys
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"D:\成本核算动态模型\参考表格\E281")
BOM_FILE = BASE / "E281项目 报价BOM V01-11.3.xlsx"
QUOTE_FILE = BASE / "吉利E281高压财务可行性分析-1125-客户目标价 - V001.xlsx"
OUT_DIR = Path(r"D:\成本核算动态模型\scripts\e281_extracted")
OUT_DIR.mkdir(exist_ok=True)

def num(v, default=0):
    """安全转数字"""
    if v is None: return default
    try: return float(v)
    except (ValueError, TypeError): return default

# ═══════════════════════════════════════════════════
# Part 0: 从报价核算文件读取二次物料明细（有完整价格数据）
# ═══════════════════════════════════════════════════
print("=== 加载报价核算文件（读取二次物料明细）===")
wb_q_mat = load_workbook(str(QUOTE_FILE), data_only=True, read_only=True)

# 报价核算的二次物料明细: 15列
# Row 0: 注释行, Row 1: Header
# Col: 0=组件描述(零件号), 1=物料名称, 2=用量, 3=单位, 4=供应商, 5=SAP号,
#      6=备注, 7=责任人, 8=单价, 9=铝重(g), 10=铜重(g), 11=非铜(元),
#      12=币别, 13=价格类型, 14=备注
print("\n--- 读取二次物料明细（报价核算文件）---")
ws_mat = wb_q_mat["二次物料明细"]
material_db = {}
for i, row in enumerate(ws_mat.iter_rows(min_row=1, values_only=True)):
    if i == 0: continue  # skip 注释行
    if i == 1:
        print(f"  Header: {[str(c or '')[:15] for c in row[:15]]}")
        continue
    if not row[0]: continue
    part_no = str(row[0]).strip()
    material_db[part_no] = {
        "name": str(row[1] or "").strip(),
        "unit": str(row[3] or "").strip(),
        "supplier": str(row[4] or "").strip(),
        "sapNo": str(row[5] or "").strip(),
        "remark": str(row[6] or "").strip(),
        "unitPrice": num(row[8]),      # col 8: 单价（非导线部分）
        "alWeight_g": num(row[9]),     # col 9: 铝重(g)
        "cuWeight_g": num(row[10]),    # col 10: 铜重(g)
        "nonMetal": num(row[11]),      # col 11: 非铜(元)
    }
print(f"  物料主数据: {len(material_db)} 条")
wb_q_mat.close()

# ═══════════════════════════════════════════════════
# Part 1: 从 BOM 文件提取 BOM 数据
# ═══════════════════════════════════════════════════
print("\n=== 加载 BOM 文件 ===")
wb_bom = load_workbook(str(BOM_FILE), data_only=True, read_only=True)
print(f"  Sheets: {wb_bom.sheetnames}")

# 1b. 识别线束 sheet（10位数字零件号）并提取 BOM
print("\n--- 提取各线束 BOM ---")
harness_sheets = [s for s in wb_bom.sheetnames if re.match(r'^\d{10}$', s)]
print(f"  线束 sheets: {harness_sheets}")

# 物料类型映射
def classify_item(mat_type, part_name):
    """根据二次物料明细的类型字段 + 零件名称判断 itemCategory"""
    t = (mat_type or "").strip()
    n = (part_name or "").strip()
    if "导线" in t or "导线" in n or "屏蔽" in n:
        return "wire"
    if "连接器" in t or "连接器" in n or "插头" in n or "插座" in n:
        return "connector"
    if "端子" in t or "端子" in n:
        if "IPT" in n.upper() or "焊接" in n:
            return "ipt_terminal"
        return "terminal"
    if "支架" in n or "橡胶" in n or "卡扣" in n or "护套" in n:
        return "bracket_rubber"
    if "胶带" in n or "套管" in n or "扎带" in n or "热缩" in n or "波纹" in n:
        return "tape_tube"
    return "other"

bom_data = {}  # harnessId -> list of items

for sheet_name in harness_sheets:
    ws = wb_bom[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Row 2: metadata (col D=harnessNo, col E=harnessName)
    harness_id = str(rows[1][3] or sheet_name).strip() if len(rows) > 1 else sheet_name
    harness_name = str(rows[1][4] or "").strip() if len(rows) > 1 else ""

    # Row 4: headers, Row 5+: data
    # Columns: 0=NO, 1=Function, 2=PartNumber, 3=PartName, 4=SemiFinished,
    #          5=WireNO, 6=PIN, 7=OPTION, 8=SPEC, 9=Quantity, 10=Unit, 11=Remark
    items = []
    for r in rows[4:]:  # skip header rows (0-3)
        if not r[2]: continue  # no part number = skip
        part_no = str(r[2]).strip()
        part_name = str(r[3] or "").strip()
        is_semi = str(r[4] or "").strip().upper() == "Y"
        spec = str(r[8] or "").strip()
        qty = num(r[9])
        unit = str(r[10] or "").strip()

        if qty == 0: continue

        # Lookup from 二次物料明细
        mat = material_db.get(part_no, {})
        mat_type = mat.get("type", "")
        supplier = mat.get("supplier", "")
        sap_no = mat.get("sapNo", "")
        unit_price = mat.get("unitPrice", 0)
        cu_g = mat.get("cuWeight_g", 0)
        al_g = mat.get("alWeight_g", 0)
        non_metal = mat.get("nonMetal", 0)

        category = classify_item(mat_type, part_name)

        item = {
            "partNo": part_no,
            "partName": part_name,
            "itemCategory": category,
            "qty": qty,
            "unit": unit,
            "unitPrice": unit_price,
            "amount": round(unit_price * qty, 10),
            "isSemiFinished": is_semi,
        }
        if sap_no: item["sapNo"] = sap_no
        if spec: item["spec"] = spec
        if supplier: item["supplier"] = supplier
        if str(r[1] or "").strip(): item["functionText"] = str(r[1]).strip()

        # 导线类型：额外字段
        if category == "wire":
            item["copperWeightPerUnit"] = cu_g / 1000  # g -> kg
            item["aluminumWeightPerUnit"] = al_g / 1000
            item["nonMetalCostPerUnit"] = non_metal / 1000  # 元/千单位 -> 元/单位
            # 导线的 unitPrice 和 amount 由引擎动态计算，设为 0
            item["unitPrice"] = 0
            item["amount"] = 0

        items.append(item)

    bom_data[harness_id] = items
    print(f"  {harness_id} ({harness_name}): {len(items)} items")

wb_bom.close()

# 保存 BOM JSON
with open(OUT_DIR / "e281_bom.json", "w", encoding="utf-8") as f:
    json.dump(bom_data, f, ensure_ascii=False, indent=2)
print(f"\n  BOM 数据已保存: {OUT_DIR / 'e281_bom.json'}")

# ═══════════════════════════════════════════════════
# Part 2: 从报价核算文件提取项目/线束数据
# ═══════════════════════════════════════════════════
print("\n=== 加载报价核算文件 ===")
wb_q = load_workbook(str(QUOTE_FILE), data_only=True, read_only=True)
print(f"  Sheets: {wb_q.sheetnames}")

# 2a. 配置明细 → 装车比 + 前后工时
print("\n--- 配置明细 ---")
ws_cfg = wb_q["配置明细"]
cfg_rows = list(ws_cfg.iter_rows(min_row=1, values_only=True))
# 找到表头行（含"零件号"的行）
header_idx = None
for i, r in enumerate(cfg_rows):
    row_str = " ".join(str(c or "") for c in r[:5])
    if "零件号" in row_str or "零件编号" in row_str:
        header_idx = i
        print(f"  Header at row {i}: {[str(c or '')[:20] for c in r[:20]]}")
        break

harness_meta = {}  # harnessId -> {name, ratio, frontHours, backHours}
if header_idx is not None:
    # 根据 agent 探索结果：装车比在前面的列，工时在后面
    # 需要找到具体列位置
    hdr = cfg_rows[header_idx]
    hdr_str = [str(c or "").strip() for c in hdr]
    print(f"  Full header: {hdr_str[:25]}")

    # 找关键列索引
    def find_col(keywords, hdr_list):
        for i, h in enumerate(hdr_list):
            for kw in keywords:
                if kw in h: return i
        return None

    col_id = find_col(["零件号", "零件编号"], hdr_str)
    col_name = find_col(["零件名称", "名称"], hdr_str)
    col_ratio = find_col(["数量", "用量", "装车比", "配置"], hdr_str)
    col_front = find_col(["公共制成", "公共制程"], hdr_str)
    col_back = find_col(["后工程", "总装工时"], hdr_str)
    col_mat = find_col(["材料成本", "材料"], hdr_str)

    print(f"  列索引: id={col_id}, name={col_name}, ratio={col_ratio}, front={col_front}, back={col_back}, mat={col_mat}")

    for r in cfg_rows[header_idx+1:]:
        if not r[col_id]: continue
        hid = str(r[col_id]).strip()
        if not re.match(r'^\d{10}$', hid): continue
        harness_meta[hid] = {
            "name": str(r[col_name] or "").strip() if col_name else "",
            "ratio": num(r[col_ratio]) if col_ratio else 0,
            "frontHours": num(r[col_front]) if col_front else 0,
            "backHours": num(r[col_back]) if col_back else 0,
            "materialCost": num(r[col_mat]) if col_mat else None,
        }
        print(f"  {hid}: ratio={harness_meta[hid]['ratio']}, front={harness_meta[hid]['frontHours']}, back={harness_meta[hid]['backHours']}")

# 2b. 包装物流费用
print("\n--- 包装物流费用 ---")
ws_pkg = wb_q["包装物流费用"]
pkg_rows = list(ws_pkg.iter_rows(min_row=1, values_only=True))
# 找表头
pkg_header_idx = None
for i, r in enumerate(pkg_rows):
    row_str = " ".join(str(c or "") for c in r[:8])
    if "零件号" in row_str or "客户零件号" in row_str:
        pkg_header_idx = i
        print(f"  Header at row {i}: {[str(c or '')[:15] for c in r[:15]]}")
        break

packaging_data = {}  # harnessId -> {innerPack, outerPack, freight, exFreight, shortHaul, thirdParty, storage}
if pkg_header_idx is not None:
    phdr = [str(c or "").strip() for c in pkg_rows[pkg_header_idx]]
    print(f"  Full header: {phdr[:15]}")

    col_pid = find_col(["零件号", "客户零件号"], phdr)
    col_inner = find_col(["内包装"], phdr)
    col_outer = find_col(["外包装"], phdr)
    col_freight_main = find_col(["运费"], phdr)
    col_ex_freight = find_col(["超额"], phdr)
    col_short = find_col(["短驳"], phdr)
    col_third = find_col(["三方仓"], phdr)
    col_store = find_col(["仓储"], phdr)

    print(f"  列索引: id={col_pid}, inner={col_inner}, outer={col_outer}, short={col_short}, third={col_third}, store={col_store}")

    for r in pkg_rows[pkg_header_idx+1:]:
        if not r[col_pid]: continue
        pid = str(r[col_pid]).strip()
        if not re.match(r'^\d{10}$', pid): continue
        packaging_data[pid] = {
            "innerPack": num(r[col_inner]) if col_inner is not None else 0,
            "outerPack": num(r[col_outer]) if col_outer is not None else 0,
            "freight": num(r[col_freight_main]) if col_freight_main is not None else 0,
            "exFreight": num(r[col_ex_freight]) if col_ex_freight is not None else 0,
            "shortHaul": num(r[col_short]) if col_short is not None else 0,
            "thirdParty": num(r[col_third]) if col_third is not None else 0,
            "storage": num(r[col_store]) if col_store is not None else 0,
        }
        p = packaging_data[pid]
        print(f"  {pid}: inner={p['innerPack']}, outer={p['outerPack']}, short={p['shortHaul']}, 3rd={p['thirdParty']}, store={p['storage']}")

# 2c. 客户报价逻辑 → 验证数据
print("\n--- 客户报价逻辑 (验证用) ---")
ws_quote = wb_q["客户报价逻辑"]
quote_rows = list(ws_quote.iter_rows(min_row=1, values_only=True))
# Row 2 = header, Row 4-14 = data
qhdr = [str(c or "").strip() for c in quote_rows[1]] if len(quote_rows) > 1 else []
print(f"  Header: {qhdr[:20]}")

verification = {}
for r in quote_rows[3:14]:  # rows 4-14 (0-indexed: 3-13)
    if not r[1]: continue
    hid = str(r[1]).strip()
    verification[hid] = {
        "name": str(r[2] or "").strip(),
        "ratio": num(r[3]),
        "cuWeight": num(r[4]),
        "alWeight": num(r[5]),
        "materialCost": num(r[6]),
        "wasteCost": num(r[7]),
        "totalHours": num(r[8]),
        "directLabor": num(r[9]),
        "mfgCost": num(r[10]),
        "laborPlusMfg": num(r[11]),
        "mgmtFee": num(r[12]),
        "profit": num(r[13]),
        "exFactoryPrice": num(r[14]),
        "packCost": num(r[15]),
        "freightCost": num(r[16]),
        "deliveredPrice": num(r[17]),
    }
    v = verification[hid]
    print(f"  {hid} ({v['name']}): mat={v['materialCost']}, exFactory={v['exFactoryPrice']}, delivered={v['deliveredPrice']}")

# 2d. 运营工时费报价基准 → 内部费率
print("\n--- 运营工时费报价基准 ---")
ws_rate = wb_q["运营工时费报价基准"]
rate_rows = list(ws_rate.iter_rows(min_row=1, values_only=True))
# 打印前30行以便定位
for i, r in enumerate(rate_rows[:35]):
    vals = [str(c or "")[:25] for c in r[:10]]
    print(f"  Row {i}: {vals}")

# 2e. 项目评估汇总 → 年产量计划
print("\n--- 项目评估汇总 ---")
ws_eval = wb_q["项目评估汇总（昆山90%）"]
eval_rows = list(ws_eval.iter_rows(min_row=1, values_only=True))
# 打印前15行找产量数据
for i, r in enumerate(eval_rows[:15]):
    vals = [str(c or "")[:20] for c in r[:15]]
    print(f"  Row {i}: {vals}")

wb_q.close()

# ═══════════════════════════════════════════════════
# Part 3: 合并输出
# ═══════════════════════════════════════════════════
print("\n=== 合并输出 ===")

# 合并 harness 数据
harness_output = []
for hid in harness_sheets:
    meta = harness_meta.get(hid, {})
    pkg = packaging_data.get(hid, {})
    ver = verification.get(hid, {})
    harness_output.append({
        "harnessId": hid,
        "harnessName": meta.get("name") or ver.get("name", ""),
        "vehicleRatio": meta.get("ratio") or ver.get("ratio", 0),
        "frontHours": meta.get("frontHours", 0),
        "backHours": meta.get("backHours", 0),
        "totalHours": ver.get("totalHours", 0),
        "innerPack": pkg.get("innerPack", 0),
        "outerPack": pkg.get("outerPack", 0),
        "freight": pkg.get("freight", 0),
        "exFreight": pkg.get("exFreight", 0),
        "shortHaul": pkg.get("shortHaul", 0),
        "thirdParty": pkg.get("thirdParty", 0),
        "storage": pkg.get("storage", 0),
        # 验证字段
        "verify_materialCost": ver.get("materialCost", 0),
        "verify_exFactoryPrice": ver.get("exFactoryPrice", 0),
        "verify_deliveredPrice": ver.get("deliveredPrice", 0),
    })

with open(OUT_DIR / "e281_harness.json", "w", encoding="utf-8") as f:
    json.dump(harness_output, f, ensure_ascii=False, indent=2)
print(f"  线束数据已保存: {OUT_DIR / 'e281_harness.json'}")

# 项目级数据
project_data = {
    "projectCode": "E281",
    "projectName": "吉利E281高压线束",
    "customer": "吉利汽车",
    "platform": "E281",
    "lifecycleYears": 6,
    "status": "quoted",
    "costRates": {
        "laborRate": 35,
        "mfgRate": 46.69,
        "wasteRate": 0.01,
        "mgmtRate": 0.06,
        "profitRate": 0.056627,
    },
    "metalPrices": {
        "copper": 76450,
        "aluminum": 18910,
    },
    # 年产量从项目评估汇总提取（待确认具体行列）
    "volumes": [
        {"year": 1, "volume": 85000},
        {"year": 2, "volume": 135000},
        {"year": 3, "volume": 125000},
        {"year": 4, "volume": 114000},
        {"year": 5, "volume": 94000},
        {"year": 6, "volume": 47000},
    ],
    "annualDropRate": 0.02,
}

with open(OUT_DIR / "e281_project.json", "w", encoding="utf-8") as f:
    json.dump(project_data, f, ensure_ascii=False, indent=2)
print(f"  项目数据已保存: {OUT_DIR / 'e281_project.json'}")

# 验证：BOM 材料成本汇总
print("\n=== 验证：BOM 材料成本 vs 报价核算 ===")
COPPER_PRICE = 76.45   # 元/kg
ALUMINUM_PRICE = 18.91  # 元/kg

for hid, items in bom_data.items():
    total_mat = 0
    for item in items:
        if item["itemCategory"] == "wire":
            # 导线成本 = (铝重×铝价 + 铜重×铜价 + 非金属) × 数量
            wire_price = (item["aluminumWeightPerUnit"] * ALUMINUM_PRICE +
                         item["copperWeightPerUnit"] * COPPER_PRICE +
                         item["nonMetalCostPerUnit"])
            total_mat += wire_price * item["qty"]
        else:
            total_mat += item["amount"]
    ver = verification.get(hid, {})
    expected = ver.get("materialCost", 0)
    diff = total_mat - expected if expected else 0
    status = "✅" if abs(diff) < 0.1 or expected == 0 else f"⚠️ diff={diff:.4f}"
    print(f"  {hid}: BOM计算={total_mat:.2f}, 报价核算={expected}, {status}")

print("\n=== 完成 ===")
