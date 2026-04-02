from __future__ import annotations

import argparse
import json
import os
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_KEYS = [
    "no",
    "function",
    "partNumber",
    "partName",
    "semiFinished",
    "wireNo",
    "pin",
    "option",
    "spec",
    "quantity",
    "unit",
    "remark",
    "otherRemark",
    "subPartNumber",
    "subPartName",
    "subPartQuantity",
    "subPartUnit",
]

GROUP_LABELS = {
    "battery_end": "接电池端",
    "edrive_end": "接电驱端",
    "accm_end": "接ACCM端",
    "ptc_end": "接PTC端",
    "branch_splitter": "分线器",
    "charge_socket": "充电插座",
    "dc_charge_end": "快充端",
    "ac_charge_end": "慢充端",
    "electronic_lock": "电子锁",
    "low_voltage_inline": "低压连接器",
    "dc_ground": "DC 接地端子",
    "ac_ground": "AC 接地端子",
    "connector_misc": "其他连接器",
    "wires": "导线",
    "sync_brackets": "支架类（同步开发件）",
    "sync_rubber": "橡胶件类（同步开发件）",
    "materials": "其他物料",
}

GROUP_SECTIONS = {
    "wires": "wire",
    "sync_brackets": "sync",
    "sync_rubber": "sync",
    "materials": "material",
}

GROUP_ORDER = [
    "battery_end",
    "edrive_end",
    "accm_end",
    "ptc_end",
    "branch_splitter",
    "charge_socket",
    "dc_charge_end",
    "ac_charge_end",
    "electronic_lock",
    "low_voltage_inline",
    "dc_ground",
    "ac_ground",
    "connector_misc",
    "wires",
    "sync_brackets",
    "sync_rubber",
    "materials",
]

DEFAULT_VERSION_SPECS = OrderedDict(
    {
        "quote": {"label": "报价 BOM", "pattern": "V03-12.4"},
        "fixed": {"label": "定点 BOM", "pattern": "V06-2026.01.20"},
        "tt": {"label": "TT BOM", "pattern": "TT_"},
    }
)
ACTIVE_VERSION_SPECS = OrderedDict(DEFAULT_VERSION_SPECS)


def source_keys() -> tuple[str, ...]:
    return tuple(ACTIVE_VERSION_SPECS.keys())


def display_source_keys() -> tuple[str, ...]:
    return tuple(reversed(source_keys()))


def base_source() -> str:
    return source_keys()[0]


def version_label(source: str) -> str:
    return ACTIVE_VERSION_SPECS[source]["label"]


def version_pattern(source: str) -> str:
    return ACTIVE_VERSION_SPECS[source]["pattern"]

END_GROUP_RULES = [
    ("battery_end", ("接电池端", "电池端")),
    ("edrive_end", ("接电驱端", "接电驱", "电驱端")),
    ("accm_end", ("接ACCM端", "ACCM端")),
    ("ptc_end", ("接PTC", "PTC端")),
    ("branch_splitter", ("二分四分线器", "分线器")),
    ("charge_socket", ("组合式充电插座", "充电插座")),
    ("dc_charge_end", ("快充连接器（直流）", "快充端连接器", "DC 10PIN低压信号连接器", "DC 8PIN低压信号连接器")),
    ("ac_charge_end", ("慢充端连接器（交流）", "慢充连接器（交流）", "AC 5PIN低压信号连接器", "AC 6PIN低压信号连接器")),
    ("electronic_lock", ("电子锁低压连接器总成", "电子锁低压连接器", "电子锁")),
    ("low_voltage_inline", ("低压INLINE连接器总成", "低压inline连接器总成", "低压连接器总成")),
    ("dc_ground", ("DC接地端子",)),
    ("ac_ground", ("AC接地端子",)),
]

CONNECTOR_KEYWORDS = (
    "连接器",
    "护套",
    "端子",
    "屏蔽",
    "挡板",
    "尾盖",
    "密封圈",
    "线卡",
    "插头",
    "插座",
    "主体",
    "组合件",
    "固定件",
    "线密封圈",
    "电子锁",
    "低压",
    "壳体",
)

SYNC_BRACKET_KEYWORDS = (
    "支架",
    "安装支架",
    "金属支架",
    "塑料支架",
)

SYNC_RUBBER_KEYWORDS = (
    "橡胶件",
    "橡胶",
    "胶件",
)


ASSEMBLY_MAPPING_KEYWORDS = (
    "低压",
    "高压",
    "直流",
    "交流",
    "ODP",
    "AC",
    "DC",
    "电池",
    "电驱",
    "充电",
    "插件",
    "插头",
    "插座",
    "互锁",
    "屏蔽",
    "防水",
    "端子",
)


@dataclass
class ParsedSheet:
    harness_id: str
    harness_name: str
    sheet_name: str
    source: str
    items: list[dict[str, Any]]


def harness_key(value: Any) -> str:
    text = collapse_text(value)
    if not text:
        return ""
    match = re.search(r"(\d{10})", text)
    return match.group(1) if match else text


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def collapse_text(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    text = collapse_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).upper()


def numeric_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if abs(number - round(number)) < 1e-9:
        return int(round(number))
    return round(number, 6)


def first_digits(sheet_name: str) -> str | None:
    match = re.match(r"^(\d{10})", str(sheet_name))
    return match.group(1) if match else None


def discover_workbook(pattern: str) -> Path:
    for name in os.listdir("."):
        if name.endswith(".xlsx") and not name.startswith("~$") and pattern in name:
            return Path(name)
    raise FileNotFoundError(f"Workbook containing '{pattern}' not found in current directory.")


def read_ksk_lookup(path: Path) -> dict[str, dict[str, list[dict[str, Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next(name for name in workbook.sheetnames if "KSK" in name)
    worksheet = workbook[sheet_name]
    lookup: dict[str, dict[str, list[dict[str, Any]]]] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        harness_id = harness_key(row[0] if len(row) > 0 else "")
        if not harness_id:
            continue

        part_number = collapse_text(row[3] if len(row) > 3 else "")
        part_name = collapse_text(row[4] if len(row) > 4 else "")
        item_key = item_key_for_row(part_number, part_name)
        if not item_key:
            continue

        record = {
            "assemblyRef": collapse_text(row[2] if len(row) > 2 else ""),
            "sapNo": collapse_text(row[6] if len(row) > 6 else ""),
            "supplier": collapse_text(row[12] if len(row) > 12 else ""),
            "otherRemark": collapse_text(row[13] if len(row) > 13 else ""),
        }
        lookup.setdefault(harness_id, {}).setdefault(item_key, []).append(record)

    return lookup


def detect_end_group(function_text: str) -> str | None:
    haystack = collapse_text(function_text).upper()
    if not haystack:
        return None
    for group_key, keywords in END_GROUP_RULES:
        for keyword in keywords:
            if keyword.upper() in haystack:
                return group_key
    return None


def is_wire_row(part_number: str, part_name: str, unit: str) -> bool:
    haystack = " ".join(filter(None, [part_number, part_name])).upper()
    if "导线" in haystack:
        return True
    if part_number and re.search(r"/\d+(?:\.\d+)?/", part_number):
        return True
    if unit == "M" and ("屏蔽" in haystack or "导线" in haystack):
        return True
    return False


def is_connector_row(function_text: str, part_number: str, part_name: str, unit: str) -> bool:
    if detect_end_group(function_text):
        return True
    haystack = " ".join(filter(None, [function_text, part_number, part_name])).upper()
    if unit == "SET":
        return True
    return any(keyword.upper() in haystack for keyword in CONNECTOR_KEYWORDS)


def classify_row(function_text: str, part_number: str, part_name: str, unit: str) -> str:
    if is_wire_row(part_number, part_name, unit):
        return "wire"
    if is_connector_row(function_text, part_number, part_name, unit):
        return "connector"
    return "material"


def detect_material_group(part_number: str, part_name: str) -> str:
    part_no = collapse_text(part_number).upper()
    part_name_text = collapse_text(part_name)
    haystack = " ".join(filter(None, [part_no, part_name_text]))

    if any(keyword in haystack for keyword in SYNC_BRACKET_KEYWORDS) or re.search(r"(?:^|[-_/])(HB|ZJ)(?:[-_/]|$)", part_no):
        return "sync_brackets"
    if any(keyword in haystack for keyword in SYNC_RUBBER_KEYWORDS) or re.search(r"(?:^|[-_/])XJ(?:[-_/]|$)", part_no):
        return "sync_rubber"
    return "materials"


def item_key_for_row(part_number: str, part_name: str) -> str:
    return normalize_key(part_number) or normalize_key(part_name)


def item_display_key(item: dict[str, Any]) -> str:
    return item.get("partNumber") or item.get("partName") or item.get("itemKey") or "UNKNOWN"


def cleaned_part_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", collapse_text(value).upper())


def part_code_chunks(value: Any) -> set[str]:
    text = collapse_text(value).upper()
    raw_tokens = re.findall(r"[A-Z]+\d+[A-Z\d]*|\d+[A-Z]+[A-Z\d]*|[A-Z]{2,}|\d{2,}", text)
    chunks: set[str] = set()
    for token in raw_tokens:
        chunks.add(token)
        if len(token) >= 6:
            chunks.add(token[:6])
        if len(token) >= 8:
            chunks.add(token[:8])
    return chunks


def shared_prefix_length(left: str, right: str) -> int:
    prefix = 0
    for left_char, right_char in zip(left, right):
        if left_char != right_char:
            break
        prefix += 1
    return prefix


def item_keywords(item: dict[str, Any]) -> set[str]:
    text_parts = [
        collapse_text(item.get("partName")),
        " ".join(item.get("functions") or []),
        " ".join(item.get("remarks") or []),
    ]
    haystack = " ".join(part for part in text_parts if part).upper()
    return {keyword.upper() for keyword in ASSEMBLY_MAPPING_KEYWORDS if keyword.upper() in haystack}


def is_connector_assembly_item(item: dict[str, Any]) -> bool:
    unit = collapse_text(item.get("unit")).upper()
    text = " ".join(
        part
        for part in (
            collapse_text(item.get("partNumber")),
            collapse_text(item.get("partName")),
            " ".join(item.get("functions") or []),
        )
        if part
    ).upper()
    if unit == "SET":
        return True
    return any(keyword in text for keyword in ("总成", "连接器", "插座", "插件"))


def assembly_mapping_score(assembly_item: dict[str, Any], fixed_item: dict[str, Any]) -> int:
    assembly_code = cleaned_part_code(assembly_item.get("partNumber"))
    fixed_code = cleaned_part_code(fixed_item.get("partNumber"))
    score = 0

    prefix_len = shared_prefix_length(assembly_code, fixed_code)
    if prefix_len >= 10:
        score += 100
    elif prefix_len >= 8:
        score += 75
    elif prefix_len >= 6:
        score += 45
    elif prefix_len >= 4:
        score += 20

    if assembly_code and fixed_code and (assembly_code in fixed_code or fixed_code in assembly_code):
        score += 60

    shared_chunks = part_code_chunks(assembly_item.get("partNumber")) & part_code_chunks(fixed_item.get("partNumber"))
    score += min(len(shared_chunks), 4) * 12
    score += len(item_keywords(assembly_item) & item_keywords(fixed_item)) * 10

    if collapse_text(assembly_item.get("partName")) == "连接器总成" and score < 30:
        return 0
    return score


def assign_connector_assembly_parts(
    assembly_items: list[dict[str, Any]],
    target_items: list[dict[str, Any]],
) -> tuple[dict[str, list[dict[str, Any]]], set[str]]:
    if not assembly_items or not target_items:
        return {}, set()

    assignments: list[int | None] = []
    if len(assembly_items) == 1:
        assignments = [0 for _ in target_items]
    else:
        for target_item in target_items:
            scores = [assembly_mapping_score(assembly_item, target_item) for assembly_item in assembly_items]
            best_score = max(scores) if scores else 0
            if best_score >= 15 and scores.count(best_score) == 1:
                assignments.append(scores.index(best_score))
            else:
                assignments.append(None)

        cursor = 0
        while cursor < len(assignments):
            if assignments[cursor] is not None:
                cursor += 1
                continue

            gap_start = cursor
            while cursor < len(assignments) and assignments[cursor] is None:
                cursor += 1
            gap_end = cursor - 1

            prev_assignment = assignments[gap_start - 1] if gap_start > 0 else None
            next_assignment = assignments[cursor] if cursor < len(assignments) else None
            fill_value = None
            if prev_assignment is not None and prev_assignment == next_assignment:
                fill_value = prev_assignment
            elif prev_assignment is not None and next_assignment is None:
                fill_value = prev_assignment
            elif prev_assignment is None and next_assignment is not None:
                fill_value = next_assignment

            if fill_value is not None:
                for index in range(gap_start, gap_end + 1):
                    assignments[index] = fill_value

    mapped: dict[str, list[dict[str, Any]]] = {}
    used_target_keys: set[str] = set()
    for target_item, assignment in zip(target_items, assignments):
        if assignment is None:
            continue
        assembly_item = assembly_items[assignment]
        mapped.setdefault(assembly_item["itemKey"], []).append(target_item)
        used_target_keys.add(target_item["itemKey"])

    return mapped, used_target_keys


def init_version_dict(default: Any) -> dict[str, Any]:
    return {source: default() if callable(default) else default for source in source_keys()}


def row_item(row: dict[str, Any], source: str) -> dict[str, Any] | None:
    return (row.get("versions") or {}).get(source)


def row_parts(row: dict[str, Any], source: str) -> list[dict[str, Any]]:
    return list(((row.get("partLists") or {}).get(source) or []))


def row_source_count(row: dict[str, Any]) -> int:
    return sum(1 for source in source_keys() if row_presence_for_source(row, source))


def row_presence_for_source(row: dict[str, Any], source: str) -> bool:
    return bool(row_item(row, source) or row_parts(row, source))


def row_match_state(row: dict[str, Any]) -> str:
    source_count = row_source_count(row)
    if row.get("rowType") == "assembly_bundle":
        return "assembly_bundle"
    if source_count >= 3:
        return "full_match"
    if source_count == 2:
        return "partial_match"
    for source in source_keys():
        if row_presence_for_source(row, source):
            return f"{source}_only"
    return "empty"


def build_initial_rows(group_items: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    rows: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    for source in source_keys():
        items = group_items.get(source, [])
        for item in items:
            row = rows.setdefault(
                item["itemKey"],
                {
                    "itemKey": item["itemKey"],
                    "rowType": "standard",
                    "versions": init_version_dict(None),
                    "partLists": {version: [] for version in source_keys() if version != base_source()},
                },
            )
            row["versions"][source] = item
    return list(rows.values())


def merge_connector_assembly_rows(aligned: list[dict[str, Any]]) -> list[dict[str, Any]]:
    assembly_rows = [
        row
        for row in aligned
        if row_item(row, base_source())
        and is_connector_assembly_item(row_item(row, base_source()))
        and any(not row_item(row, source) for source in source_keys() if source != base_source())
    ]
    parts_map_by_source: dict[str, dict[str, list[dict[str, Any]]]] = {}
    used_keys_by_source: dict[str, set[str]] = {}
    for source in source_keys():
        if source == base_source():
            continue
        target_items = [row_item(row, source) for row in aligned if row_item(row, source) and not row_item(row, base_source())]
        candidate_items = [row_item(row, base_source()) for row in assembly_rows if not row_item(row, source)]
        mapped_parts, used_keys = assign_connector_assembly_parts(candidate_items, target_items)
        parts_map_by_source[source] = mapped_parts
        used_keys_by_source[source] = used_keys

    if not any(parts_map_by_source.values()):
        for row in aligned:
            row["matchState"] = row_match_state(row)
            row["sourceCount"] = row_source_count(row)
        return aligned

    merged: list[dict[str, Any]] = []
    for row in aligned:
        base_item = row_item(row, base_source())
        mapped_part_lists = {
            source: parts_map_by_source[source].get(base_item["itemKey"], []) if base_item else []
            for source in source_keys()
            if source != base_source()
        }

        if base_item and any(mapped_part_lists.values()):
            version_items = init_version_dict(None)
            for source in source_keys():
                version_items[source] = row_item(row, source)
            assembly_row = {
                "itemKey": row["itemKey"],
                "rowType": "assembly_bundle",
                "versions": version_items,
                "partLists": {
                    source: list(mapped_part_lists.get(source, []))
                    for source in source_keys()
                    if source != base_source()
                },
            }
            assembly_row["matchState"] = row_match_state(assembly_row)
            assembly_row["sourceCount"] = row_source_count(assembly_row)
            merged.append(assembly_row)
            continue

        next_versions = init_version_dict(None)
        for source in source_keys():
            item = row_item(row, source)
            if source != base_source() and item and item["itemKey"] in used_keys_by_source[source] and not base_item:
                next_versions[source] = None
            else:
                next_versions[source] = item
        if not any(next_versions.values()):
            continue

        next_row = {
            "itemKey": row["itemKey"],
            "rowType": "standard",
            "versions": next_versions,
            "partLists": {source: [] for source in source_keys() if source != base_source()},
        }
        next_row["matchState"] = row_match_state(next_row)
        next_row["sourceCount"] = row_source_count(next_row)
        merged.append(next_row)

    return merged


def summarize_alignment(aligned: list[dict[str, Any]]) -> dict[str, Any]:
    stats = {
        "matched": 0,
        "fullMatch": 0,
        "partialMatch": 0,
        "onlyCounts": {source: 0 for source in source_keys()},
        "assemblyToParts": 0,
        "assemblyPartCount": 0,
    }
    for row in aligned:
        status = row.get("matchState") or row_match_state(row)
        if status == "full_match":
            stats["matched"] += 1
            stats["fullMatch"] += 1
        elif status == "partial_match":
            stats["matched"] += 1
            stats["partialMatch"] += 1
        elif status == "assembly_bundle":
            stats["matched"] += 1
            if row_source_count(row) >= 3:
                stats["fullMatch"] += 1
            elif row_source_count(row) == 2:
                stats["partialMatch"] += 1
            stats["assemblyToParts"] += 1
            stats["assemblyPartCount"] += sum(len(row_parts(row, source)) for source in source_keys() if source != base_source())
        elif status.endswith("_only"):
            only_source = status[:-5]
            if only_source in stats["onlyCounts"]:
                stats["onlyCounts"][only_source] += 1
    return stats


def parse_sheet_rows(sheet_name: str, source: str, rows: list[tuple[Any, ...]]) -> ParsedSheet:
    harness_id = first_digits(sheet_name) or sheet_name
    harness_name = collapse_text(rows[1][4]) if len(rows) > 1 else harness_id
    items: list[dict[str, Any]] = []
    current_connector_group: str | None = None
    blank_streak = 0

    for row_index, row in enumerate(rows[4:], start=5):
        values = list(row[: len(HEADER_KEYS)])
        if len(values) < len(HEADER_KEYS):
            values.extend([None] * (len(HEADER_KEYS) - len(values)))

        if not any(value not in (None, "") for value in values):
            blank_streak += 1
            if items and blank_streak >= 20:
                break
            continue
        blank_streak = 0

        row_map = {HEADER_KEYS[idx]: values[idx] for idx in range(len(HEADER_KEYS))}
        function_text = normalize_text(row_map["function"])
        part_number = collapse_text(row_map["partNumber"])
        part_name = collapse_text(row_map["partName"])
        unit = collapse_text(row_map["unit"]).upper()

        if not any([function_text, part_number, part_name]):
            continue

        kind = classify_row(function_text, part_number, part_name, unit)
        detected_group = detect_end_group(function_text)
        if kind == "connector":
            if detected_group:
                group_key = detected_group
                current_connector_group = detected_group
            elif not function_text and current_connector_group:
                group_key = current_connector_group
            else:
                group_key = "connector_misc"
        else:
            group_key = "wires" if kind == "wire" else detect_material_group(part_number, part_name)

        item = {
            "rowNumber": row_index,
            "sequence": numeric_value(row_map["no"]),
            "kind": kind,
            "groupKey": group_key,
            "groupLabel": GROUP_LABELS[group_key],
            "functionRaw": function_text,
            "functionBrief": collapse_text(function_text.split("\n")[1] if "\n" in function_text and len(function_text.split("\n")) > 1 else function_text),
            "partNumber": part_number,
            "partName": part_name,
            "quantity": numeric_value(row_map["quantity"]),
            "unit": unit,
            "remark": collapse_text(row_map["remark"]),
            "otherRemark": collapse_text(row_map["otherRemark"]),
            "wireNo": collapse_text(row_map["wireNo"]),
            "spec": collapse_text(row_map["spec"]),
            "itemKey": item_key_for_row(part_number, part_name),
        }
        items.append(item)

    return ParsedSheet(
        harness_id=harness_id,
        harness_name=harness_name,
        sheet_name=sheet_name,
        source=source,
        items=items,
    )


def read_workbook(path: Path, source: str) -> dict[str, ParsedSheet]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, ParsedSheet] = {}
    for sheet_name in workbook.sheetnames:
        harness_id = first_digits(sheet_name)
        if not harness_id:
            continue
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        parsed = parse_sheet_rows(sheet_name, source, rows)
        sheets[harness_id] = parsed
    return sheets


def aggregate_items(items: list[dict[str, Any]], ksk_lookup: dict[str, list[dict[str, Any]]] | None = None) -> list[dict[str, Any]]:
    ordered: "OrderedDict[str, dict[str, Any]]" = OrderedDict()
    for item in items:
        key = item["itemKey"] or f"{item['groupKey']}::{item['rowNumber']}"
        if key not in ordered:
            ordered[key] = {
                "itemKey": key,
                "partNumber": item["partNumber"],
                "partName": item["partName"],
                "unit": item["unit"],
                "quantity": 0,
                "rowNumbers": [],
                "functions": [],
                "remarks": [],
                "otherRemarks": [],
                "wireNos": [],
                "suppliers": [],
                "sapNos": [],
                "assemblyRefs": [],
                "kind": item["kind"],
                "groupKey": item["groupKey"],
            }

        bucket = ordered[key]
        bucket["rowNumbers"].append(item["rowNumber"])
        if item["quantity"] is not None:
            bucket["quantity"] = round(float(bucket["quantity"]) + float(item["quantity"]), 6)
        for source_key, target_key in (
            ("functionBrief", "functions"),
            ("remark", "remarks"),
            ("otherRemark", "otherRemarks"),
            ("wireNo", "wireNos"),
        ):
            value = item[source_key]
            if value and value not in bucket[target_key]:
                bucket[target_key].append(value)

        master_matches = (ksk_lookup or {}).get(key, [])
        for match in master_matches:
            for source_key, target_key in (
                ("supplier", "suppliers"),
                ("sapNo", "sapNos"),
                ("assemblyRef", "assemblyRefs"),
                ("otherRemark", "otherRemarks"),
            ):
                value = match[source_key]
                if value and value not in bucket[target_key]:
                    bucket[target_key].append(value)

        if item["remark"] and item["remark"] not in bucket["suppliers"]:
            bucket["suppliers"].append(item["remark"])

    aggregated = list(ordered.values())
    for item in aggregated:
        quantity = item["quantity"]
        if abs(quantity - round(quantity)) < 1e-9:
            item["quantity"] = int(round(quantity))
        else:
            item["quantity"] = round(quantity, 4)
    return aggregated


def align_group_items(
    group_items: dict[str, list[dict[str, Any]]],
    section: str = "material",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    aligned = build_initial_rows(group_items)
    if section == "connector":
        aligned = merge_connector_assembly_rows(aligned)
    else:
        for row in aligned:
            row["matchState"] = row_match_state(row)
            row["sourceCount"] = row_source_count(row)
    return aligned, summarize_alignment(aligned)


def build_group_view(
    group_key: str,
    group_items: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    section = GROUP_SECTIONS.get(group_key, "connector")
    aligned, stats = align_group_items(group_items, section)
    return {
        "key": group_key,
        "label": GROUP_LABELS[group_key],
        "section": section,
        "itemCounts": {source: len(group_items.get(source, [])) for source in source_keys()},
        "matchedCount": stats["matched"],
        "fullMatchCount": stats["fullMatch"],
        "partialMatchCount": stats["partialMatch"],
        "onlyCounts": stats["onlyCounts"],
        "assemblyToPartsCount": stats["assemblyToParts"],
        "assemblyPartCount": stats["assemblyPartCount"],
        "aligned": aligned,
    }


def comparison_summary(groups: list[dict[str, Any]]) -> dict[str, Any]:
    connector_groups = [group for group in groups if group["section"] == "connector"]
    wire_group = next((group for group in groups if group["section"] == "wire"), None)
    sync_groups = [group for group in groups if group["section"] == "sync"]
    material_groups = [group for group in groups if group["section"] in ("material", "sync")]
    return {
        "groupCount": len(groups),
        "connectorGroupCount": len(connector_groups),
        "syncGroupCount": len(sync_groups),
        "itemCounts": {
            source: sum(group["itemCounts"][source] for group in groups)
            for source in source_keys()
        },
        "matchedCount": sum(group["matchedCount"] for group in groups),
        "fullMatchCount": sum(group["fullMatchCount"] for group in groups),
        "partialMatchCount": sum(group["partialMatchCount"] for group in groups),
        "onlyCounts": {
            source: sum(group["onlyCounts"][source] for group in groups)
            for source in source_keys()
        },
        "assemblyToPartsCount": sum(group.get("assemblyToPartsCount", 0) for group in groups),
        "assemblyPartCount": sum(group.get("assemblyPartCount", 0) for group in groups),
        "wireMatchedCount": wire_group["matchedCount"] if wire_group else 0,
        "syncMatchedCount": sum(group["matchedCount"] for group in sync_groups),
        "materialMatchedCount": sum(group["matchedCount"] for group in material_groups),
    }


def build_comparison(
    sheets_by_source: dict[str, ParsedSheet],
    lookups_by_source: dict[str, dict[str, list[dict[str, Any]]]],
) -> dict[str, Any]:
    group_maps: dict[str, dict[str, list[dict[str, Any]]]] = {source: {} for source in source_keys()}
    for source in source_keys():
        sheet = sheets_by_source[source]
        lookup = lookups_by_source.get(source, {})
        for item in aggregate_items(sheet.items, lookup):
            group_maps[source].setdefault(item["groupKey"], []).append(item)

    present_keys = [
        key
        for key in GROUP_ORDER
        if any(key in group_maps[source] for source in source_keys())
    ]
    groups = [
        build_group_view(group_key, {source: group_maps[source].get(group_key, []) for source in source_keys()})
        for group_key in present_keys
    ]

    return {
        "harnessId": sheets_by_source[base_source()].harness_id,
        "harnessName": next((sheet.harness_name for sheet in sheets_by_source.values() if sheet.harness_name), sheets_by_source[base_source()].harness_id),
        "versionOrder": list(display_source_keys()),
        "compareOrder": list(source_keys()),
        "baseSource": base_source(),
        "sources": {
            source: {
                "label": version_label(source),
                "sheet": sheets_by_source[source].sheet_name,
                "itemCount": len(sheets_by_source[source].items),
            }
            for source in source_keys()
        },
        "summary": comparison_summary(groups),
        "groups": groups,
    }


def build_output(paths_by_source: dict[str, Path]) -> dict[str, Any]:
    sheets_by_source = {source: read_workbook(path, source) for source, path in paths_by_source.items()}
    lookups_by_source = {source: read_ksk_lookup(path) for source, path in paths_by_source.items()}

    harness_ids = sorted(set().union(*(set(sheets.keys()) for sheets in sheets_by_source.values())))
    comparisons: dict[str, Any] = {}
    unmatched: list[dict[str, Any]] = []

    for harness_id in harness_ids:
        harness_sheets = {source: sheets_by_source[source].get(harness_id) for source in source_keys()}
        if all(harness_sheets.values()):
            comparisons[harness_id] = build_comparison(
                {source: harness_sheets[source] for source in source_keys() if harness_sheets[source]},
                {
                    source: lookups_by_source[source].get(harness_id, {})
                    for source in source_keys()
                },
            )
        else:
            unmatched.append(
                {
                    "harnessId": harness_id,
                    "sheets": {
                        source: harness_sheets[source].sheet_name if harness_sheets[source] else ""
                        for source in source_keys()
                    },
                }
            )

    payload = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "generator": "g281_generate_bom_validation.py",
            "version": "0.5.1",
            "comparisonMethod": "per-harness -> multi-version lanes -> connector end groups / wires / sync-dev parts / materials",
            "versionOrder": list(display_source_keys()),
            "compareOrder": list(source_keys()),
            "baseSource": base_source(),
            "versionLabels": {source: version_label(source) for source in source_keys()},
            "workbooks": {source: paths_by_source[source].name for source in source_keys()},
            "harnessCount": len(comparisons),
            "unmatchedHarnessCount": len(unmatched),
        },
        "harnessOrder": sorted(comparisons.keys()),
        "comparisons": comparisons,
        "unmatchedHarnesses": unmatched,
    }

    for source in ("quote", "fixed", "tt"):
        if source in paths_by_source:
            payload["meta"][f"{source}Workbook"] = paths_by_source[source].name
    return payload


def parse_extra_version(spec_text: str) -> tuple[str, dict[str, str], Path]:
    parts = [part.strip() for part in str(spec_text or "").split("|")]
    if len(parts) != 3 or not all(parts):
        raise argparse.ArgumentTypeError("Extra version must use format key|label|path")
    key, label, path_text = parts
    if key in DEFAULT_VERSION_SPECS:
        raise argparse.ArgumentTypeError(f"Extra version key '{key}' conflicts with built-in versions.")
    return key, {"label": label, "pattern": path_text}, Path(path_text)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate aligned BOM validation JSON for G281 multi-version BOM workbooks.")
    parser.add_argument("--quote", type=Path, default=None, help="Quote BOM workbook path.")
    parser.add_argument("--fixed", type=Path, default=None, help="Fixed BOM workbook path.")
    parser.add_argument("--tt", type=Path, default=None, help="TT BOM workbook path.")
    parser.add_argument(
        "--extra-version",
        action="append",
        default=[],
        metavar="KEY|LABEL|PATH",
        help="Append an extra BOM version in order, for example review|评审 BOM|C:\\path\\review.xlsx",
    )
    parser.add_argument("--out", type=Path, default=Path("g281_data_bom_validation.json"), help="Output JSON path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    ACTIVE_VERSION_SPECS.clear()
    ACTIVE_VERSION_SPECS.update(DEFAULT_VERSION_SPECS)
    paths_by_source = {
        "quote": args.quote or discover_workbook(version_pattern("quote")),
        "fixed": args.fixed or discover_workbook(version_pattern("fixed")),
        "tt": args.tt or discover_workbook(version_pattern("tt")),
    }
    for raw_spec in args.extra_version:
        key, spec, path = parse_extra_version(raw_spec)
        ACTIVE_VERSION_SPECS[key] = spec
        paths_by_source[key] = path
    payload = build_output(paths_by_source)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {args.out}")
    print(f"Harnesses: {payload['meta']['harnessCount']}")
    print(f"Unmatched harnesses: {payload['meta']['unmatchedHarnessCount']}")


if __name__ == "__main__":
    main()
