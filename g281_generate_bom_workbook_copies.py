from __future__ import annotations

import argparse
import json
import math
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.numbers import BUILTIN_FORMATS


CONFIG_SHEET_NAME = "配置清单"
DEFAULT_SOURCE_DIR = Path("BOM核对")
DEFAULT_OUTPUT_PATH = Path("g281_data_bom_workbook_copies.json")
VERSION_SPECS = [
    {
        "key": "quote",
        "label": "报价版",
        "matchTokens": ["BOM", "V01-11.3"],
    },
    {
        "key": "fixed",
        "label": "定点版",
        "matchTokens": ["BOM", "V05-2026.01.04"],
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Copy BOM workbook sheets (excluding 配置清单) to JSON, preserving content and layout "
            "for later Excel-like rendering in frontend."
        )
    )
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=DEFAULT_SOURCE_DIR,
        help="Directory containing BOM workbooks.",
    )
    parser.add_argument(
        "--quote-workbook",
        type=Path,
        default=None,
        help="Optional explicit quote BOM workbook path.",
    )
    parser.add_argument(
        "--fixed-workbook",
        type=Path,
        default=None,
        help="Optional explicit fixed BOM workbook path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Output JSON path.",
    )
    return parser.parse_args()


def json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isfinite(value):
            return value
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def clean_descriptor_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("Values must be of type"):
        return None
    return value


def discover_workbook(source_dir: Path, tokens: list[str]) -> Path:
    files = sorted(path for path in source_dir.glob("*.xlsx") if not path.name.startswith("~$"))
    for path in files:
        if all(token in path.name for token in tokens):
            return path
    raise FileNotFoundError(
        f"Unable to discover workbook under {source_dir} using tokens={tokens}. "
        f"Candidates={[path.name for path in files]}"
    )


def serialize_color(color: Any) -> dict[str, Any] | None:
    if color is None:
        return None
    color_type = clean_descriptor_value(color.type) if isinstance(color.type, str) else None
    rgb = clean_descriptor_value(color.rgb) if isinstance(color.rgb, str) else None
    indexed = clean_descriptor_value(color.indexed) if isinstance(color.indexed, int) else None
    theme = clean_descriptor_value(color.theme) if isinstance(color.theme, int) else None
    auto = clean_descriptor_value(color.auto) if isinstance(color.auto, bool) else None
    tint = clean_descriptor_value(color.tint) if isinstance(color.tint, (int, float)) else None
    return {
        "type": color_type,
        "rgb": rgb,
        "indexed": indexed,
        "theme": theme,
        "tint": json_scalar(tint),
        "auto": auto,
    }


def serialize_side(side: Any) -> dict[str, Any]:
    if side is None:
        return {"style": None, "color": None}
    return {
        "style": side.style,
        "color": serialize_color(side.color),
    }


def serialize_font(font: Any) -> dict[str, Any]:
    return {
        "name": font.name,
        "size": json_scalar(font.sz),
        "bold": font.b,
        "italic": font.i,
        "underline": font.u,
        "strike": font.strike,
        "color": serialize_color(font.color),
        "family": font.family,
        "charset": font.charset,
        "scheme": font.scheme,
        "vertAlign": font.vertAlign,
    }


def serialize_fill(fill: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "className": fill.__class__.__name__,
        "fillType": getattr(fill, "fill_type", None),
    }
    if hasattr(fill, "patternType"):
        payload["patternType"] = fill.patternType
    if hasattr(fill, "fgColor"):
        payload["fgColor"] = serialize_color(fill.fgColor)
    if hasattr(fill, "bgColor"):
        payload["bgColor"] = serialize_color(fill.bgColor)
    if hasattr(fill, "degree"):
        payload["degree"] = json_scalar(fill.degree)
    if hasattr(fill, "stop"):
        payload["stops"] = [
            {"position": json_scalar(stop.position), "color": serialize_color(stop.color)}
            for stop in list(fill.stop or [])
        ]
    return payload


def serialize_border(border: Any) -> dict[str, Any]:
    return {
        "left": serialize_side(border.left),
        "right": serialize_side(border.right),
        "top": serialize_side(border.top),
        "bottom": serialize_side(border.bottom),
        "diagonal": serialize_side(border.diagonal),
        "diagonalUp": border.diagonalUp,
        "diagonalDown": border.diagonalDown,
        "outline": border.outline,
        "vertical": serialize_side(border.vertical),
        "horizontal": serialize_side(border.horizontal),
    }


def serialize_alignment(alignment: Any) -> dict[str, Any]:
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "textRotation": alignment.textRotation,
        "wrapText": alignment.wrapText,
        "shrinkToFit": alignment.shrinkToFit,
        "indent": alignment.indent,
        "relativeIndent": alignment.relativeIndent,
        "justifyLastLine": alignment.justifyLastLine,
        "readingOrder": alignment.readingOrder,
    }


def serialize_protection(protection: Any) -> dict[str, Any]:
    return {
        "locked": protection.locked,
        "hidden": protection.hidden,
    }


def resolve_number_format(workbook: Any, num_fmt_id: int | None) -> str | None:
    if num_fmt_id is None:
        return None
    if num_fmt_id in BUILTIN_FORMATS:
        return BUILTIN_FORMATS[num_fmt_id]
    if num_fmt_id >= 164:
        custom_index = num_fmt_id - 164
        if 0 <= custom_index < len(workbook._number_formats):  # pylint: disable=protected-access
            return workbook._number_formats[custom_index]  # pylint: disable=protected-access
    return None


def serialize_style_from_id(workbook: Any, style_id: int) -> dict[str, Any]:
    style_array = workbook._cell_styles[style_id]  # pylint: disable=protected-access
    font = workbook._fonts[style_array.fontId]  # pylint: disable=protected-access
    fill = workbook._fills[style_array.fillId]  # pylint: disable=protected-access
    border = workbook._borders[style_array.borderId]  # pylint: disable=protected-access
    alignment = workbook._alignments[style_array.alignmentId]  # pylint: disable=protected-access
    protection = workbook._protections[style_array.protectionId]  # pylint: disable=protected-access

    return {
        "styleId": style_id,
        "numberFormatId": style_array.numFmtId,
        "numberFormat": resolve_number_format(workbook, style_array.numFmtId),
        "font": serialize_font(font),
        "fill": serialize_fill(fill),
        "border": serialize_border(border),
        "alignment": serialize_alignment(alignment),
        "protection": serialize_protection(protection),
        "quotePrefix": bool(style_array.quotePrefix),
        "pivotButton": bool(style_array.pivotButton),
    }


def should_include_cell(cell: Any) -> bool:
    if isinstance(cell, MergedCell):
        return False
    if cell.value is not None:
        return True
    if cell.comment is not None or cell.hyperlink is not None:
        return True
    if cell.has_style and int(cell.style_id or 0) != 0:
        return True
    return False


def serialize_cell(cell: Any, display_cell: Any | None = None) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "address": cell.coordinate,
        "row": cell.row,
        "column": cell.column,
        "dataType": cell.data_type,
        "styleId": int(cell.style_id or 0),
    }
    if cell.data_type == "f":
        formula_text = str(cell.value or "")
        payload["formula"] = formula_text if formula_text.startswith("=") else f"={formula_text}"
        if display_cell is not None:
            payload["displayValue"] = json_scalar(display_cell.value)
    else:
        payload["value"] = json_scalar(cell.value)

    if cell.number_format not in (None, ""):
        payload["numberFormat"] = cell.number_format
    if cell.hyperlink is not None:
        payload["hyperlink"] = {
            "target": cell.hyperlink.target,
            "location": cell.hyperlink.location,
            "tooltip": cell.hyperlink.tooltip,
            "display": cell.hyperlink.display,
        }
    if cell.comment is not None:
        payload["comment"] = {
            "author": cell.comment.author,
            "text": cell.comment.text,
        }
    return payload


def serialize_row_dimension(row_index: int, dimension: Any) -> dict[str, Any]:
    return {
        "row": int(row_index),
        "height": json_scalar(dimension.height),
        "hidden": bool(dimension.hidden),
        "outlineLevel": dimension.outlineLevel,
        "collapsed": bool(dimension.collapsed),
        "styleId": int(dimension.style_id or 0),
        "customHeight": bool(dimension.customHeight),
        "customFormat": bool(dimension.customFormat),
    }


def serialize_col_dimension(key: str, dimension: Any) -> dict[str, Any]:
    outline_level = getattr(dimension, "outline_level", None)
    if outline_level is None:
        outline_level = getattr(dimension, "outlineLevel", None)
    return {
        "key": key,
        "min": dimension.min,
        "max": dimension.max,
        "width": json_scalar(dimension.width),
        "hidden": bool(dimension.hidden),
        "outlineLevel": outline_level,
        "collapsed": bool(dimension.collapsed),
        "bestFit": bool(dimension.bestFit),
        "styleId": int(dimension.style_id or 0),
        "customWidth": bool(dimension.customWidth),
    }


def serialize_sheet(
    sheet: Any,
    workbook_sheet_index: int,
    style_table: dict[str, dict[str, Any]],
    workbook: Any,
    display_sheet: Any | None = None,
) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    for cell in sorted(sheet._cells.values(), key=lambda item: (item.row, item.column)):  # pylint: disable=protected-access
        if not should_include_cell(cell):
            continue
        style_id = str(int(cell.style_id or 0))
        if style_id not in style_table:
            style_table[style_id] = serialize_style_from_id(workbook, int(style_id))
        display_cell = display_sheet[cell.coordinate] if display_sheet is not None else None
        cells.append(serialize_cell(cell, display_cell))

    row_dimensions = [
        serialize_row_dimension(row_index, dimension)
        for row_index, dimension in sorted(sheet.row_dimensions.items(), key=lambda item: item[0])
    ]
    col_dimensions = [
        serialize_col_dimension(key, dimension)
        for key, dimension in sorted(sheet.column_dimensions.items(), key=lambda item: item[0])
    ]
    for entry in row_dimensions:
        style_id = str(int(entry.get("styleId", 0)))
        if style_id not in style_table:
            style_table[style_id] = serialize_style_from_id(workbook, int(style_id))
    for entry in col_dimensions:
        style_id = str(int(entry.get("styleId", 0)))
        if style_id not in style_table:
            style_table[style_id] = serialize_style_from_id(workbook, int(style_id))

    freeze_pane = sheet.freeze_panes
    if freeze_pane is not None:
        freeze_pane = freeze_pane.coordinate if hasattr(freeze_pane, "coordinate") else str(freeze_pane)

    sheet_view = sheet.sheet_view
    return {
        "workbookSheetIndex": workbook_sheet_index,
        "sheetName": sheet.title,
        "sheetState": sheet.sheet_state,
        "isHidden": sheet.sheet_state in ("hidden", "veryHidden"),
        "sheetOrderKey": workbook_sheet_index,
        "dimensionRef": sheet.calculate_dimension(),
        "maxRow": sheet.max_row,
        "maxColumn": sheet.max_column,
        "sheetFormat": {
            "defaultRowHeight": json_scalar(sheet.sheet_format.defaultRowHeight),
            "defaultColWidth": json_scalar(sheet.sheet_format.defaultColWidth),
            "baseColWidth": json_scalar(sheet.sheet_format.baseColWidth),
            "outlineLevelRow": sheet.sheet_format.outlineLevelRow,
            "outlineLevelCol": sheet.sheet_format.outlineLevelCol,
        },
        "sheetView": {
            "showGridLines": sheet_view.showGridLines,
            "showRowColHeaders": sheet_view.showRowColHeaders,
            "zoomScale": sheet_view.zoomScale,
            "topLeftCell": sheet_view.topLeftCell,
        },
        "freezePane": freeze_pane,
        "mergedRanges": [str(cell_range) for cell_range in sheet.merged_cells.ranges],
        "rowDimensions": row_dimensions,
        "columnDimensions": col_dimensions,
        "hiddenRows": [entry["row"] for entry in row_dimensions if entry["hidden"]],
        "hiddenColumns": [entry["key"] for entry in col_dimensions if entry["hidden"]],
        "cells": cells,
    }


def serialize_workbook(version_key: str, label: str, workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    display_workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    style_table: dict[str, dict[str, Any]] = {"0": serialize_style_from_id(workbook, 0)}
    sheets = []
    excluded_sheets = []

    for workbook_sheet_index, sheet in enumerate(workbook.worksheets):
        if sheet.title.strip() == CONFIG_SHEET_NAME:
            excluded_sheets.append(sheet.title)
            continue
        display_sheet = display_workbook[sheet.title] if sheet.title in display_workbook.sheetnames else None
        sheets.append(serialize_sheet(sheet, workbook_sheet_index, style_table, workbook, display_sheet))

    workbook.close()
    display_workbook.close()

    return {
        "versionKey": version_key,
        "versionLabel": label,
        "sourceFileName": workbook_path.name,
        "sourcePath": str(workbook_path.as_posix()),
        "excludedSheets": excluded_sheets,
        "sheetOrder": [sheet["sheetName"] for sheet in sheets],
        "hiddenSheets": [sheet["sheetName"] for sheet in sheets if sheet["isHidden"]],
        "styleTable": style_table,
        "sheets": sheets,
    }


def contract_shape() -> dict[str, Any]:
    return {
        "name": "g281.bomWorkbookCopies",
        "version": "1.0.0",
        "goal": (
            "Store BOM workbook copies as versioned JSON payloads that can be restored in frontend "
            "as Excel-like full-sheet views."
        ),
        "exclusionRule": f"Exclude sheet title exactly equal to '{CONFIG_SHEET_NAME}'.",
        "requiredFrontEndFields": [
            "sheetOrder",
            "freezePane",
            "mergedRanges",
            "rowDimensions",
            "columnDimensions",
            "cells.value",
            "cells.formula",
            "styleTable",
        ],
        "topLevelShape": {
            "contract": "object",
            "generatedAt": "ISO datetime string",
            "generator": "string",
            "versionOrder": "string[]",
            "sourceDirectory": "string",
            "versions.quote": "WorkbookPayload",
            "versions.fixed": "WorkbookPayload",
        },
        "workbookPayloadShape": {
            "versionKey": "quote|fixed",
            "versionLabel": "string",
            "sourceFileName": "string",
            "sourcePath": "string",
            "excludedSheets": "string[]",
            "sheetOrder": "string[]",
            "hiddenSheets": "string[]",
            "styleTable": "{styleId: StylePayload}",
            "sheets": "SheetPayload[]",
        },
        "sheetPayloadShape": {
            "workbookSheetIndex": "number",
            "sheetName": "string",
            "sheetState": "visible|hidden|veryHidden",
            "dimensionRef": "string",
            "maxRow": "number",
            "maxColumn": "number",
            "freezePane": "string|null",
            "mergedRanges": "string[]",
            "rowDimensions": "RowDimension[]",
            "columnDimensions": "ColumnDimension[]",
            "cells": "CellPayload[]",
        },
        "cellPayloadShape": {
            "address": "A1 notation",
            "row": "number",
            "column": "number",
            "dataType": "openpyxl cell data type",
            "value": "scalar (when not formula)",
            "formula": "excel formula string (when formula)",
            "styleId": "number",
            "numberFormat": "string?",
            "hyperlink": "object?",
            "comment": "object?",
        },
    }


def resolve_source_paths(args: argparse.Namespace) -> tuple[Path, Path]:
    source_dir = args.source_dir.resolve()
    quote_path = args.quote_workbook.resolve() if args.quote_workbook else None
    fixed_path = args.fixed_workbook.resolve() if args.fixed_workbook else None

    if quote_path is None:
        quote_spec = next(spec for spec in VERSION_SPECS if spec["key"] == "quote")
        quote_path = discover_workbook(source_dir, quote_spec["matchTokens"])
    if fixed_path is None:
        fixed_spec = next(spec for spec in VERSION_SPECS if spec["key"] == "fixed")
        fixed_path = discover_workbook(source_dir, fixed_spec["matchTokens"])
    return quote_path, fixed_path


def main() -> None:
    args = parse_args()
    source_dir = args.source_dir.resolve()
    output_path = args.output.resolve()

    quote_path, fixed_path = resolve_source_paths(args)

    quote_spec = next(spec for spec in VERSION_SPECS if spec["key"] == "quote")
    fixed_spec = next(spec for spec in VERSION_SPECS if spec["key"] == "fixed")

    payload = {
        "contract": contract_shape(),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "generator": "g281_generate_bom_workbook_copies.py",
        "versionOrder": ["quote", "fixed"],
        "sourceDirectory": str(source_dir.as_posix()),
        "versions": {
            "quote": serialize_workbook("quote", quote_spec["label"], quote_path),
            "fixed": serialize_workbook("fixed", fixed_spec["label"], fixed_path),
        },
    }

    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generated {output_path}")


if __name__ == "__main__":
    main()
