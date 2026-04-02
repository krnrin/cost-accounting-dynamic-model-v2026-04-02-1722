from __future__ import annotations

import argparse
import json
import os
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_DIR = Path("BOM\u6838\u5bf9")
ASSESSMENT_SHEET = "\u9879\u76ee\u8bc4\u4f30\u6c47\u603b\uff08\u6606\u5c7190%\uff09"
YEAR_COLUMNS = ["F", "G", "H", "I", "J", "K"]
YEAR_VALUES = [2026, 2027, 2028, 2029, 2030, 2031]
SNAPSHOT_SHEETS = [
    ASSESSMENT_SHEET,
    "\u8fd0\u8425\u5de5\u65f6\u8d39\u62a5\u4ef7\u57fa\u51c6",
    "\u8bbe\u5907\u6295\u8d44\u660e\u7ec6",
    "\u9879\u76ee\u4e13\u7528\u6a21\u5177",
    "\u9879\u76ee\u5de5\u88c5\u6295\u5165 ",
    "\u7814\u53d1\u8d39\u7528 ",
    "\u5305\u88c5\u7269\u6d41\u8d39\u7528",
    "\u914d\u7f6e\u660e\u7ec6",
]

VERSION_SPECS = {
    "quote": {
        "pattern": "\u62a5\u4ef7\u6838\u7b97",
        "label": "\u62a5\u4ef7\u7248",
    },
    "fixed": {
        "pattern": "\u5b9a\u70b9\u6838\u7b97",
        "label": "\u5b9a\u70b9\u7248",
    },
}

TOTAL_CELLS = {
    "volume": "E5",
    "revenue": "E9",
    "profit": "E10",
    "margin": "E11",
    "cost": "E14",
    "material": "E15",
    "directLabor": "E16",
    "equipment": "E20",
    "manufacturing": "E23",
    "rnd": "E31",
    "packaging": "E32",
}

ANNUAL_ROWS = {
    "asp": 6,
    "revenue": 9,
    "cost": 14,
    "material": 15,
    "directLabor": 16,
    "equipment": 20,
    "manufacturing": 23,
    "rnd": 31,
    "packaging": 32,
}


def collapse_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def serialize_cell_value(value: Any) -> float | str | bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    numeric = number_or_none(value)
    if numeric is not None:
        return numeric
    text = collapse_text(value)
    return text or None


def number_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = collapse_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def round_number(value: Any, digits: int = 12) -> float | None:
    numeric = number_or_none(value)
    if numeric is None:
        return None
    return round(numeric, digits)


def discover_workbook(pattern: str) -> Path:
    base_dir = WORKBOOK_DIR if WORKBOOK_DIR.exists() else Path(".")
    matches = [
        base_dir / name
        for name in os.listdir(base_dir)
        if name.endswith(".xlsx") and not name.startswith("~$") and pattern in name
    ]
    if not matches:
        raise FileNotFoundError(f"Workbook containing '{pattern}' not found under {base_dir}.")
    return sorted(matches)[0]


def read_row(ws, row_index: int) -> list[float | None]:
    return [round_number(ws[f"{column}{row_index}"].value) for column in YEAR_COLUMNS]


def clean_series(values: list[float | None]) -> list[float]:
    return [0.0 if value is None else float(value) for value in values]


def formula_text(value: Any) -> str | None:
    if value is None:
        return None
    text = getattr(value, "text", None)
    if isinstance(text, str) and text.strip():
        return text.strip()
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return None


def capture_sheet_snapshot(data_worksheet: Any, formula_worksheet: Any) -> dict[str, Any]:
    cells: list[dict[str, Any]] = []
    for row_index in range(1, formula_worksheet.max_row + 1):
        for column_index in range(1, formula_worksheet.max_column + 1):
            data_cell = data_worksheet.cell(row=row_index, column=column_index)
            formula_cell = formula_worksheet.cell(row=row_index, column=column_index)
            value = serialize_cell_value(data_cell.value)
            formula_value = formula_text(formula_cell.value)
            if value is None and formula_value is None:
                continue
            cells.append(
                {
                    "address": formula_cell.coordinate,
                    "row": row_index,
                    "column": column_index,
                    "dataType": formula_cell.data_type,
                    "value": value,
                    "formula": formula_value,
                }
            )
    return {
        "sheetName": formula_worksheet.title,
        "maxRow": formula_worksheet.max_row,
        "maxColumn": formula_worksheet.max_column,
        "cells": cells,
    }


def capture_assessment_workbook_seed(
    data_workbook: Any, formula_workbook: Any, workbook_path: Path
) -> dict[str, Any]:
    sheet_order = [
        sheet_name
        for sheet_name in SNAPSHOT_SHEETS
        if sheet_name in data_workbook.sheetnames and sheet_name in formula_workbook.sheetnames
    ]
    return {
        "workbookName": workbook_path.name,
        "sourceFileName": workbook_path.name,
        "sourcePath": str(workbook_path.resolve()),
        "sheetOrder": sheet_order,
        "sheets": [
            capture_sheet_snapshot(data_workbook[sheet_name], formula_workbook[sheet_name])
            for sheet_name in sheet_order
        ],
    }


def extract_version_payload(version_key: str, workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    worksheet = workbook[ASSESSMENT_SHEET]
    workbook_formula = load_workbook(workbook_path, data_only=False, read_only=True)

    totals = {key: round_number(worksheet[cell].value) or 0.0 for key, cell in TOTAL_CELLS.items()}
    annual = {key: clean_series(read_row(worksheet, row_index)) for key, row_index in ANNUAL_ROWS.items()}
    annual_volume = clean_series(read_row(worksheet, 5))

    total_volume = totals["volume"] or sum(annual_volume)
    per_set = {
        "revenue": round_number(totals["revenue"] / total_volume) or 0.0,
        "cost": round_number(totals["cost"] / total_volume) or 0.0,
        "profit": round_number(totals["profit"] / total_volume) or 0.0,
        "margin": round_number(totals["margin"]) or 0.0,
        "material": round_number(totals["material"] / total_volume) or 0.0,
        "directLabor": round_number(totals["directLabor"] / total_volume) or 0.0,
        "equipment": round_number(totals["equipment"] / total_volume) or 0.0,
        "manufacturing": round_number(totals["manufacturing"] / total_volume) or 0.0,
        "rnd": round_number(totals["rnd"] / total_volume) or 0.0,
        "packaging": round_number(totals["packaging"] / total_volume) or 0.0,
    }

    annual_cost = [
        round_number((annual["cost"][index] or 0.0) * (annual_volume[index] or 0.0)) or 0.0
        for index in range(len(YEAR_VALUES))
    ]
    annual_profit = [
        round_number((annual["revenue"][index] or 0.0) - annual_cost[index]) or 0.0
        for index in range(len(YEAR_VALUES))
    ]
    annual_margin = [
        round_number(annual_profit[index] / annual["revenue"][index]) if annual["revenue"][index] else 0.0
        for index in range(len(YEAR_VALUES))
    ]

    seed = capture_assessment_workbook_seed(workbook, workbook_formula, workbook_path)
    workbook.close()
    workbook_formula.close()

    return {
        "key": version_key,
        "label": VERSION_SPECS[version_key]["label"],
        "workbook": workbook_path.name,
        "sheetName": ASSESSMENT_SHEET,
        "years": YEAR_VALUES,
        "volumes": annual_volume,
        "asp": annual["asp"],
        "totals": totals,
        "perSet": per_set,
        "annual": {
            "revenue": annual["revenue"],
            "cost": annual_cost,
            "profit": annual_profit,
            "margin": annual_margin,
            "costPerSet": annual["cost"],
            "materialPerSet": annual["material"],
            "directLaborPerSet": annual["directLabor"],
            "equipmentPerSet": annual["equipment"],
            "manufacturingPerSet": annual["manufacturing"],
            "rndPerSet": annual["rnd"],
            "packagingPerSet": annual["packaging"],
        },
        "cells": {
            "totals": TOTAL_CELLS,
            "annualRows": ANNUAL_ROWS,
        },
        "assessmentWorkbookSeed": seed,
    }


def build_payload() -> dict[str, Any]:
    versions = {}
    for version_key, spec in VERSION_SPECS.items():
        workbook_path = discover_workbook(spec["pattern"])
        versions[version_key] = extract_version_payload(version_key, workbook_path)

    generated_at = datetime.now(timezone(timedelta(hours=8))).isoformat(timespec="seconds")
    return {
        "meta": {
            "generator": "g281_generate_financial_versions.py",
            "generatedAt": generated_at,
            "workbookDirectory": str(WORKBOOK_DIR),
            "sheetName": ASSESSMENT_SHEET,
            "note": "\u62a5\u4ef7/\u5b9a\u70b9\u7cbe\u786e\u53e3\u5f84\u6765\u81ea\u300a\u9879\u76ee\u8bc4\u4f30\u6c47\u603b\uff08\u6606\u5c7190%\uff09\u300b\uff0c\u7528\u4e8e\u7a0b\u5e8f\u57fa\u7ebf\u9a8c\u8bc1\u4e0e\u7248\u672c\u5316 ASP/\u6210\u672c\u53c2\u8003\u3002",
        },
        "versionOrder": list(VERSION_SPECS.keys()),
        "versions": versions,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate exact financial version data from E281 quote/fixed workbooks.")
    parser.add_argument("--out", default="g281_data_financial_versions.json", help="Output JSON path.")
    args = parser.parse_args()

    payload = build_payload()
    output_path = Path(args.out)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
