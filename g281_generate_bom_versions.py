from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook


DETAIL_SHEET = "二次物料明细"
OUTPUT_PATH = Path("g281_data_bom_versions.json")
MASTER_PATH = Path("g281_data_master.json")
TZ = timezone(timedelta(hours=8))

WIRE_KEYWORDS = ("导线", "线缆", "电缆")
TAPE_KEYWORDS = ("胶带",)


def discover_workbook(keyword: str) -> Path:
    matches = [
        path
        for path in Path(".").glob("*.xlsx")
        if keyword in path.name and not path.name.startswith("~$")
    ]
    if not matches:
        raise FileNotFoundError(f"Workbook containing '{keyword}' not found in current directory.")
    return sorted(matches)[0]


def to_text(value: object) -> str:
    return str(value or "").strip()


def to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def same_number(left: float | None, right: float | None, tolerance: float = 1e-9) -> bool:
    if left is None and right is None:
        return True
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def classify_meter_item(part_number: str, part_name: str, unit: str) -> str:
    if unit.upper() != "M":
        return "other"
    haystack = f"{part_number} {part_name}"
    if any(keyword in haystack for keyword in WIRE_KEYWORDS):
        return "wire"
    if any(keyword in haystack for keyword in TAPE_KEYWORDS):
        return "tape"
    return "tube"


def round_or_none(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def load_master_defaults() -> dict[str, float]:
    payload = json.loads(MASTER_PATH.read_text(encoding="utf-8"))
    defaults = payload.get("bomDefaults", {})
    return {
        "wireDrawing": float(defaults.get("wireDrawing", 0) or 0),
        "wireEat": float(defaults.get("wireEat", 0) or 0),
        "wireHidden": float(defaults.get("wireHidden", 0) or 0),
        "tapeDiameter": float(defaults.get("tapeDiameter", 0) or 0),
        "tapeWidth": float(defaults.get("tapeWidth", 0) or 0),
        "tapeOverlap": float(defaults.get("tapeOverlap", 0) or 0),
    }


def load_detail_metrics(workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[DETAIL_SHEET]
    metrics = {
        "totalMeter": 0.0,
        "wireMeter": 0.0,
        "tapeMeter": 0.0,
        "tubeMeter": 0.0,
        "meterRows": 0,
        "wireRows": 0,
        "tapeRows": 0,
        "tubeRows": 0,
        "samples": [],
    }
    for row_index in range(3, sheet.max_row + 1):
        part_number = to_text(sheet.cell(row_index, 1).value)
        part_name = to_text(sheet.cell(row_index, 2).value)
        quantity = to_float(sheet.cell(row_index, 3).value)
        unit = to_text(sheet.cell(row_index, 4).value)
        if quantity is None or unit.upper() != "M":
            continue
        category = classify_meter_item(part_number, part_name, unit)
        metrics["totalMeter"] += quantity
        metrics["meterRows"] += 1
        if category == "wire":
            metrics["wireMeter"] += quantity
            metrics["wireRows"] += 1
        elif category == "tape":
            metrics["tapeMeter"] += quantity
            metrics["tapeRows"] += 1
        else:
            metrics["tubeMeter"] += quantity
            metrics["tubeRows"] += 1
        if len(metrics["samples"]) < 8:
            metrics["samples"].append(
                {
                    "partNumber": part_number,
                    "partName": part_name,
                    "quantity": round(quantity, 6),
                    "unit": unit,
                    "category": category,
                    "sourceCell": f"C{row_index}",
                }
            )
    return metrics


def tape_per_mm(diameter: float, width: float, overlap_pct: float) -> float:
    pitch = width * (1 - overlap_pct / 100)
    circumference = math.pi * diameter
    if pitch <= 0:
        return 0
    return math.sqrt(circumference * circumference + pitch * pitch) / pitch


def derive_overlap(
    tape_diameter: float,
    tape_width: float,
    base_overlap: float,
    wire_factor: float,
    tape_factor: float,
) -> float:
    if tape_diameter <= 0 or tape_width <= 0 or wire_factor <= 0 or tape_factor <= 0:
        return base_overlap
    base_per_mm = tape_per_mm(tape_diameter, tape_width, base_overlap)
    if base_per_mm <= 1:
        return base_overlap
    target_per_mm = base_per_mm * (tape_factor / wire_factor)
    if target_per_mm <= 1:
        return base_overlap
    circumference = math.pi * tape_diameter
    denominator = max(target_per_mm * target_per_mm - 1, 1e-9)
    pitch = circumference / math.sqrt(denominator)
    overlap = (1 - pitch / tape_width) * 100
    if not math.isfinite(overlap):
        return base_overlap
    return max(0, min(95, overlap))


def derive_draft(
    base_defaults: dict[str, float],
    current_metrics: dict[str, object],
    quote_metrics: dict[str, object],
) -> dict[str, float]:
    base_drawing = base_defaults["wireDrawing"]
    base_eat = base_defaults["wireEat"]
    base_hidden = base_defaults["wireHidden"]
    base_diameter = base_defaults["tapeDiameter"]
    base_width = base_defaults["tapeWidth"]
    base_overlap = base_defaults["tapeOverlap"]
    base_total_mm = base_drawing + base_eat + base_hidden

    quote_wire = float(quote_metrics.get("wireMeter", 0) or 0)
    current_wire = float(current_metrics.get("wireMeter", 0) or 0)
    quote_tape = float(quote_metrics.get("tapeMeter", 0) or 0)
    current_tape = float(current_metrics.get("tapeMeter", 0) or 0)

    wire_factor = current_wire / quote_wire if quote_wire > 0 else 1
    tape_factor = current_tape / quote_tape if quote_tape > 0 else wire_factor
    derived_total_mm = base_total_mm * wire_factor
    derived_drawing = max(derived_total_mm - base_eat - base_hidden, 0)
    derived_overlap = derive_overlap(base_diameter, base_width, base_overlap, wire_factor, tape_factor)

    return {
        "bomWireDrawing": round(derived_drawing, 3),
        "bomWireEat": round(base_eat, 3),
        "bomWireHidden": round(base_hidden, 3),
        "bomTapeDiameter": round(base_diameter, 3),
        "bomTapeWidth": round(base_width, 3),
        "bomTapeOverlap": round(derived_overlap, 3),
    }


def build_snapshot(
    kind: str,
    label: str,
    workbook_path: Path,
    base_defaults: dict[str, float],
    current_metrics: dict[str, object],
    quote_metrics: dict[str, object],
) -> dict[str, object]:
    quote_total = float(quote_metrics.get("totalMeter", 0) or 0)
    quote_wire = float(quote_metrics.get("wireMeter", 0) or 0)
    quote_tape = float(quote_metrics.get("tapeMeter", 0) or 0)
    quote_tube = float(quote_metrics.get("tubeMeter", 0) or 0)
    total_meter = float(current_metrics.get("totalMeter", 0) or 0)
    wire_meter = float(current_metrics.get("wireMeter", 0) or 0)
    tape_meter = float(current_metrics.get("tapeMeter", 0) or 0)
    tube_meter = float(current_metrics.get("tubeMeter", 0) or 0)

    material_factor = total_meter / quote_total if quote_total > 0 else 1
    wire_factor = wire_meter / quote_wire if quote_wire > 0 else 1
    tape_factor = tape_meter / quote_tape if quote_tape > 0 else 1
    tube_factor = tube_meter / quote_tube if quote_tube > 0 else 1

    return {
        "kind": kind,
        "label": label,
        "workbook": workbook_path.name,
        "detailSheet": DETAIL_SHEET,
        "materialFactor": round(material_factor, 6),
        "wireFactor": round(wire_factor, 6),
        "tapeFactor": round(tape_factor, 6),
        "tubeFactor": round(tube_factor, 6),
        "totalMeter": round(total_meter, 6),
        "wireMeter": round(wire_meter, 6),
        "tapeMeter": round(tape_meter, 6),
        "tubeMeter": round(tube_meter, 6),
        "meterRows": int(current_metrics.get("meterRows", 0) or 0),
        "wireRows": int(current_metrics.get("wireRows", 0) or 0),
        "tapeRows": int(current_metrics.get("tapeRows", 0) or 0),
        "tubeRows": int(current_metrics.get("tubeRows", 0) or 0),
        "draft": derive_draft(base_defaults, current_metrics, quote_metrics),
        "samples": current_metrics.get("samples", []),
        "sources": {
            "detailSheet": f"{workbook_path.name}!{DETAIL_SHEET}",
            "metricRule": "按《二次物料明细》中单位=M的零件统计导线/胶带/套管总用量；材料系数按总 M 类用量相对报价版的比值估算。",
        },
    }


def compare_tt_harness_lengths(reference_path: Path, current_path: Path) -> dict[str, object]:
    old_wb = load_workbook(reference_path, data_only=True, read_only=True)
    new_wb = load_workbook(current_path, data_only=True, read_only=True)
    changes: list[dict[str, object]] = []

    for sheet_name in new_wb.sheetnames:
        if not sheet_name.isdigit() or sheet_name not in old_wb.sheetnames:
            continue
        old_sheet = old_wb[sheet_name]
        new_sheet = new_wb[sheet_name]
        max_row = max(old_sheet.max_row, new_sheet.max_row)
        for row_index in range(5, max_row + 1):
            old_value = to_float(old_sheet.cell(row_index, 10).value)
            new_value = to_float(new_sheet.cell(row_index, 10).value)
            if same_number(old_value, new_value):
                continue
            unit = to_text(new_sheet.cell(row_index, 11).value or old_sheet.cell(row_index, 11).value)
            if unit.upper() != "M":
                continue
            changes.append(
                {
                    "harnessNo": sheet_name,
                    "cell": f"J{row_index}",
                    "partNumber": to_text(new_sheet.cell(row_index, 3).value or old_sheet.cell(row_index, 3).value),
                    "partName": to_text(new_sheet.cell(row_index, 4).value or old_sheet.cell(row_index, 4).value),
                    "function": to_text(new_sheet.cell(row_index, 2).value or old_sheet.cell(row_index, 2).value),
                    "sap": to_text(new_sheet.cell(row_index, 6).value or old_sheet.cell(row_index, 6).value),
                    "oldQty": round_or_none(old_value),
                    "newQty": round_or_none(new_value),
                    "deltaQty": round_or_none((new_value or 0) - (old_value or 0)),
                    "unit": unit,
                }
            )

    harnesses = sorted({item["harnessNo"] for item in changes})
    old_total = sum(float(item["oldQty"] or 0) for item in changes)
    new_total = sum(float(item["newQty"] or 0) for item in changes)
    return {
        "referenceWorkbook": reference_path.name,
        "currentWorkbook": current_path.name,
        "changedHarnessCount": len(harnesses),
        "changedRowCount": len(changes),
        "oldQtyTotal": round(old_total, 6),
        "newQtyTotal": round(new_total, 6),
        "deltaQtyTotal": round(new_total - old_total, 6),
        "harnesses": harnesses,
        "rows": changes,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate BOM version snapshot JSON from quote, fixed, and TT workbooks.")
    parser.add_argument("--quote", type=Path, default=None, help="Quote BOM workbook path.")
    parser.add_argument("--fixed", type=Path, default=None, help="Fixed BOM workbook path.")
    parser.add_argument("--tt", type=Path, default=None, help="TT BOM workbook path.")
    parser.add_argument("--tt-reference", type=Path, default=None, help="Reference TT workbook path for actual-length comparison.")
    parser.add_argument("--out", type=Path, default=OUTPUT_PATH, help="Output JSON path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_defaults = load_master_defaults()
    quote_bom = args.quote or discover_workbook("报价BOM")
    fixed_bom = args.fixed or discover_workbook("定点BOM")
    tt_reference = args.tt_reference or discover_workbook("G281 TT.xlsx")
    tt_current = args.tt or discover_workbook("G281 TT_")

    quote_metrics = load_detail_metrics(quote_bom)
    fixed_metrics = load_detail_metrics(fixed_bom)
    tt_metrics = load_detail_metrics(tt_current)
    tt_changes = compare_tt_harness_lengths(tt_reference, tt_current)

    payload = {
        "meta": {
            "generator": "g281_generate_bom_versions.py",
            "generatedAt": datetime.now(TZ).isoformat(timespec="seconds"),
            "quoteWorkbook": quote_bom.name,
            "fixedWorkbook": fixed_bom.name,
            "ttWorkbook": tt_current.name,
            "ttReferenceWorkbook": tt_reference.name,
            "layer": "JSON 数据层",
            "note": "BOM 版本快照来自报价 BOM、定点 BOM 和 TT 实际开线长度回填工作簿；材料系数按《二次物料明细》M 类总用量口径估算。",
        },
        "versionSnapshots": {
            "quote": build_snapshot("quote", "报价版BOM", quote_bom, base_defaults, quote_metrics, quote_metrics),
            "fixed": build_snapshot("fixed", "定点版BOM", fixed_bom, base_defaults, fixed_metrics, quote_metrics),
            "tt": build_snapshot("tt", "TT版BOM", tt_current, base_defaults, tt_metrics, quote_metrics),
        },
        "ttActualLengthChanges": tt_changes,
    }
    payload["versionSnapshots"]["tt"]["actualLengthChangeSummary"] = {
        "changedHarnessCount": tt_changes["changedHarnessCount"],
        "changedRowCount": tt_changes["changedRowCount"],
        "deltaQtyTotal": tt_changes["deltaQtyTotal"],
    }
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(args.out)


if __name__ == "__main__":
    main()
