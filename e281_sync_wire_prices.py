from __future__ import annotations

import re
import shutil
from collections import Counter
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import win32com.client  # type: ignore[import-not-found]
except ImportError:  # pragma: no cover - optional at runtime
    win32com = None


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
TMP_DIR = ROOT / "tmp" / "spreadsheets"
VERSION_TAG = "V02-2026.03.26"

DETAIL_SHEET = "二次物料明细"
BOM_SHEET = "KSK线束BOM明细"
ASSEMBLY_SHEET = "总成散件清单"
CONFIG_SHEET = "配置明细"
WIRE_PRICE_SHEET = "导线价格"

WIRE_HEADERS = [
    "组件描述",
    "物料名称",
    "供应商",
    "铝重",
    "铜重",
    "非铜",
    "币别",
    "价格类型",
    "导线单价",
    "来源核算表",
    "计算逻辑备注",
    "来源明细行",
]

BOM_EXTRA_HEADERS = {
    "Z": "供应商",
    "AA": "责任人",
    "AB": "单价",
    "AC": "铝重",
    "AD": "铜重",
    "AE": "非铜",
    "AF": "导线单价",
    "AG": "材料成本",
    "AH": "价格类型",
}

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="F3F3F3")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

WIRE_FORMULA_RE = re.compile(
    r"=O\d+/1000\*(?P<al>[^+]+)\+P\d+/1000\*(?P<cu>[^+]+)\+Q\d+/1000",
    re.IGNORECASE,
)
CELL_REF_RE = re.compile(r"^\$?([A-Z]{1,3})\$?(\d+)$")
NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")


@dataclass
class BasePriceInfo:
    aluminum: float
    copper: float
    aluminum_note: str
    copper_note: str


@dataclass
class WorkbookResult:
    output_path: Path
    wire_row_key: str
    detail_row: int
    bom_row: int
    assembly_row: int | None
    base_cell_al: str
    base_cell_cu: str


def copy_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)  # noqa: SLF001
    if source.comment:
        target.comment = copy(source.comment)


def find_last_data_row(ws, key_column: str) -> int:
    for row in range(ws.max_row, 1, -1):
        if ws[f"{key_column}{row}"].value not in (None, ""):
            return row
    return 1


def extract_numeric(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        match = NUMBER_RE.search(value.replace(",", ""))
        if match:
            return float(match.group())
    return None


def resolve_formula_token(token: str, ws) -> float | None:
    token = token.strip()
    numeric = extract_numeric(token)
    if numeric is not None and CELL_REF_RE.fullmatch(token) is None:
        return numeric
    match = CELL_REF_RE.fullmatch(token)
    if not match:
        return None
    cell_value = ws[token.replace("$", "")].value
    return extract_numeric(cell_value)


def format_number(value: float | None) -> str:
    if value is None:
        return "未识别"
    return f"{value:.4f}".rstrip("0").rstrip(".")


def collect_wire_base_info(workbook) -> BasePriceInfo:
    bom_ws = workbook[BOM_SHEET]
    config_ws = workbook[CONFIG_SHEET]
    last_bom_row = find_last_data_row(bom_ws, "C")
    al_values: list[float] = []
    cu_values: list[float] = []
    cu_candidates: set[float] = set()

    for row in range(2, last_bom_row + 1):
        formula = bom_ws[f"R{row}"].value
        if not isinstance(formula, str):
            continue
        match = WIRE_FORMULA_RE.match(formula)
        if not match:
            continue
        al_value = resolve_formula_token(match.group("al"), bom_ws)
        cu_value = resolve_formula_token(match.group("cu"), bom_ws)
        if al_value is not None:
            al_values.append(al_value)
        if cu_value is not None:
            cu_values.append(cu_value)
            cu_candidates.add(cu_value)

    dominant_al = Counter(al_values).most_common(1)[0][0] if al_values else 18.91
    dominant_cu = Counter(cu_values).most_common(1)[0][0] if cu_values else 76.45

    config_al = extract_numeric(config_ws["I25"].value)
    config_cu = extract_numeric(config_ws["I26"].value)
    note_al_parts = [f"默认值按 {BOM_SHEET} 原导线公式主值 {format_number(dominant_al)} 初始化"]
    note_cu_parts = [f"默认值按 {BOM_SHEET} 原导线公式主值 {format_number(dominant_cu)} 初始化"]

    if config_al is not None and abs(config_al - dominant_al) > 1e-9:
        note_al_parts.append(f"另检测到 {CONFIG_SHEET}!I25={format_number(config_al)}")
    elif config_al is not None:
        note_al_parts.append(f"同时匹配 {CONFIG_SHEET}!I25")

    if config_cu is not None and abs(config_cu - dominant_cu) > 1e-9:
        note_cu_parts.append(f"另检测到 {CONFIG_SHEET}!I26={format_number(config_cu)}")
    elif config_cu is not None:
        note_cu_parts.append(f"同时匹配 {CONFIG_SHEET}!I26")

    if len(cu_candidates) > 1:
        others = ", ".join(format_number(v) for v in sorted(cu_candidates) if abs(v - dominant_cu) > 1e-9)
        if others:
            note_cu_parts.append(f"原公式中还出现 {others}")

    return BasePriceInfo(
        aluminum=dominant_al,
        copper=dominant_cu,
        aluminum_note="；".join(note_al_parts),
        copper_note="；".join(note_cu_parts),
    )


def is_wire_row(ws, row: int) -> bool:
    description = str(ws[f"B{row}"].value or "")
    weight_values = [extract_numeric(ws[f"{col}{row}"].value) or 0 for col in ("J", "K", "L")]
    return "导线" in description or any(value != 0 for value in weight_values)


def collect_wire_rows(detail_ws) -> list[int]:
    last_row = find_last_data_row(detail_ws, "A")
    return [row for row in range(3, last_row + 1) if detail_ws[f"A{row}"].value and is_wire_row(detail_ws, row)]


def build_wire_sheet(workbook, source_name: str, base_info: BasePriceInfo, wire_rows: Iterable[int]):
    if WIRE_PRICE_SHEET in workbook.sheetnames:
        old_sheet = workbook[WIRE_PRICE_SHEET]
        workbook.remove(old_sheet)

    detail_index = workbook.sheetnames.index(DETAIL_SHEET)
    ws = workbook.create_sheet(WIRE_PRICE_SHEET, detail_index + 1)

    ws["A1"] = "导线价格联动表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "来源核算表"
    ws["B2"] = source_name
    ws["A3"] = "铝基价"
    ws["B3"] = base_info.aluminum
    ws["A4"] = "铜基价"
    ws["B4"] = base_info.copper
    ws["D2"] = "计算逻辑"
    ws["E2"] = "导线单价 = 铝重/1000*铝基价 + 铜重/1000*铜基价 + 非铜/1000"
    ws["D3"] = "铝基价来源"
    ws["E3"] = base_info.aluminum_note
    ws["D4"] = "铜基价来源"
    ws["E4"] = base_info.copper_note

    for address in ("A2", "A3", "A4", "D2", "D3", "D4"):
        ws[address].font = HEADER_FONT
        ws[address].fill = HEADER_FILL
        ws[address].alignment = CENTER
    for address in ("B3", "B4"):
        ws[address].fill = INPUT_FILL
        ws[address].number_format = "0.0000"
    for address in ("E2", "E3", "E4"):
        ws[address].fill = NOTE_FILL
        ws[address].alignment = LEFT_WRAP

    ws["B3"].comment = Comment(base_info.aluminum_note, "Codex")
    ws["B4"].comment = Comment(base_info.copper_note, "Codex")

    header_row = 6
    for col_idx, header in enumerate(WIRE_HEADERS, start=1):
        cell = ws.cell(header_row, col_idx, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    detail_ws = workbook[DETAIL_SHEET]
    current_row = header_row + 1
    for detail_row in wire_rows:
        ws[f"A{current_row}"] = f"='{DETAIL_SHEET}'!A{detail_row}"
        ws[f"B{current_row}"] = f"='{DETAIL_SHEET}'!B{detail_row}"
        ws[f"C{current_row}"] = f"='{DETAIL_SHEET}'!E{detail_row}"
        ws[f"D{current_row}"] = f"='{DETAIL_SHEET}'!J{detail_row}"
        ws[f"E{current_row}"] = f"='{DETAIL_SHEET}'!K{detail_row}"
        ws[f"F{current_row}"] = f"='{DETAIL_SHEET}'!L{detail_row}"
        ws[f"G{current_row}"] = f"='{DETAIL_SHEET}'!M{detail_row}"
        ws[f"H{current_row}"] = f"='{DETAIL_SHEET}'!N{detail_row}"
        ws[f"I{current_row}"] = f"=IFERROR(D{current_row}/1000*$B$3+E{current_row}/1000*$B$4+F{current_row}/1000,0)"
        ws[f"J{current_row}"] = source_name
        ws[f"K{current_row}"] = (
            f"逻辑抽自 {source_name} 原 {BOM_SHEET} 导线成本公式，"
            "现改为集中维护铜/铝基价。"
        )
        ws[f"L{current_row}"] = detail_row

        for col in "ABCDEFGHIJKL":
            ws[f"{col}{current_row}"].alignment = LEFT_WRAP if col in "ABJK" else CENTER
        ws[f"I{current_row}"].number_format = "0.000000"
        current_row += 1

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:L{max(current_row - 1, 6)}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[6].height = 24

    widths = {
        "A": 28,
        "B": 36,
        "C": 18,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 14,
        "J": 20,
        "K": 42,
        "L": 10,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    return ws


def update_detail_sheet(detail_ws, wire_rows: Iterable[int]) -> None:
    for row in wire_rows:
        detail_ws[f"I{row}"] = f"=IFERROR(VLOOKUP(A{row},{WIRE_PRICE_SHEET}!$A:$L,9,0),0)"


def ensure_bom_extra_columns(bom_ws, last_row: int) -> None:
    width_map = {"Z": "J", "AA": "M", "AB": "N", "AC": "O", "AD": "P", "AE": "Q", "AF": "R", "AG": "S", "AH": "X"}

    for target_col, header in BOM_EXTRA_HEADERS.items():
        bom_ws[f"{target_col}1"] = header
        copy_style(bom_ws[f"{width_map[target_col]}1"], bom_ws[f"{target_col}1"])
        if bom_ws[f"{target_col}1"].value != header:
            bom_ws[f"{target_col}1"] = header
        source_width = bom_ws.column_dimensions[width_map[target_col]].width
        bom_ws.column_dimensions[target_col].width = source_width

    for row in range(2, last_row + 1):
        for target_col, source_col in width_map.items():
            copy_style(bom_ws[f"{source_col}{row}"], bom_ws[f"{target_col}{row}"])


def update_bom_sheet(bom_ws, last_row: int) -> None:
    ensure_bom_extra_columns(bom_ws, last_row)
    for row in range(2, last_row + 1):
        bom_ws[f"N{row}"] = (
            f'=IF(COUNTIF({WIRE_PRICE_SHEET}!$A:$A,$C{row})>0,0,'
            f'IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,9,0),0))'
        )
        bom_ws[f"R{row}"] = f'=IFERROR(VLOOKUP($C{row},{WIRE_PRICE_SHEET}!$A:$L,9,0),0)'
        bom_ws[f"S{row}"] = f"=(N{row}+R{row})*H{row}"

        bom_ws[f"Z{row}"] = f'=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,5,0),"")'
        bom_ws[f"AA{row}"] = f'=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,8,0),"")'
        bom_ws[f"AB{row}"] = f"=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,9,0),0)"
        bom_ws[f"AC{row}"] = f"=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,10,0),0)"
        bom_ws[f"AD{row}"] = f"=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,11,0),0)"
        bom_ws[f"AE{row}"] = f"=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,12,0),0)"
        bom_ws[f"AF{row}"] = f'=IFERROR(VLOOKUP($C{row},{WIRE_PRICE_SHEET}!$A:$L,9,0),0)'
        bom_ws[f"AG{row}"] = (
            f"=(IF(COUNTIF({WIRE_PRICE_SHEET}!$A:$A,$C{row})>0,0,AB{row})+AF{row})*H{row}"
        )
        bom_ws[f"AH{row}"] = f'=IFERROR(VLOOKUP($C{row},{DETAIL_SHEET}!$A:$O,14,0),"")'


def update_assembly_sheet(assembly_ws) -> int | None:
    last_row = find_last_data_row(assembly_ws, "D")
    first_formula_row: int | None = None
    for row in range(2, last_row + 1):
        if assembly_ws[f"D{row}"].value in (None, ""):
            continue
        price_cell = assembly_ws[f"C{row}"]
        if isinstance(price_cell, MergedCell):
            continue
        existing_value = price_cell.value
        fallback = '""' if existing_value in (None, "") else str(existing_value)
        assembly_key_formula = f"LEFT($B{row},FIND(CHAR(10),$B{row}&CHAR(10))-1)"
        assembly_ws[f"C{row}"] = (
            f"=IFERROR(VLOOKUP($D{row},{DETAIL_SHEET}!$A:$O,9,0),"
            f"IFERROR(VLOOKUP({assembly_key_formula},{DETAIL_SHEET}!$A:$O,9,0),{fallback}))"
        )
        if first_formula_row is None:
            first_formula_row = row
    return first_formula_row


def set_workbook_calc_flags(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"


def recalculate_with_excel(paths: list[Path]) -> None:
    if not paths or win32com is None:
        return

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    excel.Calculation = -4105  # xlCalculationAutomatic
    try:
        for path in paths:
            workbook = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False)
            workbook.ForceFullCalculation = True
            workbook.RefreshAll()
            excel.CalculateFullRebuild()
            workbook.Save()
            workbook.Close(SaveChanges=True)
    finally:
        excel.Quit()


def validate_linkage(result: WorkbookResult) -> dict[str, float | str | None]:
    formula_wb = load_workbook(result.output_path, data_only=False)
    value_wb = load_workbook(result.output_path, data_only=True)

    detail_formula = formula_wb[DETAIL_SHEET][f"I{result.detail_row}"].value
    bom_formula = formula_wb[BOM_SHEET][f"R{result.bom_row}"].value
    unit_formula = formula_wb[BOM_SHEET][f"AB{result.bom_row}"].value
    agg_formula = formula_wb[BOM_SHEET][f"AG{result.bom_row}"].value
    assembly_formula = None
    if result.assembly_row:
        assembly_formula = formula_wb[ASSEMBLY_SHEET][f"C{result.assembly_row}"].value

    detail_value = value_wb[DETAIL_SHEET][f"I{result.detail_row}"].value
    bom_value = value_wb[BOM_SHEET][f"R{result.bom_row}"].value
    unit_value = value_wb[BOM_SHEET][f"AB{result.bom_row}"].value

    if win32com is None:
        changed = None
    else:
        TMP_DIR.mkdir(parents=True, exist_ok=True)
        temp_path = TMP_DIR / f"validate_{result.output_path.name}"
        shutil.copy2(result.output_path, temp_path)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.Calculation = -4105
        try:
            workbook = excel.Workbooks.Open(str(temp_path), UpdateLinks=0, ReadOnly=False)
            wire_ws = workbook.Worksheets(WIRE_PRICE_SHEET)
            detail_ws = workbook.Worksheets(DETAIL_SHEET)
            before = detail_ws.Range(f"I{result.detail_row}").Value
            wire_ws.Range(result.base_cell_cu).Value = float(wire_ws.Range(result.base_cell_cu).Value) + 1
            excel.CalculateFullRebuild()
            after = detail_ws.Range(f"I{result.detail_row}").Value
            changed = None if before is None or after is None else float(after) - float(before)
            workbook.Close(SaveChanges=False)
        finally:
            excel.Quit()
            if temp_path.exists():
                temp_path.unlink()

    return {
        "detail_formula": detail_formula,
        "bom_formula": bom_formula,
        "unit_formula": unit_formula,
        "agg_formula": agg_formula,
        "assembly_formula": assembly_formula,
        "detail_value": detail_value,
        "bom_value": bom_value,
        "unit_value": unit_value,
        "base_change_effect": changed,
    }


def build_output_name(source_path: Path) -> Path:
    return OUTPUT_DIR / f"{source_path.stem}_导线联动_{VERSION_TAG}{source_path.suffix}"


def process_workbook(source_path: Path) -> WorkbookResult:
    workbook = load_workbook(source_path, data_only=False)
    detail_ws = workbook[DETAIL_SHEET]
    bom_ws = workbook[BOM_SHEET]
    assembly_ws = workbook[ASSEMBLY_SHEET]

    base_info = collect_wire_base_info(workbook)
    wire_rows = collect_wire_rows(detail_ws)
    if not wire_rows:
        raise RuntimeError(f"{source_path.name} 未找到可提取的导线行。")

    wire_ws = build_wire_sheet(workbook, source_path.name, base_info, wire_rows)
    update_detail_sheet(detail_ws, wire_rows)

    last_bom_row = find_last_data_row(bom_ws, "C")
    update_bom_sheet(bom_ws, last_bom_row)
    first_assembly_row = update_assembly_sheet(assembly_ws)
    set_workbook_calc_flags(workbook)

    output_path = build_output_name(source_path)
    workbook.save(output_path)

    wire_key = detail_ws[f"A{wire_rows[0]}"].value
    bom_row = next(
        row for row in range(2, last_bom_row + 1) if bom_ws[f"C{row}"].value == wire_key
    )
    return WorkbookResult(
        output_path=output_path,
        wire_row_key=wire_key,
        detail_row=wire_rows[0],
        bom_row=bom_row,
        assembly_row=first_assembly_row,
        base_cell_al="B3",
        base_cell_cu="B4",
    )


def detect_source_workbooks() -> list[Path]:
    files = [
        path
        for path in ROOT.glob("*E281*核算*.xlsx")
        if not path.name.startswith("~$") and "导线联动" not in path.stem
    ]
    return sorted(files)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    source_files = detect_source_workbooks()
    if not source_files:
        raise RuntimeError("当前目录未找到 E281 核算源文件。")

    results = [process_workbook(path) for path in source_files]
    recalculate_with_excel([result.output_path for result in results])

    for result in results:
        validation = validate_linkage(result)
        print(f"[OK] {result.output_path.name}")
        print(f"  导线样本: {result.wire_row_key}")
        print(f"  导线页改单价后影响: {validation['base_change_effect']}")
        print(f"  二次物料单价: {validation['detail_value']}")
        print(f"  BOM导线单价: {validation['bom_value']}")
        print(f"  BOM显式单价: {validation['unit_value']}")
        print(f"  公式检查: {validation['detail_formula']}")
        print(f"  公式检查: {validation['bom_formula']}")
        print(f"  公式检查: {validation['agg_formula']}")
        if validation["assembly_formula"] is not None:
            print(f"  总成散件参考价公式: {validation['assembly_formula']}")


if __name__ == "__main__":
    main()
