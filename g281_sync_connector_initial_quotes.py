from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


DETAIL_SHEET = "二次物料明细"
ASSEMBLY_SHEET = "总成散件清单"
PROTOCOL_JSON = Path("g281_data_connector_protocol_status.json")


def discover_workbook(keyword: str) -> Path:
    matches = [
        path
        for path in Path(".").glob("*.xlsx")
        if keyword in path.name and not path.name.startswith("~$")
    ]
    if not matches:
        raise FileNotFoundError(f"未找到包含“{keyword}”的工作簿")
    return sorted(matches)[0]


def normalize_code(value: object) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = text.split(" / ")[0].split("\n")[0].strip()
    text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"\s+", "", text)
    return text


def load_detail_lookup(workbook: Path) -> dict[str, dict[str, object]]:
    wb = load_workbook(workbook, data_only=True, read_only=True)
    ws = wb[DETAIL_SHEET]
    lookup: dict[str, dict[str, object]] = {}
    for row_index in range(3, ws.max_row + 1):
        desc = ws.cell(row_index, 1).value
        name = ws.cell(row_index, 2).value
        unit = ws.cell(row_index, 4).value
        if desc is None:
            continue
        price = ws.cell(row_index, 9).value
        if price in (None, ""):
            continue
        text = f"{desc or ''} {name or ''}"
        if str(unit or "").strip().upper() != "SET" and not any(keyword in text for keyword in ("连接器", "总成", "插座", "IPT")):
            continue
        code = normalize_code(desc)
        if not code or code in lookup:
            continue
        lookup[code] = {
            "row": row_index,
            "desc": str(desc).strip(),
            "name": str(name or "").strip(),
            "supplier": str(ws.cell(row_index, 5).value or "").strip(),
            "price": float(price),
            "priceType": str(ws.cell(row_index, 14).value or "").strip(),
        }
    return lookup


def load_assembly_aliases(workbook: Path) -> dict[str, set[str]]:
    wb = load_workbook(workbook, data_only=True, read_only=True)
    ws = wb[ASSEMBLY_SHEET]
    aliases: dict[str, set[str]] = {}
    for row_index in range(2, ws.max_row + 1):
        function_cell = ws.cell(row_index, 2).value
        part_number = ws.cell(row_index, 4).value
        left = normalize_code(function_cell)
        right = normalize_code(part_number)
        if not left or not right:
            continue
        aliases.setdefault(left, set()).add(right)
        aliases.setdefault(right, set()).add(left)
    return aliases


def protocol_codes(row: dict[str, object]) -> list[str]:
    codes = [
        normalize_code(row.get("functionRaw")),
        normalize_code(row.get("functionBrief")),
        normalize_code(row.get("partNumber")),
    ]
    deduped: list[str] = []
    for code in codes:
        if code and code not in deduped:
            deduped.append(code)
    return deduped


def enrich_protocol_rows(
    payload: dict[str, object],
    detail_lookup: dict[str, dict[str, object]],
    aliases: dict[str, set[str]],
    workbook_name: str,
) -> int:
    matched = 0
    rows = payload.get("rows") or []
    for row in rows:
        if not isinstance(row, dict):
            continue
        match = None
        match_type = ""
        for code in protocol_codes(row):
            match = detail_lookup.get(code)
            if match:
                match_type = "direct"
                break
        if not match:
            for code in protocol_codes(row):
                for alias in aliases.get(code, set()):
                    match = detail_lookup.get(alias)
                    if match:
                        match_type = "assembly_bridge"
                        break
                if match:
                    break
        if match:
            row["initialQuote"] = match["price"]
            row["initialQuoteSource"] = f"{workbook_name}!{DETAIL_SHEET}!I{match['row']}"
            row["initialQuoteCode"] = match["desc"]
            row["initialQuotePriceType"] = match["priceType"]
            row["initialQuoteMatchType"] = match_type
            matched += 1
        else:
            row["initialQuote"] = None
            row["initialQuoteSource"] = ""
            row["initialQuoteCode"] = ""
            row["initialQuotePriceType"] = ""
            row["initialQuoteMatchType"] = "none"
    return matched


def main() -> None:
    workbook = discover_workbook("报价核算")
    payload = json.loads(PROTOCOL_JSON.read_text(encoding="utf-8"))
    detail_lookup = load_detail_lookup(workbook)
    aliases = load_assembly_aliases(workbook)
    matched = enrich_protocol_rows(payload, detail_lookup, aliases, workbook.name)
    payload.setdefault("meta", {})
    payload["meta"]["initialQuoteWorkbook"] = workbook.name
    payload["meta"]["initialQuoteSheet"] = DETAIL_SHEET
    payload["meta"]["initialQuoteRule"] = "报价核算二次物料明细优先，必要时通过总成散件清单做总成号桥接；未匹配到则留空。"
    payload["meta"]["initialQuoteMatchedRows"] = matched
    PROTOCOL_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(PROTOCOL_JSON)
    print(f"matched_rows={matched}")


if __name__ == "__main__":
    main()
