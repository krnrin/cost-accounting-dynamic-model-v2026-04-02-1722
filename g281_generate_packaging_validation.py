from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_INDEX = {
    "summary": 0,
    "assessment": 1,
    "packaging": 10,
}

NUMBER_RE = re.compile(r"(-?\d+(?:\.\d+)?)")


def discover_workbook(pattern: str) -> Path:
    matches = [
        Path(name)
        for name in os.listdir(".")
        if name.endswith(".xlsx") and not name.startswith("~$") and pattern in name and "导线联动" not in name
    ]
    if not matches:
        raise FileNotFoundError(f"Workbook containing '{pattern}' not found in current directory.")
    return sorted(matches)[0]


def collapse_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\s+", " ", text)


def numeric_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = collapse_text(value).replace(",", "")
    if not text or text.startswith("#"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def round_number(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def note_join(*parts: str) -> str:
    return "；".join(part for part in parts if collapse_text(part))


def make_item(
    item_key: str,
    label: str,
    unit: str,
    source_sheet: str,
    source_cell: str,
    raw_value: Any,
    formula_value: Any = "",
    *,
    numeric_override: float | None = None,
    note: str = "",
    display_label: str = "",
) -> dict[str, Any]:
    numeric = round_number(numeric_override if numeric_override is not None else numeric_value(raw_value))
    raw_text = collapse_text(raw_value)
    formula_text = collapse_text(formula_value)
    value: Any = numeric if numeric is not None else (raw_text or "-")
    value_type = "number" if numeric is not None else ("error" if raw_text.startswith("#") else "text")
    return {
        "itemKey": item_key,
        "label": label,
        "displayLabel": display_label or label,
        "unit": unit,
        "value": value,
        "numericValue": numeric,
        "valueType": value_type,
        "sourceSheet": source_sheet,
        "sourceCell": source_cell,
        "source": f"{source_sheet}!{source_cell}",
        "formula": formula_text,
        "note": collapse_text(note),
        "rawValue": raw_text,
    }


def workbook_context(path: Path) -> dict[str, Any]:
    value_book = load_workbook(path, read_only=True, data_only=True)
    formula_book = load_workbook(path, read_only=True, data_only=False)
    sheets = {}
    for alias, index in SHEET_INDEX.items():
        title = value_book.sheetnames[index]
        sheets[alias] = {
            "title": title,
            "values": value_book[title],
            "formulas": formula_book[formula_book.sheetnames[index]],
        }
    return {
        "path": path,
        "sheets": sheets,
    }


def packaging_row_item(ctx: dict[str, Any], row: int) -> dict[str, Any]:
    ws_values = ctx["sheets"]["packaging"]["values"]
    ws_formulas = ctx["sheets"]["packaging"]["formulas"]
    customer_part = collapse_text(ws_values[f"A{row}"].value)
    sap = collapse_text(ws_values[f"B{row}"].value)
    description = collapse_text(ws_values[f"C{row}"].value)
    usage = round_number(numeric_value(ws_values[f"D{row}"].value))
    inner_pack = round_number(numeric_value(ws_values[f"E{row}"].value))
    outer_pack = round_number(numeric_value(ws_values[f"F{row}"].value))
    freight = round_number(numeric_value(ws_values[f"G{row}"].value))
    extra_freight = round_number(numeric_value(ws_values[f"H{row}"].value))
    short_haul_other = round_number(numeric_value(ws_values[f"I{row}"].value))
    third_party_warehouse = round_number(numeric_value(ws_values[f"J{row}"].value))
    storage = round_number(numeric_value(ws_values[f"K{row}"].value))
    total = round_number(numeric_value(ws_values[f"L{row}"].value))
    note = collapse_text(ws_values[f"M{row}"].value)

    breakdown = note_join(
        f"用量 {usage}" if usage is not None else "",
        f"内包装 {inner_pack}" if inner_pack is not None else "",
        f"外包装 {outer_pack}" if outer_pack is not None else "",
        f"运费 {freight}" if freight is not None else "",
        f"超额运费 {extra_freight}" if extra_freight is not None else "",
        f"短驳/其他 {short_haul_other}" if short_haul_other is not None else "",
        f"三方仓 {third_party_warehouse}" if third_party_warehouse is not None else "",
        f"仓储费 {storage}" if storage is not None else "",
        note,
    )

    return make_item(
        f"detail.{customer_part or row}",
        f"{customer_part} · {description}".strip(" ·"),
        "元/套",
        ctx["sheets"]["packaging"]["title"],
        f"L{row}",
        ws_values[f"L{row}"].value,
        ws_formulas[f"L{row}"].value,
        numeric_override=total,
        display_label=sap or customer_part or f"第{row}行",
        note=breakdown,
    )


def summary_cell_item(
    ctx: dict[str, Any],
    cell: str,
    item_key: str,
    label: str,
    unit: str,
    note: str = "",
) -> dict[str, Any]:
    ws_values = ctx["sheets"]["packaging"]["values"]
    ws_formulas = ctx["sheets"]["packaging"]["formulas"]
    return make_item(
        item_key,
        label,
        unit,
        ctx["sheets"]["packaging"]["title"],
        cell,
        ws_values[cell].value,
        ws_formulas[cell].value,
        note=note,
    )


def assessment_item(ctx: dict[str, Any], cell: str, item_key: str, label: str, unit: str, note: str = "") -> dict[str, Any]:
    ws_values = ctx["sheets"]["assessment"]["values"]
    ws_formulas = ctx["sheets"]["assessment"]["formulas"]
    return make_item(
        item_key,
        label,
        unit,
        ctx["sheets"]["assessment"]["title"],
        cell,
        ws_values[cell].value,
        ws_formulas[cell].value,
        note=note,
    )


def summary_item(ctx: dict[str, Any], cell: str, item_key: str, label: str, unit: str, note: str = "") -> dict[str, Any]:
    ws_values = ctx["sheets"]["summary"]["values"]
    ws_formulas = ctx["sheets"]["summary"]["formulas"]
    return make_item(
        item_key,
        label,
        unit,
        ctx["sheets"]["summary"]["title"],
        cell,
        ws_values[cell].value,
        ws_formulas[cell].value,
        note=note,
    )


def build_snapshot(ctx: dict[str, Any], kind: str) -> dict[str, Any]:
    ws = ctx["sheets"]["packaging"]["values"]
    inner_pack = numeric_value(ws["E25"].value) or 0
    outer_pack = numeric_value(ws["F25"].value) or 0
    freight = numeric_value(ws["G25"].value) or 0
    extra_freight = numeric_value(ws["H25"].value) or 0
    short_haul_other = numeric_value(ws["I25"].value) or 0
    third_party_warehouse = numeric_value(ws["J25"].value) or 0
    storage = numeric_value(ws["K25"].value) or 0
    total = numeric_value(ws["L25"].value) or 0

    return {
        "kind": kind,
        "label": "报价包装版" if kind == "quote" else "定点包装版",
        "packInner": round_number(inner_pack + outer_pack),
        "packFreight": round_number(freight + extra_freight),
        "packWarehouse": round_number(third_party_warehouse + storage),
        "packOther": round_number(short_haul_other),
        "packTotal": round_number(total),
        "sources": {
            "packInner": f"{ctx['sheets']['packaging']['title']}!E25 + F25",
            "packFreight": f"{ctx['sheets']['packaging']['title']}!G25 + H25",
            "packWarehouse": f"{ctx['sheets']['packaging']['title']}!J25 + K25",
            "packOther": f"{ctx['sheets']['packaging']['title']}!I25",
            "packTotal": f"{ctx['sheets']['packaging']['title']}!L25",
        },
        "note": "程序应用值按包装物流费用页单套成本拆分：内外包装=内包装+外包装，运输费=运费+超额运费，仓储费=三方仓+仓储费，短驳/其他=短驳及其他。",
    }


def workbook_groups(ctx: dict[str, Any], kind: str) -> dict[str, list[dict[str, Any]]]:
    snapshot = build_snapshot(ctx, kind)
    scenario_items = [
        make_item(
            "apply.packInner",
            "内外包装",
            "元/套",
            ctx["sheets"]["packaging"]["title"],
            "E25+F25",
            snapshot["packInner"],
            "",
            note="程序字段 = 内包装 + 外包装。",
        ),
        make_item(
            "apply.packFreight",
            "运输费",
            "元/套",
            ctx["sheets"]["packaging"]["title"],
            "G25+H25",
            snapshot["packFreight"],
            "",
            note="程序字段 = 运费 + 超额运费。",
        ),
        make_item(
            "apply.packWarehouse",
            "仓储费",
            "元/套",
            ctx["sheets"]["packaging"]["title"],
            "J25+K25",
            snapshot["packWarehouse"],
            "",
            note="程序字段 = 三方仓费用 + 仓储费。",
        ),
        make_item(
            "apply.packOther",
            "短驳 / 其他",
            "元/套",
            ctx["sheets"]["packaging"]["title"],
            "I25",
            snapshot["packOther"],
            "",
            note="程序字段 = 短驳（合肥仓库到客户）+ 其他。",
        ),
        make_item(
            "apply.packTotal",
            "程序包装物流合计",
            "元/套",
            ctx["sheets"]["packaging"]["title"],
            "L25",
            snapshot["packTotal"],
            "",
            note="四个程序字段汇总后应与包装物流费用页单套合计一致。",
        ),
        assessment_item(ctx, "F32", "assessment.packagingPerSet", "项目评估单套包装物流", "元/PCS", "来源：项目评估汇总引用包装物流费用!L25。"),
    ]

    breakdown_items = [
        summary_cell_item(ctx, "E25", "breakdown.innerPack", "内包装", "元/套"),
        summary_cell_item(ctx, "F25", "breakdown.outerPack", "外包装", "元/套"),
        summary_cell_item(ctx, "G25", "breakdown.freight", "运费", "元/套"),
        summary_cell_item(ctx, "H25", "breakdown.extraFreight", "超额运费", "元/套"),
        summary_cell_item(ctx, "I25", "breakdown.shortHaulOther", "短驳 / 其他", "元/套"),
        summary_cell_item(ctx, "J25", "breakdown.thirdWarehouse", "三方仓费用", "元/套"),
        summary_cell_item(ctx, "K25", "breakdown.storage", "仓储费", "元/套"),
        summary_cell_item(ctx, "L25", "breakdown.total", "包装物流单套合计", "元/套"),
        assessment_item(ctx, "F32", "breakdown.assessmentPackaging", "项目评估包装物流引用", "元/PCS", "项目评估汇总 F32 引用包装物流费用!L25。"),
    ]

    detail_items = []
    ws_values = ctx["sheets"]["packaging"]["values"]
    row = 4
    while True:
        marker = collapse_text(ws_values[f"A{row}"].value)
        if not marker:
            row += 1
            if row > 80:
                break
            continue
        if marker == "单套成本":
            break
        detail_items.append(packaging_row_item(ctx, row))
        row += 1

    return {
        "scenario": [
            {
                "key": "scenario.mapping",
                "label": "程序包装映射",
                "meta": "把包装物流费用页单套拆分映射到程序当前 4 个包装输入项。",
                "items": scenario_items,
            }
        ],
        "breakdown": [
            {
                "key": "breakdown.perSet",
                "label": "单套包装拆分",
                "meta": "来源：包装物流费用 + 项目评估/项目汇总引用。",
                "items": breakdown_items,
            }
        ],
        "details": [
            {
                "key": "details.parts",
                "label": "零件包装明细",
                "meta": "来源：包装物流费用逐件明细。",
                "items": detail_items,
            }
        ],
    }


def align_items(quote_items: list[dict[str, Any]], fixed_items: list[dict[str, Any]]) -> dict[str, Any]:
    quote_map = {item["itemKey"]: item for item in quote_items}
    fixed_map = {item["itemKey"]: item for item in fixed_items}
    order = [item["itemKey"] for item in quote_items]
    for item in fixed_items:
        if item["itemKey"] not in quote_map:
            order.append(item["itemKey"])

    aligned = []
    summary = {
        "quoteCount": len(quote_items),
        "fixedCount": len(fixed_items),
        "matchedCount": 0,
        "quoteOnlyCount": 0,
        "fixedOnlyCount": 0,
        "differenceCount": 0,
    }

    for item_key in order:
        quote = quote_map.get(item_key)
        fixed = fixed_map.get(item_key)
        if quote and fixed:
            status = "matched"
            summary["matchedCount"] += 1
        elif quote:
            status = "quote_only"
            summary["quoteOnlyCount"] += 1
        else:
            status = "fixed_only"
            summary["fixedOnlyCount"] += 1

        if quote and fixed:
            quote_number = quote.get("numericValue")
            fixed_number = fixed.get("numericValue")
            if quote_number is not None and fixed_number is not None:
                if abs(float(quote_number) - float(fixed_number)) > 1e-6:
                    summary["differenceCount"] += 1
            elif quote.get("value") != fixed.get("value") or quote.get("note") != fixed.get("note"):
                summary["differenceCount"] += 1
        else:
            summary["differenceCount"] += 1

        aligned.append(
            {
                "status": status,
                "quote": quote,
                "fixed": fixed,
            }
        )

    return {
        "aligned": aligned,
        "summary": summary,
    }


def build_scope(scope_id: str, scope_label: str, hint: str, quote_groups: list[dict[str, Any]], fixed_groups: list[dict[str, Any]]) -> dict[str, Any]:
    fixed_map = {group["key"]: group for group in fixed_groups}
    order = [group["key"] for group in quote_groups]
    for group in fixed_groups:
        if group["key"] not in order:
            order.append(group["key"])

    groups = []
    scope_summary = {
        "quoteCount": 0,
        "fixedCount": 0,
        "matchedCount": 0,
        "quoteOnlyCount": 0,
        "fixedOnlyCount": 0,
        "differenceCount": 0,
    }

    for key in order:
        quote_group = next((group for group in quote_groups if group["key"] == key), None)
        fixed_group = fixed_map.get(key)
        quote_items = quote_group["items"] if quote_group else []
        fixed_items = fixed_group["items"] if fixed_group else []
        aligned = align_items(quote_items, fixed_items)
        groups.append(
            {
                "key": key,
                "label": quote_group["label"] if quote_group else fixed_group["label"],
                "meta": quote_group.get("meta") if quote_group else fixed_group.get("meta", ""),
                "aligned": aligned["aligned"],
                "summary": aligned["summary"],
            }
        )
        for summary_key in scope_summary:
            scope_summary[summary_key] += aligned["summary"][summary_key]

    return {
        "scopeId": scope_id,
        "scopeLabel": scope_label,
        "hint": hint,
        "groups": groups,
        "summary": scope_summary,
    }


def build_payload(quote_workbook: Path, fixed_workbook: Path) -> dict[str, Any]:
    quote_ctx = workbook_context(quote_workbook)
    fixed_ctx = workbook_context(fixed_workbook)
    quote_groups = workbook_groups(quote_ctx, "quote")
    fixed_groups = workbook_groups(fixed_ctx, "fixed")

    scope_order = ["scenario", "breakdown", "details"]
    scope_meta = {
        "scenario": {
            "label": "程序映射",
            "hint": "应用报价版 / 定点版后，会直接回填主页面的内外包装、运输费、仓储费、短驳/其他 4 个输入项。",
        },
        "breakdown": {
            "label": "单套拆分",
            "hint": "保留包装物流费用页单套成本拆分，以及项目评估/项目汇总对它的引用关系。",
        },
        "details": {
            "label": "明细对照",
            "hint": "逐个零件对照包装物流费用页中的包装、仓储和短驳成本。",
        },
    }

    comparisons = {
        scope_id: build_scope(
            scope_id,
            scope_meta[scope_id]["label"],
            scope_meta[scope_id]["hint"],
            quote_groups[scope_id],
            fixed_groups[scope_id],
        )
        for scope_id in scope_order
    }

    return {
        "meta": {
            "quoteWorkbook": quote_workbook.name,
            "fixedWorkbook": fixed_workbook.name,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "layer": "JSON 数据层",
        },
        "scopeOrder": scope_order,
        "versionSnapshots": {
            "quote": build_snapshot(quote_ctx, "quote"),
            "fixed": build_snapshot(fixed_ctx, "fixed"),
        },
        "comparisons": comparisons,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate packaging validation data from quote/fixed E281 workbooks.")
    parser.add_argument("--quote", type=Path, default=None, help="Quote workbook path")
    parser.add_argument("--fixed", type=Path, default=None, help="Fixed workbook path")
    parser.add_argument("--out", type=Path, default=Path("g281_data_packaging_validation.json"), help="Output JSON path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    quote_path = args.quote or discover_workbook("报价核算")
    fixed_path = args.fixed or discover_workbook("定点核算")
    payload = build_payload(quote_path, fixed_path)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(args.out)


if __name__ == "__main__":
    main()
