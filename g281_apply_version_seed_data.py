from __future__ import annotations

import json
from copy import deepcopy
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries


ROOT = Path(__file__).resolve().parent
MASTER_PATH = ROOT / "g281_data_master.json"
TMP_DIR = ROOT / "tmp"
BOM_DIR = ROOT / "BOM核对"


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value)


def find_xlsx_sheet(workbook, sheet_name: str):
    target = sheet_name.strip()
    for worksheet in workbook.worksheets:
        if worksheet.title.strip() == target:
            return worksheet
    raise KeyError(f"Sheet not found: {sheet_name}")


def find_xls_sheet(workbook, sheet_name: str):
    target = sheet_name.strip()
    for index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(index)
        if sheet.name.strip() == target:
            return sheet
    raise KeyError(f"Sheet not found: {sheet_name}")


def build_cell(row: int, column: int, value: Any = None, formula: str | None = None) -> dict[str, Any]:
    address = f"{get_column_letter(column)}{row}"
    payload: dict[str, Any] = {
        "address": address,
        "row": row,
        "column": column,
        "dataType": "f" if formula else None,
        "styleId": 0,
    }
    if formula:
        payload["formula"] = formula if formula.startswith("=") else f"={formula}"
    elif value not in (None, ""):
        payload["value"] = json_scalar(value)
    return payload


def build_sheet(
    sheet_name: str,
    max_row: int,
    max_column: int,
    cells: list[dict[str, Any]],
    merged_ranges: list[str] | None = None,
    row_dimensions: list[dict[str, Any]] | None = None,
    column_dimensions: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    return {
        "sheetName": sheet_name,
        "sheetState": "visible",
        "maxRow": max_row,
        "maxColumn": max_column,
        "freezePane": None,
        "mergedRanges": merged_ranges or [],
        "rowDimensions": row_dimensions or [],
        "columnDimensions": column_dimensions or [],
        "cells": [cell for cell in cells if cell.get("formula") or cell.get("value") not in (None, "")],
    }


def build_matrix_sheet(sheet_name: str, rows: list[list[Any]]) -> dict[str, Any]:
    max_row = len(rows)
    max_column = max((len(row) for row in rows), default=1)
    cells: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            if value in (None, ""):
                continue
            if isinstance(value, str) and value.startswith("="):
                cells.append(build_cell(row_index, column_index, formula=value))
            else:
                cells.append(build_cell(row_index, column_index, value=value))
    return build_sheet(sheet_name, max_row=max_row or 1, max_column=max_column or 1, cells=cells)


def build_sheet_from_cell_map(sheet_name: str, cells_map: dict[str, Any]) -> dict[str, Any]:
    max_row = 1
    max_column = 1
    cells: list[dict[str, Any]] = []
    for address, value in cells_map.items():
        column_letters = "".join(ch for ch in address if ch.isalpha())
        row_digits = "".join(ch for ch in address if ch.isdigit())
        if not column_letters or not row_digits:
            continue
        row = int(row_digits)
        column = column_index_from_string(column_letters)
        max_row = max(max_row, row)
        max_column = max(max_column, column)
        if isinstance(value, str) and value.startswith("="):
            cells.append(build_cell(row, column, formula=value))
        else:
            cells.append(build_cell(row, column, value=value))
    return build_sheet(sheet_name, max_row=max_row, max_column=max_column, cells=cells)


def extract_xlsx_sheet(
    workbook_path: Path,
    sheet_name: str,
    range_ref: str | None = None,
    max_rows: int | None = None,
    max_columns: int | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False)
    worksheet = find_xlsx_sheet(workbook, sheet_name)

    if range_ref:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    else:
        min_col, min_row = 1, 1
        max_col = max_columns or worksheet.max_column
        max_row = max_rows or worksheet.max_row

    max_col = min(max_col, max_columns) if max_columns else max_col
    max_row = min(max_row, max_rows) if max_rows else max_row

    cells: list[dict[str, Any]] = []
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cell = worksheet.cell(row, column)
            value = cell.value
            if value in (None, ""):
                continue
            if isinstance(value, str) and value.startswith("="):
                cells.append(build_cell(row, column, formula=value))
            else:
                cells.append(build_cell(row, column, value=value))

    row_dimensions: list[dict[str, Any]] = []
    for row in range(min_row, max_row + 1):
        dimension = worksheet.row_dimensions.get(row)
        if not dimension:
            continue
        entry: dict[str, Any] = {"row": row}
        if dimension.height:
            entry["height"] = float(dimension.height)
        if dimension.hidden:
            entry["hidden"] = True
        if len(entry) > 1:
            row_dimensions.append(entry)

    column_dimensions: list[dict[str, Any]] = []
    for column in range(min_col, max_col + 1):
        letter = get_column_letter(column)
        dimension = worksheet.column_dimensions.get(letter)
        if not dimension:
            continue
        entry: dict[str, Any] = {"min": column, "max": column}
        if dimension.width:
            entry["width"] = float(dimension.width)
        if dimension.hidden:
            entry["hidden"] = True
        if len(entry) > 2:
            column_dimensions.append(entry)

    merged_ranges: list[str] = []
    for merged in worksheet.merged_cells.ranges:
        bounds = range_boundaries(str(merged))
        if bounds[0] >= min_col and bounds[1] >= min_row and bounds[2] <= max_col and bounds[3] <= max_row:
            merged_ranges.append(str(merged))

    return build_sheet(
        worksheet.title,
        max_row=max_row,
        max_column=max_col,
        cells=cells,
        merged_ranges=merged_ranges,
        row_dimensions=row_dimensions,
        column_dimensions=column_dimensions,
    )


def extract_xls_sheet(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    workbook = xlrd.open_workbook(workbook_path, formatting_info=True)
    sheet = find_xls_sheet(workbook, sheet_name)
    cells: list[dict[str, Any]] = []
    for row in range(sheet.nrows):
        for column in range(sheet.ncols):
            value = sheet.cell_value(row, column)
            if value in ("", None):
                continue
            cells.append(build_cell(row + 1, column + 1, value=value))
    merged_ranges = [
        f"{get_column_letter(left + 1)}{top + 1}:{get_column_letter(right)}{bottom}"
        for top, bottom, left, right in sheet.merged_cells
    ]
    return build_sheet(
        sheet.name,
        max_row=sheet.nrows or 1,
        max_column=sheet.ncols or 1,
        cells=cells,
        merged_ranges=merged_ranges,
    )


def build_workbook_seed(workbook_name: str, source_file_name: str, sheets: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "workbookName": workbook_name,
        "sourceFileName": source_file_name,
        "sourcePath": source_file_name,
        "versionKey": "manual-seed",
        "versionLabel": workbook_name,
        "sheetOrder": [sheet["sheetName"] for sheet in sheets],
        "hiddenSheets": [],
        "styleTable": {},
        "sheets": sheets,
    }


def text_join(values: list[Any], separator: str = " / ") -> str:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
    return separator.join(ordered)


def connector_status_label(statuses: list[str]) -> str:
    normalized = [str(value or "").strip() for value in statuses if str(value or "").strip()]
    if not normalized:
        return ""
    if all(value == "confirmed" for value in normalized):
        return "已达成"
    if all(value == "quoted_pending" for value in normalized):
        return "待确认"
    if all(value == "dev_pending" for value in normalized):
        return "开发中"
    if all(value == "no_reply" for value in normalized):
        return "暂无回复"
    return "部分达成"


def connector_price_text(rows: list[dict[str, Any]], field_name: str) -> Any:
    values: list[float] = []
    for row in rows:
        value = row.get(field_name)
        if value in ("", None):
            continue
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            continue
        if not any(abs(numeric - existing) < 1e-9 for existing in values):
            values.append(numeric)
    if not values:
        return ""
    values.sort()
    if len(values) == 1:
        value = values[0]
        return int(value) if float(value).is_integer() else round(value, 4)
    return " / ".join(
        str(int(value)) if float(value).is_integer() else f"{value:.4f}".rstrip("0").rstrip(".")
        for value in values
    )


def connector_detail_text(rows: list[dict[str, Any]]) -> str:
    details: list[str] = []
    seen: set[str] = set()
    for row in rows:
        parts = [
            str(row.get("assemblyNo") or row.get("assemblyCode") or "").strip(),
            str(row.get("partNumber") or "").strip(),
            str(row.get("partName") or "").strip(),
            str(row.get("functionBrief") or "").strip(),
        ]
        text = " / ".join(part for part in parts if part)
        if not text or text in seen:
            continue
        seen.add(text)
        details.append(text)
    return "\n".join(details)


def connector_mark(statuses: list[str], target: str) -> str:
    return "Y" if any(str(value or "").strip() == target for value in statuses) else ""


def build_connector_detail_sheet(
    sheet_name: str,
    rows: list[dict[str, Any]],
    workbook_name: str,
) -> dict[str, Any]:
    matrix = [[
        "portfolioId",
        "groupLabel",
        "assemblyNo",
        "partNumber",
        "partName",
        "supplier",
        "protocolPrice",
        "progressPrice",
        "initialQuote",
        "status",
        "recommendedStage",
        "sourceWorkbook",
        "sourceSheet",
        "sourceRow",
        "detail",
    ]]
    for row in rows:
        matrix.append([
            row.get("portfolioId", ""),
            row.get("groupLabel", ""),
            row.get("assemblyNo") or row.get("assemblyCode") or "",
            row.get("partNumber", ""),
            row.get("partName", ""),
            row.get("supplierRaw") or row.get("supplier", ""),
            row.get("targetProtocolPrice", ""),
            row.get("replyPrice", ""),
            row.get("initialQuote", ""),
            connector_status_label([row.get("statusKey", "")]),
            row.get("recommendedStage", ""),
            workbook_name,
            (row.get("initialQuoteSourceMeta") or {}).get("sheet", ""),
            (row.get("protocolPriceSource") or {}).get("sourceRow", ""),
            text_join([
                row.get("functionRaw", ""),
                row.get("initialQuoteSource", ""),
            ], " | "),
        ])
    return build_matrix_sheet(sheet_name, matrix)


def connector_numeric_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return numeric if abs(numeric) > 1e-9 else None


def connector_version_row_price(row: dict[str, Any], version_key: str) -> float | None:
    ordered_fields = (
        ("initialQuote", "replyPrice", "targetProtocolPrice")
        if version_key == "quote"
        else ("replyPrice", "targetProtocolPrice", "initialQuote")
    )
    for field_name in ordered_fields:
        numeric = connector_numeric_value(row.get(field_name))
        if numeric is not None:
            return numeric
    return None


def connector_average_price(rows: list[dict[str, Any]], version_key: str) -> float | None:
    values = [
        numeric
        for numeric in (connector_version_row_price(row, version_key) for row in rows)
        if numeric is not None
    ]
    if not values:
        return None
    return sum(values) / len(values)


def quoted_sheet_ref(sheet_name: str, cell_ref: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!{cell_ref}"


def build_packaging_template_sheet(workbook_path: Path) -> dict[str, Any]:
    sheet = deepcopy(
        extract_xlsx_sheet(
            workbook_path,
            "包装物流费用",
            max_rows=70,
            max_columns=13,
        )
    )
    sheet["maxColumn"] = max(int(sheet.get("maxColumn") or 13), 17)
    column_dimensions = list(sheet.get("columnDimensions") or [])
    for column in range(14, 18):
        column_dimensions.append({"min": column, "max": column, "hidden": True})
    sheet["columnDimensions"] = column_dimensions
    cells = list(sheet.get("cells") or [])
    cells.extend([
        build_cell(25, 14, formula="=E25+F25"),
        build_cell(25, 15, formula="=G25+H25"),
        build_cell(25, 16, formula="=J25+K25"),
        build_cell(25, 17, formula="=I25"),
    ])
    sheet["cells"] = cells
    return sheet


def set_version(master: dict[str, Any], group: str, key: str, patch: dict[str, Any]) -> None:
    master.setdefault("versions", {}).setdefault(group, {})
    base = deepcopy(master["versions"][group].get(key, {}))
    base.update(patch)
    master["versions"][group][key] = base


def build_metal_versions(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_metal.json")
    for version_key in ("quote", "fixed"):
        item = seed["versions"][version_key]
        workbook_path = ROOT / item["workbook"]
        excerpt = extract_xlsx_sheet(workbook_path, "配置明细", range_ref="G24:K28")
        copper_value = round(float(item["copperPrice"]) * 1000, 3)
        aluminum_value = round(float(item["aluminumPrice"]) * 1000, 3)
        summary = build_matrix_sheet("版本录入", [
            ["字段", "值", "说明"],
            ["铜价(元/吨)", copper_value, f"源：配置明细!I26 = {item['copperPrice']} 元/kg"],
            ["铝价(元/吨)", aluminum_value, "源：配置明细!I28 公式常数 18.21 元/kg"],
        ])
        workbook_seed = build_workbook_seed(
            workbook_name=f"{item['label']}铜铝基价模板",
            source_file_name=item["workbook"],
            sheets=[summary, excerpt],
        )
        set_version(master, "metal", version_key, {
            "label": "报价版" if version_key == "quote" else "定点版",
            "copperPrice": copper_value,
            "aluminumPrice": aluminum_value,
            "sourceNote": item["sourceNote"],
            "templateSource": f"{item['workbook']} / 配置明细",
            "templateNote": seed["source"]["sourceNote"],
            "templateRawInputs": {
                "copperPrice": str(copper_value),
                "aluminumPrice": str(aluminum_value),
            },
            "templateFieldAddressMap": {
                "copperPrice": "B2",
                "aluminumPrice": "B3",
            },
            "templateWorkbookSeed": workbook_seed,
        })
        if version_key == "quote":
            master["copperPrice"] = copper_value
            master["aluminumPrice"] = aluminum_value


def build_labor_versions(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_labor.json")
    mapping = {"base": "报价版", "optimize": "定点版"}
    for version_key, label in mapping.items():
        item = seed["versions"][version_key]
        source_sheets = [
            build_sheet_from_cell_map(entry["sheetName"], entry["cells"])
            for entry in item["workbookSnapshotSeed"]["sheets"]
        ]
        summary = build_matrix_sheet("版本录入", [
            ["字段", "值", "来源"],
            ["直接人工工时(h/套)", item["directHours"], item["sources"]["directHours"]],
            ["直接人工费率(元/h)", item["directRate"], item["sources"]["directRate"]],
            ["制造工时(h/套)", item["manufacturingHours"], item["sources"]["manufacturingHours"]],
            ["制造费率(元/h)", item["manufacturingRate"], item["sources"]["manufacturingRate"]],
        ])
        workbook_seed = build_workbook_seed(
            workbook_name=f"{label}工时模板",
            source_file_name=item["workbookSnapshotSeed"]["workbookName"],
            sheets=[summary, *source_sheets],
        )
        set_version(master, "labor", version_key, {
            "label": label,
            "directHours": item["directHours"],
            "directRate": item["directRate"],
            "manufacturingHours": item["manufacturingHours"],
            "manufacturingRate": item["manufacturingRate"],
            "sourceNote": item["dataNote"],
            "templateSource": item["workbookSnapshotSeed"]["relativePath"],
            "templateNote": item["workbookSnapshotSeed"]["sourceNote"],
            "templateRawInputs": {
                "directHours": str(item["directHours"]),
                "directRate": str(item["directRate"]),
                "manufacturingHours": str(item["manufacturingHours"]),
                "manufacturingRate": str(item["manufacturingRate"]),
            },
            "templateFieldAddressMap": {
                "directHours": "B2",
                "directRate": "B3",
                "manufacturingHours": "B4",
                "manufacturingRate": "B5",
            },
            "templateWorkbookSeed": workbook_seed,
        })
        if version_key == "base":
            master["baseDirectHours"] = item["directHours"]
            master["baseDirectRate"] = item["directRate"]
            master["baseMfgHours"] = item["manufacturingHours"]
            master["baseMfgRate"] = item["manufacturingRate"]
            master["baseLaborPerSet"] = item["directHours"] * item["directRate"]
            master["baseMfgPerSet"] = item["manufacturingHours"] * item["manufacturingRate"]


def build_packaging_versions(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_packaging.json")
    mapping = {
        "base": "吉利E281报价核算.xlsx",
        "optimize": "吉利E281定点核算.xlsx",
    }
    for version_key, workbook_name in mapping.items():
        item = seed["versions"][version_key]
        packaging_sheet = build_packaging_template_sheet(BOM_DIR / workbook_name)
        workbook_seed = build_workbook_seed(
            workbook_name=f"{item['label']}包装物流模板",
            source_file_name=workbook_name,
            sheets=[packaging_sheet],
        )
        set_version(master, "packaging", version_key, {
            "label": item["label"],
            "packInner": item["packInner"],
            "packFreight": item["packFreight"],
            "packWarehouse": item["packWarehouse"],
            "packOther": item["packOther"],
            "sourceNote": "",
            "templateSource": "",
            "templateNote": "",
            "templateRawInputs": {
                "packInner": "=E25+F25",
                "packFreight": "=G25+H25",
                "packWarehouse": "=J25+K25",
                "packOther": "=I25",
            },
            "templateFieldAddressMap": {
                "packInner": "N25",
                "packFreight": "O25",
                "packWarehouse": "P25",
                "packOther": "Q25",
            },
            "templateWorkbookSeed": workbook_seed,
        })
        if version_key == "base":
            master["basePackagingPerSet"] = (
                float(item["packInner"])
                + float(item["packFreight"])
                + float(item["packWarehouse"])
                + float(item["packOther"])
            )


def build_equipment_versions(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_equipment.json")
    mapping = {"base": "报价版", "shared": "定点版"}
    for version_key, label in mapping.items():
        item = seed["versions"][version_key]
        workbook_name = item["workbook"]
        workbook_path = BOM_DIR / workbook_name
        equipment_amount = item["scopes"]["equipment"]["summary"]["totalNewAmount"]
        tooling_amount = item["scopes"]["tooling"]["summary"]["totalNewAmount"]
        fixtures_amount = item["scopes"]["fixtures"]["summary"]["totalNewAmount"]
        rnd_amount = item["scopes"]["rnd"]["summary"]["totalNewAmount"]
        summary = build_matrix_sheet("版本录入", [
            ["字段", "值", "说明"],
            ["设备投资(元)", equipment_amount, "源：设备投资明细汇总"],
            ["专用模具(元)", tooling_amount, "源：项目专用模具汇总"],
            ["项目工装(元)", fixtures_amount, "源：项目工装投入汇总"],
            ["研发费用(元)", rnd_amount, "源：研发费用汇总"],
        ])
        source_sheets = [
            extract_xlsx_sheet(workbook_path, sheet_name)
            for sheet_name in ("设备投资明细", "项目专用模具", "项目工装投入 ", "研发费用 ")
        ]
        workbook_seed = build_workbook_seed(
            workbook_name=f"{label}资源投入模板",
            source_file_name=workbook_name,
            sheets=[summary, *source_sheets],
        )
        set_version(master, "equipment", version_key, {
            "label": label,
            "equipment": equipment_amount,
            "tooling": tooling_amount,
            "fixtures": fixtures_amount,
            "rnd": rnd_amount,
            "sourceNote": seed["source"]["sourceNote"],
            "templateSource": workbook_name,
            "templateNote": "保留设备投资明细、项目专用模具、项目工装投入、研发费用四张源表，首个 Sheet 为可编辑汇总。",
            "templateRawInputs": {
                "equipment": str(equipment_amount),
                "tooling": str(tooling_amount),
                "fixtures": str(fixtures_amount),
                "rnd": str(rnd_amount),
            },
            "templateFieldAddressMap": {
                "equipment": "B2",
                "tooling": "B3",
                "fixtures": "B4",
                "rnd": "B5",
            },
            "templateWorkbookSeed": workbook_seed,
        })
        if version_key == "base":
            master["capital"] = {
                "equipment": equipment_amount,
                "tooling": tooling_amount,
                "fixtures": fixtures_amount,
                "rnd": rnd_amount,
            }


def build_onetime_versions(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_onetime_customer.json")
    workbook_name = seed["source"]["fixedWorkbook"]
    workbook_path = BOM_DIR / workbook_name
    source_sheets = [extract_xls_sheet(workbook_path, name) for name in seed["source"]["sheetNames"]]
    for version_key, label, source_note in (
        ("quote", "报价版", "未提供报价版一次性费用源表，沿用定点模板结构并保留 0 元默认值。"),
        ("fixed", "定点版", seed["versions"]["fixed"]["sourceNote"]),
    ):
        amount_total = 0.0 if version_key == "quote" else float(seed["versions"]["fixed"]["amountTotal"])
        summary = build_matrix_sheet("版本录入", [
            ["字段", "值", "说明"],
            ["客户支付总额(元)", amount_total, source_note],
        ])
        workbook_seed = build_workbook_seed(
            workbook_name=f"{label}一次性费用模板",
            source_file_name=workbook_name,
            sheets=[summary, *source_sheets],
        )
        set_version(master, "oneTimeCustomer", version_key, {
            "label": label,
            "amountTotal": amount_total,
            "sourceNote": source_note,
            "templateSource": workbook_name,
            "templateNote": "保留 4.1工装费 / 4.2.试验费 / 4.3.研发费 三个 Sheet，首个 Sheet 为可编辑汇总。",
            "templateRawInputs": {
                "amountTotal": str(amount_total),
            },
            "templateFieldAddressMap": {
                "amountTotal": "B2",
            },
            "templateWorkbookSeed": workbook_seed,
        })


def build_connector_meta(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_connector.json")
    source_text = "报价/定点核算《二次物料明细》《总成散件清单》《KSK线束BOM明细》"
    note_text = "连接器价格执行仍按 batch/protocol/sample 档位管理；quote/fixed 来源明细已抽取到 tmp/version_seed_connector.json 供后续继续扩展模板来源。"
    for version_key in ("batch", "protocol", "sample"):
        set_version(master, "connector", version_key, {
            "templateSource": source_text,
            "templateNote": note_text,
            "sourceNote": f"{source_text}；协议/进度价来自 g281_data_connector_protocol_status.json。",
            "seedSummary": {
                "quoteRows": len(seed.get("quote", {}).get("rows", [])),
                "fixedRows": len(seed.get("fixed", {}).get("rows", [])),
                "unmatchedPortfolios": seed.get("suggestedIntegration", {}).get("mismatch", {}),
            },
        })


def build_connector_meta_v2(master: dict[str, Any]) -> None:
    seed = load_json(TMP_DIR / "version_seed_connector.json")
    source_text = "报价/定点核算《二次物料明细》《总成散件清单》《KSK线束BOM明细》"
    note_text = "首个 Sheet 为可编辑汇总，后附报价版与定点版连接器明细；版本管理统一按报价版 / 定点版执行，逐项可临时切换为进度价。"
    quote_seed = seed.get("quote", {})
    fixed_seed = seed.get("fixed", {})
    quote_rows = list(quote_seed.get("rows", []))
    fixed_rows = list(fixed_seed.get("rows", []))
    quote_by_portfolio: dict[str, list[dict[str, Any]]] = {}
    fixed_by_portfolio: dict[str, list[dict[str, Any]]] = {}

    for row in quote_rows:
        quote_by_portfolio.setdefault(str(row.get("portfolioId") or "").strip(), []).append(row)
    for row in fixed_rows:
        fixed_by_portfolio.setdefault(str(row.get("portfolioId") or "").strip(), []).append(row)

    connector_items = list(master.get("connectorPortfolio", {}).get("items", []))
    portfolio_base_cost_per_set = float(master.get("connectorPortfolio", {}).get("baseCostPerSet") or 0)
    summary_rows = [[
        "序号",
        "连接器ID",
        "总成号",
        "名称",
        "供应商",
        "报价版协议价",
        "报价版进度价",
        "报价版初始报价",
        "报价版状态",
        "定点版协议价",
        "定点版进度价",
        "定点版初始报价",
        "定点版状态",
        "执行档位",
        "报价版已达成",
        "报价版待确认",
        "报价版开发中",
        "定点版已达成",
        "定点版待确认",
        "定点版开发中",
        "报价版散件明细",
        "定点版散件明细",
    ]]
    field_address_map: dict[str, str] = {}
    quote_template_inputs: dict[str, str] = {}
    fixed_template_inputs: dict[str, str] = {}
    quote_total_base_cost = 0.0
    fixed_total_cost = 0.0

    for index, item in enumerate(connector_items, start=1):
        portfolio_id = str(item.get("id") or "").strip()
        quote_item_rows = quote_by_portfolio.get(portfolio_id, [])
        fixed_item_rows = fixed_by_portfolio.get(portfolio_id, [])
        quote_statuses = [str(row.get("statusKey") or "").strip() for row in quote_item_rows]
        fixed_statuses = [str(row.get("statusKey") or "").strip() for row in fixed_item_rows]
        row_number = index + 1
        field_address_map[f"connector_stage__{portfolio_id}"] = f"N{row_number}"
        quote_template_inputs[f"connector_stage__{portfolio_id}"] = "报价版"
        fixed_template_inputs[f"connector_stage__{portfolio_id}"] = "定点版"

        share = float(item.get("share") or 0)
        base_cost = float(item.get("baseCost") or 0) or portfolio_base_cost_per_set * share
        quote_average = connector_average_price(quote_item_rows, "quote")
        fixed_average = connector_average_price(fixed_item_rows, "fixed")
        if quote_average is None:
            quote_average = connector_average_price(fixed_item_rows, "quote")
        if fixed_average is None:
            fixed_average = quote_average
        if quote_average is None:
            quote_average = 0.0
        if fixed_average is None:
            fixed_average = quote_average
        factor = fixed_average / quote_average if quote_average > 0 else 1.0
        if base_cost > 0:
            quote_total_base_cost += base_cost
            fixed_total_cost += base_cost * factor

        summary_rows.append([
            index,
            portfolio_id,
            text_join([
                text_join([row.get("assemblyNo") or row.get("assemblyCode") or "" for row in quote_item_rows], "\n"),
                text_join([row.get("assemblyNo") or row.get("assemblyCode") or "" for row in fixed_item_rows], "\n"),
                item.get("code", ""),
            ], "\n"),
            item.get("name", "") or item.get("label", "") or portfolio_id,
            text_join([
                text_join([row.get("supplierRaw") or row.get("supplier", "") for row in quote_item_rows]),
                text_join([row.get("supplierRaw") or row.get("supplier", "") for row in fixed_item_rows]),
                item.get("supplier", ""),
            ], " / "),
            connector_price_text(quote_item_rows, "targetProtocolPrice"),
            connector_price_text(quote_item_rows, "replyPrice"),
            connector_price_text(quote_item_rows, "initialQuote"),
            connector_status_label(quote_statuses),
            connector_price_text(fixed_item_rows, "targetProtocolPrice"),
            connector_price_text(fixed_item_rows, "replyPrice"),
            connector_price_text(fixed_item_rows, "initialQuote"),
            connector_status_label(fixed_statuses),
            "",
            connector_mark(quote_statuses, "confirmed"),
            connector_mark(quote_statuses, "quoted_pending"),
            connector_mark(quote_statuses, "dev_pending"),
            connector_mark(fixed_statuses, "confirmed"),
            connector_mark(fixed_statuses, "quoted_pending"),
            connector_mark(fixed_statuses, "dev_pending"),
            connector_detail_text(quote_item_rows),
            connector_detail_text(fixed_item_rows),
        ])

    workbook_seed = build_workbook_seed(
        workbook_name="连接器价格录入模板",
        source_file_name=f"{quote_seed.get('workbook', '')} / {fixed_seed.get('workbook', '')}",
        sheets=[
            build_matrix_sheet("版本录入", summary_rows),
            build_connector_detail_sheet("报价版连接器明细", quote_rows, str(quote_seed.get("workbook", ""))),
            build_connector_detail_sheet("定点版连接器明细", fixed_rows, str(fixed_seed.get("workbook", ""))),
        ],
    )

    fixed_factor = fixed_total_cost / quote_total_base_cost if quote_total_base_cost > 0 else 1.0
    mismatch_summary = seed.get("suggestedIntegration", {}).get("mismatch", {})
    master.setdefault("versions", {})["connector"] = {}

    for version_key, label, factor, template_inputs in (
        ("quote", "报价版", 1.0, quote_template_inputs),
        ("fixed", "定点版", fixed_factor, fixed_template_inputs),
    ):
        set_version(master, "connector", version_key, {
            "label": label,
            "factor": round(float(factor), 6),
            "sourceKey": version_key,
            "workbookVersionKeyFallback": version_key,
            "note": "连接器价格版本按核算表录入，可逐项临时切换为进度价。",
            "templateSource": source_text,
            "templateNote": note_text,
            "sourceNote": f"{source_text}；协议/进度价来自 g281_data_connector_protocol_status.json。",
            "seedSummary": {
                "quoteRows": len(quote_rows),
                "fixedRows": len(fixed_rows),
                "unmatchedPortfolios": mismatch_summary,
                "quoteBaseCostPerSet": round(quote_total_base_cost, 6),
                "fixedCostPerSet": round(fixed_total_cost, 6),
            },
            "templateRawInputs": template_inputs,
            "templateFieldAddressMap": field_address_map,
            "templateWorkbookSeed": workbook_seed,
        })


def main() -> None:
    master = load_json(MASTER_PATH)
    build_metal_versions(master)
    build_labor_versions(master)
    build_packaging_versions(master)
    build_equipment_versions(master)
    build_onetime_versions(master)
    build_connector_meta_v2(master)
    save_json(MASTER_PATH, master)
    print(MASTER_PATH)


if __name__ == "__main__":
    main()
