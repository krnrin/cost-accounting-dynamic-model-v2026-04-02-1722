# -*- coding: utf-8 -*-
"""
Extract cost data per set (单套成本) from 吉利E281报价核算.xlsx
Sheet: 项目评估汇总（昆山90%）
"""
import openpyxl
from openpyxl import load_workbook
import sys
import io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load the workbook
file_path = r'C:\Users\lyvee\.openclaw\workspace\成本核算动态模型\BOM核对\吉利E281报价核算.xlsx'
wb = load_workbook(file_path, data_only=True)

# Find the target sheet
target_sheet = None
for sheet_name in wb.sheetnames:
    if '项目评估汇总' in sheet_name or '昆山' in sheet_name:
        target_sheet = wb[sheet_name]
        break

if not target_sheet:
    target_sheet = wb[wb.sheetnames[0]]

print("="*80)
print("吉利E281 报价核算 - 单套成本数据提取")
print(f"工作表: {target_sheet.title}")
print("="*80)
print()

# Define the cost items to extract with their row numbers and descriptions
cost_items = [
    {'row': 15, 'name_cn': '材料费', 'name_en': 'Material Cost', 'desc': '单套材料成本'},
    {'row': 16, 'name_cn': '直接人工', 'name_en': 'Direct Labor', 'desc': '单套直接人工'},
    {'row': 20, 'name_cn': '设备摊销', 'name_en': 'Equipment Amortization', 'desc': '单套设备成本'},
    {'row': 23, 'name_cn': '制造费用', 'name_en': 'Manufacturing Overhead', 'desc': '制造费用(单套)'},
    {'row': 32, 'name_cn': '包装物流', 'name_en': 'Packaging & Logistics', 'desc': '单套包装、物流费用'},
    {'row': 31, 'name_cn': '研发费用', 'name_en': 'R&D Cost', 'desc': '研发费用'},
    {'row': 14, 'name_cn': '总成本', 'name_en': 'Total Cost', 'desc': '单套项目总成本'},
]

print(f"{'Row':<6} {'Cost Item (CN)':<12} {'Cost Item (EN)':<28} {'Value (Total)':>18} {'Value per PCS':>18}")
print("-"*90)

results = []
for item in cost_items:
    row_idx = item['row']
    row_data = list(target_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

    # Column E (index 4) contains the total value
    # Column F (index 5) contains the per PCS value for first year
    total_value = row_data[4] if row_data[4] is not None else 0
    per_pcs_value = row_data[5] if row_data[5] is not None else 0

    results.append({
        'row': row_idx,
        'name_cn': item['name_cn'],
        'name_en': item['name_en'],
        'total': total_value,
        'per_pcs': per_pcs_value
    })

    print(f"{row_idx:<6} {item['name_cn']:<12} {item['name_en']:<28} {total_value:>18,.2f} {per_pcs_value:>18.2f}")

print()
print("="*80)
print("数据说明:")
print("  - Value (Total): 项目周期汇总值 (Column E)")
print("  - Value per PCS: 单套成本 - 第一年2026年 (Column F)")
print("="*80)

# Also show the breakdown of manufacturing overhead
print()
print("="*80)
print("制造费用明细 (Manufacturing Overhead Breakdown)")
print("="*80)

overhead_items = [
    {'row': 24, 'name': '间接人工'},
    {'row': 25, 'name': '低值易耗品'},
    {'row': 26, 'name': '机物料消耗'},
    {'row': 27, 'name': '厂房分摊'},
    {'row': 28, 'name': '自动化仓、仓库分摊'},
    {'row': 29, 'name': '其他制费'},
    {'row': 30, 'name': '材料损耗'},
]

print(f"{'Row':<6} {'Item':<20} {'Total':>18} {'Per PCS':>18}")
print("-"*60)

for item in overhead_items:
    row_idx = item['row']
    row_data = list(target_sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    total_value = row_data[4] if row_data[4] is not None else 0
    per_pcs_value = row_data[5] if row_data[5] is not None else 0
    print(f"{row_idx:<6} {item['name']:<20} {total_value:>18,.2f} {per_pcs_value:>18.2f}")

wb.close()
