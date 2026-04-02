import openpyxl
from openpyxl import load_workbook

# Load the workbook
file_path = r'C:\Users\lyvee\.openclaw\workspace\成本核算动态模型\BOM核对\吉利E281报价核算.xlsx'
wb = load_workbook(file_path, data_only=True)

# List all sheet names
print("Available sheets:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Find the target sheet (项目评估汇总)
target_sheet = None
for sheet_name in wb.sheetnames:
    if '项目评估汇总' in sheet_name or '昆山' in sheet_name:
        target_sheet = wb[sheet_name]
        print(f"\nFound target sheet: '{sheet_name}'")
        break

if not target_sheet:
    print("\nTarget sheet not found with '项目评估汇总' or '昆山' in name.")
    print("Using first sheet instead...")
    target_sheet = wb[wb.sheetnames[0]]
    print(f"Using sheet: '{target_sheet.title}'")

# Display the first 30 rows to understand the structure
print(f"\n--- First 30 rows of sheet '{target_sheet.title}' ---")
for row_idx, row in enumerate(target_sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    print(f"Row {row_idx:2}: {row}")

print("\n" + "="*80)
print("Searching for cost data...")
print("="*80)

# Search for keywords and extract data
cost_data = {}
keywords = {
    '材料费': 'material cost',
    '直接人工': 'direct labor',
    '制造费用': 'manufacturing overhead',
    '设备摊销': 'equipment amortization',
    '包装物流': 'packaging and logistics',
    '研发费用': 'R&D cost',
    '总成本': 'total cost'
}

# Scan all rows for keywords
for row_idx, row in enumerate(target_sheet.iter_rows(min_row=1, max_row=100, values_only=True), 1):
    for cell in row:
        if cell and isinstance(cell, str):
            for keyword, english_name in keywords.items():
                if keyword in cell:
                    # Found a keyword, extract the row data
                    print(f"\nRow {row_idx}: Found '{keyword}' ({english_name})")
                    print(f"  Full row data: {row}")
                    # Find the numeric value in the same row (usually in a later column)
                    for col_idx, val in enumerate(row, 1):
                        if val and (isinstance(val, (int, float))):
                            print(f"  -> Value at column {col_idx}: {val}")
                            cost_data[keyword] = {
                                'row': row_idx,
                                'value': val,
                                'english': english_name
                            }
                            break

# Display summary
print("\n" + "="*80)
print("SUMMARY - Cost Data Per Set (单套成本)")
print("="*80)
print(f"{'Item':<20} {'Row':<6} {'Value':<15}")
print("-"*45)
for keyword, data in cost_data.items():
    print(f"{data['english']:<20} {data['row']:<6} {data['value']:<15}")

wb.close()
