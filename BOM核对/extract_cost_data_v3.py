# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook
import sys
import io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load the workbook
file_path = r'C:\Users\lyvee\.openclaw\workspace\成本核算动态模型\BOM核对\吉利E281报价核算.xlsx'
wb = load_workbook(file_path, data_only=True)

# List all sheet names
print("Available sheets:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Find the target sheet (项目评估汇总)
target_sheet = None
for sheet_name in wb.sheetnames:
    if '项目评估汇总' in sheet_name or '昆山' in sheet_name:
        target_sheet = wb[sheet_name]
        print(f"\nFound target sheet: '{sheet_name}'")
        break

if not target_sheet:
    print("\nTarget sheet not found with '项目评估汇总' or '昆山' in name.")
    print("Using first sheet instead...")
    target_sheet = wb[wb.sheetnames[0]]
    print(f"Using sheet: '{target_sheet.title}'")

# Display the first 50 rows to understand the structure
print(f"\n--- First 50 rows of sheet '{target_sheet.title}' ---")
for row_idx, row in enumerate(target_sheet.iter_rows(min_row=1, max_row=50, values_only=True), 1):
    # Only print non-empty rows or rows with relevant data
    has_data = any(cell for cell in row if cell is not None)
    if has_data:
        print(f"Row {row_idx:2}: {row}")

print("\n" + "="*80)
print("Searching for cost data keywords...")
print("="*80)

# Search for keywords and extract data
cost_data = {}
keywords_map = {
    '材料费': 'material cost',
    '直接人工': 'direct labor',
    '制造费用': 'manufacturing overhead',
    '设备摊销': 'equipment amortization',
    '包装物流': 'packaging and logistics',
    '研发费用': 'R&D cost',
    '总成本': 'total cost',
    '单套': 'per set'
}

# Scan all rows for keywords
for row_idx, row in enumerate(target_sheet.iter_rows(min_row=1, max_row=100, values_only=True), 1):
    for cell in row:
        if cell and isinstance(cell, str):
            for keyword, english_name in keywords_map.items():
                if keyword in cell:
                    # Found a keyword, extract the row data
                    print(f"\nRow {row_idx}: Found '{keyword}' ({english_name})")
                    print(f"  Full row: {row}")
                    # Find the first numeric value in the row (usually the cost value)
                    for col_idx, val in enumerate(row, 1):
                        if val is not None and isinstance(val, (int, float)) and not isinstance(val, bool):
                            print(f"  -> Numeric value at column {col_idx}: {val}")
                            if keyword not in cost_data:
                                cost_data[keyword] = []
                            cost_data[keyword].append({
                                'row': row_idx,
                                'col': col_idx,
                                'value': val,
                                'english': english_name,
                                'full_cell': cell
                            })
                            break  # Only take first numeric value per row

# Display summary
print("\n" + "="*80)
print("SUMMARY - Cost Data Per Set (单套成本)")
print("="*80)
print(f"{'Keyword':<15} {'English':<25} {'Row':<6} {'Value':<20}")
print("-"*70)
for keyword, entries in cost_data.items():
    for entry in entries:
        print(f"{keyword:<15} {entry['english']:<25} {entry['row']:<6} {entry['value']:<20}")

wb.close()
