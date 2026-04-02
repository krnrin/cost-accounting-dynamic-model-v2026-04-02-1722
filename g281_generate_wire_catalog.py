from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUTPUT_PATH = ROOT / "g281_data_wire_catalog.json"
SEARCH_DIR = ROOT / "output" / "spreadsheet"
MASTER_PATH = ROOT / "g281_data_master.json"
LINKED_DETAIL_SHEET = "二次物料明细"
WIRE_PRICE_SHEET = "导线价格"
TZ = timezone(timedelta(hours=8))
CODE_SIZE_RE = re.compile(r"/(\d+(?:\.\d+)?)(?:/|$)")
NAME_SIZE_RE = re.compile(r"(\d+(?:\.\d+)?)mm", re.IGNORECASE)
TT_REMARK_DEFAULT = "TT 导入"


def discover_linked_workbook(*keywords: str) -> Path:
    matches = []
    for path in SEARCH_DIR.glob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if all(keyword in path.name for keyword in keywords):
            matches.append(path)
    if not matches:
        keyword_text = " + ".join(keywords)
        raise FileNotFoundError(f"Workbook containing '{keyword_text}' not found under {SEARCH_DIR}.")
    return sorted(matches, key=lambda item: (item.stat().st_mtime, item.name))[-1]


def discover_tt_workbook() -> Path:
    candidates = [path for path in ROOT.glob("*TT*.xlsx") if not path.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError(f"TT workbook not found under {ROOT}.")
    preferred = [path for path in candidates if "回填" in path.name]
    pool = preferred or candidates
    return sorted(pool, key=lambda item: (item.stat().st_mtime, item.name))[-1]


def to_text(value: Any) -> str:
    return str(value or "").strip()


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def round_or_none(value: float | None, digits: int = 6) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def merge_text(left: str, right: str) -> str:
    values: list[str] = []
    for raw in (left, right):
        text = to_text(raw)
        if text and text not in values:
            values.append(text)
    return " / ".join(values)


def merge_many(*values: Any) -> str:
    merged = ""
    for value in values:
        merged = merge_text(merged, to_text(value))
    return merged


def strip_tt_suffix(code: str) -> str:
    return re.sub(r"/AL\d+$", "", to_text(code), flags=re.IGNORECASE)


def code_key(code: str) -> str:
    return strip_tt_suffix(code).upper()


def name_key(name: str) -> str:
    return re.sub(r"\s+", "", to_text(name))


def parse_section(code: str, name: str) -> float | None:
    code_match = CODE_SIZE_RE.search(to_text(code))
    if code_match:
        return float(code_match.group(1))
    name_match = NAME_SIZE_RE.search(to_text(name))
    if name_match:
        return float(name_match.group(1))
    return None


def family_key(code: str) -> str:
    normalized = code_key(code)
    if normalized.startswith("FLRY-B/0.5/"):
        return "FLRY-B/0.5"
    match = re.match(r"^(.+?)/\d+(?:\.\d+)?(?:/.*)?$", normalized)
    return match.group(1) if match else normalized


def material_family(aluminum_weight: float, copper_weight: float) -> str:
    has_aluminum = aluminum_weight > 0
    has_copper = copper_weight > 0
    if has_aluminum and has_copper:
        return "铜铝混合"
    if has_aluminum:
        return "铝基"
    if has_copper:
        return "铜基"
    return "非铜"


def material_family_from_name(name: str) -> str:
    text = to_text(name)
    has_aluminum = "铝" in text
    has_copper = "铜" in text
    if has_aluminum and has_copper:
        return "铜铝混合"
    if has_aluminum:
        return "铝基"
    if has_copper:
        return "铜基"
    return "非铜"


def relation_label(aluminum_weight: float, copper_weight: float, non_copper_cost: float) -> str:
    tags = []
    if aluminum_weight > 0:
        tags.append("铝基")
    if copper_weight > 0:
        tags.append("铜基")
    if non_copper_cost > 0:
        tags.append("非铜")
    return "+".join(tags) if tags else "待补"


def calculate_sheet_unit_price(
    aluminum_weight: float,
    copper_weight: float,
    non_copper_cost: float,
    aluminum_price: float,
    copper_price: float,
) -> float:
    return (aluminum_weight * aluminum_price / 1000.0) + (copper_weight * copper_price / 1000.0) + (non_copper_cost / 1000.0)


def looks_like_wire_record(code: str, name: str, unit: str) -> bool:
    normalized_code = code_key(code)
    normalized_unit = to_text(unit).upper()
    if normalized_code.startswith(("FHL", "FLRY")):
        return True
    return normalized_unit == "M" and "导线" in to_text(name)


def load_linked_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    detail_ws = workbook[LINKED_DETAIL_SHEET]
    wire_price_ws = workbook[WIRE_PRICE_SHEET]

    aluminum_base = to_float(wire_price_ws["B3"].value) or 0.0
    copper_base = to_float(wire_price_ws["B4"].value) or 0.0
    source_name = to_text(wire_price_ws["B2"].value) or path.name

    models: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row_index in range(3, detail_ws.max_row + 1):
        code = to_text(detail_ws.cell(row_index, 1).value)
        if not code:
            continue

        aluminum_weight = to_float(detail_ws.cell(row_index, 10).value) or 0.0
        copper_weight = to_float(detail_ws.cell(row_index, 11).value) or 0.0
        non_copper_cost = to_float(detail_ws.cell(row_index, 12).value) or 0.0
        if not any((aluminum_weight, copper_weight, non_copper_cost)):
            continue

        record = models.setdefault(
            code,
            {
                "code": code,
                "codeKey": code_key(code),
                "familyKey": family_key(code),
                "section": parse_section(code, detail_ws.cell(row_index, 2).value),
                "name": to_text(detail_ws.cell(row_index, 2).value),
                "nameKey": name_key(detail_ws.cell(row_index, 2).value),
                "unit": to_text(detail_ws.cell(row_index, 4).value) or "M",
                "supplier": to_text(detail_ws.cell(row_index, 5).value),
                "sap": to_text(detail_ws.cell(row_index, 6).value),
                "remark": to_text(detail_ws.cell(row_index, 7).value),
                "owner": to_text(detail_ws.cell(row_index, 8).value),
                "currency": to_text(detail_ws.cell(row_index, 13).value) or "CNY",
                "priceType": to_text(detail_ws.cell(row_index, 14).value),
                "usageQty": 0.0,
                "aluminumWeight": aluminum_weight,
                "copperWeight": copper_weight,
                "nonCopperCost": non_copper_cost,
                "sourceRows": [],
            },
        )

        record["usageQty"] += to_float(detail_ws.cell(row_index, 3).value) or 0.0
        record["supplier"] = merge_text(record["supplier"], detail_ws.cell(row_index, 5).value)
        record["sap"] = merge_text(record["sap"], detail_ws.cell(row_index, 6).value)
        record["remark"] = merge_text(record["remark"], detail_ws.cell(row_index, 7).value)
        record["owner"] = merge_text(record["owner"], detail_ws.cell(row_index, 8).value)
        record["priceType"] = merge_text(record["priceType"], detail_ws.cell(row_index, 14).value)
        if row_index not in record["sourceRows"]:
            record["sourceRows"].append(row_index)

    return {
        "workbook": path.name,
        "sourceWorkbook": source_name,
        "aluminumBasePrice": round(aluminum_base, 6),
        "copperBasePrice": round(copper_base, 6),
        "models": models,
    }


def load_tt_snapshot(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    detail_ws = workbook[workbook.sheetnames[1]]

    models: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row_index in range(3, detail_ws.max_row + 1):
        code = to_text(detail_ws.cell(row_index, 1).value)
        name = to_text(detail_ws.cell(row_index, 2).value)
        unit = to_text(detail_ws.cell(row_index, 4).value) or "M"
        if not code or not looks_like_wire_record(code, name, unit):
            continue

        record = models.setdefault(
            code,
            {
                "code": code,
                "codeKey": code_key(code),
                "familyKey": family_key(code),
                "section": parse_section(code, name),
                "name": name,
                "nameKey": name_key(name),
                "unit": unit,
                "supplier": to_text(detail_ws.cell(row_index, 8).value),
                "sap": to_text(detail_ws.cell(row_index, 10).value),
                "remark": to_text(detail_ws.cell(row_index, 11).value) or TT_REMARK_DEFAULT,
                "owner": "",
                "currency": "CNY",
                "priceType": TT_REMARK_DEFAULT,
                "usageQty": 0.0,
                "sourceRows": [],
            },
        )
        record["usageQty"] += to_float(detail_ws.cell(row_index, 3).value) or 0.0
        record["supplier"] = merge_text(record["supplier"], detail_ws.cell(row_index, 8).value)
        record["sap"] = merge_text(record["sap"], detail_ws.cell(row_index, 10).value)
        record["remark"] = merge_text(record["remark"], detail_ws.cell(row_index, 11).value or TT_REMARK_DEFAULT)
        if row_index not in record["sourceRows"]:
            record["sourceRows"].append(row_index)

    return {
        "workbook": path.name,
        "sourceWorkbook": path.name,
        "models": models,
    }


def build_template_indexes(*snapshots: dict[str, Any]) -> dict[str, Any]:
    exact_code: dict[str, dict[str, Any]] = {}
    normalized_code: dict[str, dict[str, Any]] = {}
    sap_map: dict[str, dict[str, Any]] = {}
    name_map: dict[str, dict[str, Any]] = {}
    family_section_map: dict[tuple[str, float], dict[str, Any]] = {}
    family_map: dict[str, list[dict[str, Any]]] = {}

    for snapshot in snapshots:
        for row in snapshot["models"].values():
            exact_code.setdefault(row["code"], row)
            normalized_code.setdefault(row["codeKey"], row)
            if row["sap"]:
                sap_map.setdefault(row["sap"], row)
            if row["nameKey"]:
                name_map.setdefault(row["nameKey"], row)
            section = row.get("section")
            if row["familyKey"] and section is not None:
                family_section_map.setdefault((row["familyKey"], float(section)), row)
                family_map.setdefault(row["familyKey"], []).append(row)

    for rows in family_map.values():
        rows.sort(key=lambda item: (float(item.get("section") or 0), item["code"]))

    return {
        "exactCode": exact_code,
        "normalizedCode": normalized_code,
        "sap": sap_map,
        "name": name_map,
        "familySection": family_section_map,
        "familyMap": family_map,
    }


def scaled_weights(source_row: dict[str, Any], ratio: float) -> dict[str, float]:
    return {
        "aluminum": round(float(source_row["aluminumWeight"] or 0.0) * ratio, 6),
        "copper": round(float(source_row["copperWeight"] or 0.0) * ratio, 6),
        "nonCopper": round(float(source_row["nonCopperCost"] or 0.0) * ratio, 6),
    }


def direct_weight_payload(method: str, label: str, note: str, source_row: dict[str, Any]) -> dict[str, Any]:
    return {
        "weights": {
            "aluminum": round_or_none(float(source_row["aluminumWeight"] or 0.0)),
            "copper": round_or_none(float(source_row["copperWeight"] or 0.0)),
            "nonCopper": round_or_none(float(source_row["nonCopperCost"] or 0.0)),
        },
        "weightSource": {
            "method": method,
            "label": label,
            "note": note,
            "templateCode": source_row["code"],
            "templateName": source_row["name"],
            "inferred": False,
        },
    }


def scaled_weight_payload(method: str, label: str, note: str, source_row: dict[str, Any], ratio: float) -> dict[str, Any]:
    return {
        "weights": scaled_weights(source_row, ratio),
        "weightSource": {
            "method": method,
            "label": label,
            "note": note,
            "templateCode": source_row["code"],
            "templateName": source_row["name"],
            "ratio": round_or_none(ratio),
            "inferred": True,
        },
    }


def surrogate_family_key(family: str) -> str | None:
    if family.startswith("FHLR"):
        return f"FHLAL{family[3:]}"
    if family.startswith("FHLAL") and not family.startswith("FHLALR"):
        return f"FHLALR{family[5:]}"
    return None


def resolve_weights(
    code: str,
    name: str,
    sap: str,
    indexes: dict[str, Any],
) -> dict[str, Any]:
    normalized_code = code_key(code)
    normalized_name = name_key(name)
    section = parse_section(code, name)
    family = family_key(code)

    exact = indexes["exactCode"].get(code)
    if exact:
        return direct_weight_payload("exact_code", "原表", "沿用报价/定点导线模板。", exact)

    normalized = indexes["normalizedCode"].get(normalized_code)
    if normalized:
        return direct_weight_payload("normalized_code", "规格映射", "按去后缀后的同型号模板映射。", normalized)

    if sap and sap in indexes["sap"]:
        return direct_weight_payload("sap_match", "SAP映射", "按相同 SAP 模板映射。", indexes["sap"][sap])

    if normalized_name and normalized_name in indexes["name"]:
        return direct_weight_payload("name_match", "名称映射", "按相同导线名称模板映射。", indexes["name"][normalized_name])

    if family and section is not None:
        family_section = indexes["familySection"].get((family, float(section)))
        if family_section:
            return direct_weight_payload("family_section", "家族映射", "按相同家族与规格模板映射。", family_section)

    if family and section is not None and family in indexes["familyMap"]:
        family_rows = indexes["familyMap"][family]
        source_row = min(
            family_rows,
            key=lambda item: abs(float(item.get("section") or 0.0) - float(section)),
        )
        source_section = float(source_row.get("section") or 0.0)
        ratio = float(section) / source_section if source_section else 1.0
        note = f"TT 专有规格，按 {source_row['code']} 同家族模板 × {ratio:.4f} 推算。"
        return scaled_weight_payload("scaled_family", "同家族推算", note, source_row, ratio)

    surrogate_family = surrogate_family_key(family)
    if surrogate_family and section is not None and surrogate_family in indexes["familyMap"]:
        family_rows = indexes["familyMap"][surrogate_family]
        source_row = min(
            family_rows,
            key=lambda item: abs(float(item.get("section") or 0.0) - float(section)),
        )
        source_section = float(source_row.get("section") or 0.0)
        ratio = float(section) / source_section if source_section else 1.0
        note = f"TT 专有规格，按 {source_row['code']} 同结构近似模板 × {ratio:.4f} 推算。"
        return scaled_weight_payload("scaled_surrogate", "同类推算", note, source_row, ratio)

    return {
        "weights": {
            "aluminum": None,
            "copper": None,
            "nonCopper": None,
        },
        "weightSource": {
            "method": "unresolved",
            "label": "待补重量",
            "note": "当前缺少可映射的重量模板，请补充导线重量基准。",
            "templateCode": None,
            "templateName": None,
            "inferred": True,
        },
    }


def load_master_metal_versions() -> dict[str, dict[str, float | None]]:
    if not MASTER_PATH.exists():
        return {"quote": {"aluminum": None, "copper": None}, "fixed": {"aluminum": None, "copper": None}, "tt": {"aluminum": None, "copper": None}}
    payload = json.loads(MASTER_PATH.read_text(encoding="utf-8"))
    versions = payload.get("versions", {}).get("metal", {})
    return {
        "quote": {
            "aluminum": round_or_none(to_float(versions.get("quote", {}).get("aluminumPrice"))),
            "copper": round_or_none(to_float(versions.get("quote", {}).get("copperPrice"))),
        },
        "fixed": {
            "aluminum": round_or_none(to_float(versions.get("fixed", {}).get("aluminumPrice"))),
            "copper": round_or_none(to_float(versions.get("fixed", {}).get("copperPrice"))),
        },
        "tt": {
            "aluminum": round_or_none(to_float(versions.get("tt", {}).get("aluminumPrice"))),
            "copper": round_or_none(to_float(versions.get("tt", {}).get("copperPrice"))),
        },
    }


def build_catalog(quote_path: Path, fixed_path: Path, tt_path: Path) -> dict[str, Any]:
    quote_snapshot = load_linked_snapshot(quote_path)
    fixed_snapshot = load_linked_snapshot(fixed_path)
    tt_snapshot = load_tt_snapshot(tt_path)
    master_metal_versions = load_master_metal_versions()
    indexes = build_template_indexes(quote_snapshot, fixed_snapshot)

    order = list(quote_snapshot["models"].keys())
    for snapshot in (fixed_snapshot, tt_snapshot):
        for code in snapshot["models"].keys():
            if code not in order:
                order.append(code)

    models = []
    inferred_model_count = 0
    for code in order:
        quote_row = quote_snapshot["models"].get(code)
        fixed_row = fixed_snapshot["models"].get(code)
        tt_row = tt_snapshot["models"].get(code)
        base_row = quote_row or fixed_row or tt_row
        if not base_row:
            continue

        weight_payload = resolve_weights(
            code=code,
            name=to_text(base_row.get("name")),
            sap=to_text(base_row.get("sap")),
            indexes=indexes,
        )
        weights = weight_payload["weights"]
        weight_source = weight_payload["weightSource"]
        if weight_source.get("inferred"):
            inferred_model_count += 1

        aluminum_weight = to_float(weights.get("aluminum")) or 0.0
        copper_weight = to_float(weights.get("copper")) or 0.0
        non_copper_cost = to_float(weights.get("nonCopper")) or 0.0

        quote_unit_price = None
        if quote_snapshot["aluminumBasePrice"] and quote_snapshot["copperBasePrice"]:
            quote_unit_price = calculate_sheet_unit_price(
                aluminum_weight,
                copper_weight,
                non_copper_cost,
                float(quote_snapshot["aluminumBasePrice"]),
                float(quote_snapshot["copperBasePrice"]),
            )

        fixed_unit_price = None
        if fixed_snapshot["aluminumBasePrice"] and fixed_snapshot["copperBasePrice"]:
            fixed_unit_price = calculate_sheet_unit_price(
                aluminum_weight,
                copper_weight,
                non_copper_cost,
                float(fixed_snapshot["aluminumBasePrice"]),
                float(fixed_snapshot["copperBasePrice"]),
            )

        models.append(
            {
                "code": code,
                "name": merge_many(base_row.get("name"), quote_row.get("name") if quote_row else "", fixed_row.get("name") if fixed_row else "", tt_row.get("name") if tt_row else ""),
                "unit": merge_many(base_row.get("unit"), quote_row.get("unit") if quote_row else "", fixed_row.get("unit") if fixed_row else "", tt_row.get("unit") if tt_row else "") or "M",
                "supplier": merge_many(
                    quote_row.get("supplier") if quote_row else "",
                    fixed_row.get("supplier") if fixed_row else "",
                    tt_row.get("supplier") if tt_row else "",
                ),
                "sap": merge_many(
                    quote_row.get("sap") if quote_row else "",
                    fixed_row.get("sap") if fixed_row else "",
                    tt_row.get("sap") if tt_row else "",
                ),
                "remark": merge_many(
                    quote_row.get("remark") if quote_row else "",
                    fixed_row.get("remark") if fixed_row else "",
                    tt_row.get("remark") if tt_row else "",
                ),
                "owner": merge_many(
                    quote_row.get("owner") if quote_row else "",
                    fixed_row.get("owner") if fixed_row else "",
                    tt_row.get("owner") if tt_row else "",
                ),
                "currency": merge_many(
                    quote_row.get("currency") if quote_row else "",
                    fixed_row.get("currency") if fixed_row else "",
                    tt_row.get("currency") if tt_row else "",
                ) or "CNY",
                "priceType": merge_many(
                    quote_row.get("priceType") if quote_row else "",
                    fixed_row.get("priceType") if fixed_row else "",
                    tt_row.get("priceType") if tt_row else "",
                ),
                "familyKey": family_key(code),
                "section": round_or_none(parse_section(code, base_row.get("name")), 4),
                "materialFamily": material_family(aluminum_weight, copper_weight) if any((aluminum_weight, copper_weight)) else material_family_from_name(base_row.get("name", "")),
                "relationLabel": relation_label(aluminum_weight, copper_weight, non_copper_cost),
                "weights": {
                    "aluminum": round_or_none(aluminum_weight),
                    "copper": round_or_none(copper_weight),
                    "nonCopper": round_or_none(non_copper_cost),
                },
                "weightSource": weight_source,
                "usage": {
                    "quote": round_or_none(float(quote_row["usageQty"])) if quote_row else None,
                    "fixed": round_or_none(float(fixed_row["usageQty"])) if fixed_row else None,
                    "tt": round_or_none(float(tt_row["usageQty"])) if tt_row else None,
                },
                "baseUnitPrice": {
                    "quote": round_or_none(quote_unit_price) if quote_unit_price is not None else None,
                    "fixed": round_or_none(fixed_unit_price) if fixed_unit_price is not None else None,
                    "tt": None,
                },
                "sourceRows": {
                    "quote": list(quote_row["sourceRows"]) if quote_row else [],
                    "fixed": list(fixed_row["sourceRows"]) if fixed_row else [],
                    "tt": list(tt_row["sourceRows"]) if tt_row else [],
                },
            }
        )

    total_usage_quote = sum(float(model["usage"]["quote"] or 0) for model in models)
    total_usage_fixed = sum(float(model["usage"]["fixed"] or 0) for model in models)
    total_usage_tt = sum(float(model["usage"]["tt"] or 0) for model in models)
    total_aluminum_weight = sum(float(model["weights"]["aluminum"] or 0) for model in models)
    total_copper_weight = sum(float(model["weights"]["copper"] or 0) for model in models)
    total_non_copper = sum(float(model["weights"]["nonCopper"] or 0) for model in models)

    return {
        "meta": {
            "generator": "g281_generate_wire_catalog.py",
            "generatedAt": datetime.now(TZ).isoformat(timespec="seconds"),
            "quoteWorkbook": quote_snapshot["workbook"],
            "fixedWorkbook": fixed_snapshot["workbook"],
            "ttWorkbook": tt_snapshot["workbook"],
            "quoteSourceWorkbook": quote_snapshot["sourceWorkbook"],
            "fixedSourceWorkbook": fixed_snapshot["sourceWorkbook"],
            "ttSourceWorkbook": tt_snapshot["sourceWorkbook"],
            "sheetUnitPriceRule": "导线基准单价 = 铝重*铝价/1000 + 铜重*铜价/1000 + 非铜/1000（导线价格表按元/kg）",
            "runtimeUnitPriceRule": "导线当前单价 = 铝重*当前铝价/1000000 + 铜重*当前铜价/1000000 + 非铜/1000（页面金属价按元/吨）",
            "usageFallbackOrder": {
                "tt": ["tt"],
                "fixed": ["fixed"],
                "quote": ["quote"],
            },
        },
        "basePrices": {
            "quote": {
                "aluminum": quote_snapshot["aluminumBasePrice"],
                "copper": quote_snapshot["copperBasePrice"],
            },
            "fixed": {
                "aluminum": fixed_snapshot["aluminumBasePrice"],
                "copper": fixed_snapshot["copperBasePrice"],
            },
            "tt": master_metal_versions["tt"],
        },
        "runtimeMetalVersions": master_metal_versions,
        "summary": {
            "modelCount": len(models),
            "usageQtyQuote": round_or_none(total_usage_quote),
            "usageQtyFixed": round_or_none(total_usage_fixed),
            "usageQtyTt": round_or_none(total_usage_tt),
            "totalAluminumWeight": round_or_none(total_aluminum_weight),
            "totalCopperWeight": round_or_none(total_copper_weight),
            "totalNonCopper": round_or_none(total_non_copper),
            "aluminumFamilyCount": sum(1 for model in models if model["materialFamily"] == "铝基"),
            "copperFamilyCount": sum(1 for model in models if model["materialFamily"] == "铜基"),
            "mixedFamilyCount": sum(1 for model in models if model["materialFamily"] == "铜铝混合"),
            "inferredModelCount": inferred_model_count,
        },
        "models": models,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate wire model catalog JSON from quote/fixed linked workbooks and the TT workbook.")
    parser.add_argument("--quote", type=Path, default=None, help="Quote linked workbook path.")
    parser.add_argument("--fixed", type=Path, default=None, help="Fixed linked workbook path.")
    parser.add_argument("--tt", type=Path, default=None, help="TT workbook path.")
    parser.add_argument("--out", type=Path, default=OUTPUT_PATH, help="Output JSON path.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    quote_path = args.quote or discover_linked_workbook("报价核算", "导线联动")
    fixed_path = args.fixed or discover_linked_workbook("定点核算", "导线联动")
    tt_path = args.tt or discover_tt_workbook()
    payload = build_catalog(quote_path, fixed_path, tt_path)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(args.out)
    print(f"Wire models: {payload['summary']['modelCount']}")
    print(f"TT usage total: {payload['summary']['usageQtyTt']}")


if __name__ == "__main__":
    main()
